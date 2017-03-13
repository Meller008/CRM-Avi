import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet import Worksheet
from openpyxl.styles import NamedStyle
from copy import copy

COLUMN_EXCEL = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']


def table_to_excel(table, path):
    wb = openpyxl.Workbook()
    ws = wb.active

    for col in range(table.columnCount()):
        head = table.horizontalHeaderItem(col).text()
        ws["%s%s" % (COLUMN_EXCEL[col], 1)] = str(head)

        for row in range(table.rowCount()):
            for col in range(table.columnCount()):
                head = table.item(row, col).text()
                ws["%s%s" % (COLUMN_EXCEL[col], row+2)] = str(head)

    wb.save(path + ".xlsx")
