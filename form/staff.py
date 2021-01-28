from os import getcwd, path, mkdir, listdir, rmdir
from shutil import copy
from form.templates import list
from PyQt5.uic import loadUi
from PyQt5.QtWidgets import QDialog, QMainWindow, QMessageBox, QTableWidgetItem, QListWidgetItem, QFileDialog, QLineEdit, QWidget, QSizePolicy, QButtonGroup, QInputDialog
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import Qt, QDate
from function import my_sql, to_excel
import openpyxl
import subprocess
from classes.my_class import User


class Staff(QMainWindow):
    def __init__(self, main=None, dc_select=False):
        super(Staff, self).__init__()
        loadUi(getcwd() + '/ui/staff.ui', self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.main = main
        self.dc_select = dc_select
        self.filter = None
        self.query_table_all = """SELECT staff_worker_info.Id, Last_Name, First_Name, DATE_FORMAT(Date_Recruitment, '%d.%m.%Y'), staff_worker_login.Login,
                                      `Leave`, Date_Leave, staff_position.Name, Date_Recruitment, Ip
                                    FROM staff_worker_info LEFT JOIN staff_position ON staff_worker_info.Position_Id = staff_position.Id
                                      LEFT JOIN staff_worker_login ON staff_worker_info.Id = staff_worker_login.Worker_Info_Id
                                    ORDER BY Last_Name"""
        self.query_table_select = self.query_table_all

        self.set_settings()
        self.set_info()

        # Быстрый фильтр
        self.le_fast_filter = QLineEdit()
        self.le_fast_filter.setPlaceholderText("Фамилия")
        self.le_fast_filter.setMaximumWidth(150)
        self.le_fast_filter.editingFinished.connect(self.fast_filter)
        dummy = QWidget()
        dummy.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.Preferred)
        self.toolBar.addWidget(dummy)
        self.toolBar.addWidget(self.le_fast_filter)

        self.access()

    def access(self):
        for item in User().access_list(self.__class__.__name__):
            a = getattr(self, item["atr1"])
            if item["atr2"]:
                a = getattr(a, item["atr2"])

            if item["value"]:
                if item["value"] == "True":
                    val = True
                elif item["value"] == "False":
                    val = False
                else:
                    try:
                        val = int(item["value"])
                    except:
                        val = item["value"]
                a(val)
            else:
                a()

    def inspection_path(self, dir_name, sql_dir_name):  # Находим путь работника
        if not hasattr(self, 'path_work'):
            query = 'SELECT `Values` FROM program_settings_path WHERE Name = "%s"' % sql_dir_name
            info_sql = my_sql.sql_select(query)
            if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False
            self.path_wor = info_sql[0][0]
            if not path.isdir("%s/%s" % (self.path_wor, dir_name)):
                try:
                    mkdir("%s/%s" % (self.path_wor, dir_name))
                    return "%s/%s" % (self.path_wor, dir_name)
                except:
                    QMessageBox.critical(self, "Ошибка файлы", "Нет доступа к корневому диалогу, файлы недоступны", QMessageBox.Ok)
                    return False
            else:
                return "%s/%s" % (self.path_wor, dir_name)

    def set_settings(self):
        self.tw_workers.horizontalHeader().resizeSection(0, 50)
        self.tw_workers.horizontalHeader().resizeSection(1, 150)
        self.tw_workers.horizontalHeader().resizeSection(2, 150)
        self.tw_workers.horizontalHeader().resizeSection(3, 80)
        self.tw_workers.horizontalHeader().resizeSection(4, 60)

    def set_info(self):
        query = "SELECT Id, Name FROM staff_position"
        self.staff_positions = my_sql.sql_select(query)
        if "mysql.connector.errors" in str(type(self.staff_positions)):
            QMessageBox.critical(self, "Ошибка sql", self.staff_positions.msg, QMessageBox.Ok)
            return False
        self.staff_workers = my_sql.sql_select(self.query_table_select)
        if "mysql.connector.errors" in str(type(self.staff_workers)):
            QMessageBox.critical(self, "Ошибка sql", self.staff_workers.msg, QMessageBox.Ok)
            return False

        self.lw_position.clear()
        for position in self.staff_positions:
            self.lw_position.addItem(position[1])
        else:
            self.lw_position.addItem("Принятые в этом месяце ООО")
            self.lw_position.addItem("Принятые в этом месяце ИП")
            self.lw_position.addItem("Уволеные в этом месяце ООО")
            self.lw_position.addItem("Уволеные в этом месяце ИП")
            self.lw_position.addItem("Уволеные в этом году")
            self.lw_position.addItem("Уволеные")
            self.lw_position.addItem("Не принятые")
            self.lw_position.addItem("Все работающие")

        self.tw_workers.setSortingEnabled(False)
        self.tw_workers.clearContents()
        self.tw_workers.setRowCount(len(self.staff_workers))
        for row in range(len(self.staff_workers)):
            for column in range(5):
                a = self.staff_workers[row][column]
                item = QTableWidgetItem(str(a))
                self.tw_workers.setItem(row, column, item)

        self.tw_workers.setSortingEnabled(True)

    def add(self):
        self.add_mat = OneStaff(self)
        self.add_mat.set_add_settings()
        self.add_mat.setWindowModality(Qt.ApplicationModal)
        self.add_mat.show()

    def double_click(self, row):
        if not self.dc_select:
            id = self.tw_workers.item(row, 0).text()
            self.add_mat = OneStaff(self, True)
            self.add_mat.set_add_settings()
            if self.add_mat.insert_info(id):
                self.add_mat.setWindowModality(Qt.ApplicationModal)
                self.add_mat.show()
        else:
            item = (self.tw_workers.item(row, 0).text(), self.tw_workers.item(row, 2).text())
            self.main.of_list_worker(item)
            self.close()
            self.destroy()

    def change(self):
        try:
            row = self.tw_workers.selectedItems()[0].row()
            self.double_click(row)
        except:
            pass

    def delete(self):
        try:
            select_work = self.tw_workers.item(self.tw_workers.selectedItems()[0].row(), 0).text()
            if select_work:
                result = QMessageBox.question(self, "Удаление", "Точно удалить работника?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if result == 16384:
                    query = "DELETE FROM staff_worker_info WHERE Id = %s"
                    info_sql = my_sql.sql_change(query, (select_work, ))
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False
                    dir = select_work
                    rmdir(self.inspection_path(dir, "Путь корень рабочие"))
                    self.set_info()
        except:
            pass

    def sorting(self, select_position):
        self.select_position = select_position  # Запоменаем выбраную должность
        self.to_date = QDate.currentDate()
        self.tw_workers.setSortingEnabled(False)

        if self.select_position == "Принятые в этом месяце ООО":  # Вставляем принятых в этом месяце
            self.tw_workers.clearContents()
            self.tw_workers.setRowCount(0)
            for row in range(len(self.staff_workers)):
                if self.staff_workers[row][5] == 0 and self.staff_workers[row][8].month == self.to_date.month() and\
                        self.staff_workers[row][8].year == self.to_date.year() and self.staff_workers[row][9] == 0:
                    self.tw_workers.insertRow(self.tw_workers.rowCount())
                    for column in range(5):
                        a = self.staff_workers[row][column]
                        item = QTableWidgetItem(str(a))
                        self.tw_workers.setItem(self.tw_workers.rowCount() - 1, column, item)

        elif self.select_position == "Принятые в этом месяце ИП":  # Вставляем принятых в этом месяце
            self.tw_workers.clearContents()
            self.tw_workers.setRowCount(0)
            for row in range(len(self.staff_workers)):
                if self.staff_workers[row][5] == 0 and self.staff_workers[row][8].month == self.to_date.month() and\
                        self.staff_workers[row][8].year == self.to_date.year() and self.staff_workers[row][9] == 1:
                    self.tw_workers.insertRow(self.tw_workers.rowCount())
                    for column in range(5):
                        a = self.staff_workers[row][column]
                        item = QTableWidgetItem(str(a))
                        self.tw_workers.setItem(self.tw_workers.rowCount() - 1, column, item)

        elif self.select_position == "Уволеные":  # Вставляем уволеных
            self.tw_workers.clearContents()
            self.tw_workers.setRowCount(0)
            for row in range(len(self.staff_workers)):
                if self.staff_workers[row][5] == 1:
                    self.tw_workers.insertRow(self.tw_workers.rowCount())
                    for column in range(5):
                        a = self.staff_workers[row][column]
                        item = QTableWidgetItem(str(a))
                        self.tw_workers.setItem(self.tw_workers.rowCount() - 1, column, item)

        elif self.select_position == "Уволеные в этом году":  # Вставляем уволеных
            self.tw_workers.clearContents()
            self.tw_workers.setRowCount(0)
            for row in range(len(self.staff_workers)):
                if self.staff_workers[row][5] == 1 and self.staff_workers[row][6].year == self.to_date.year():
                    self.tw_workers.insertRow(self.tw_workers.rowCount())
                    for column in range(5):
                        a = self.staff_workers[row][column]
                        item = QTableWidgetItem(str(a))
                        self.tw_workers.setItem(self.tw_workers.rowCount() - 1, column, item)

        elif self.select_position == "Уволеные в этом месяце ООО":  # Вставляем уволеных
            self.tw_workers.clearContents()
            self.tw_workers.setRowCount(0)
            for row in range(len(self.staff_workers)):
                if self.staff_workers[row][5] == 1 and self.staff_workers[row][6].month == self.to_date.month() and\
                        self.staff_workers[row][6].year == self.to_date.year() and self.staff_workers[row][9] == 0:
                    self.tw_workers.insertRow(self.tw_workers.rowCount())
                    for column in range(5):
                        a = self.staff_workers[row][column]
                        item = QTableWidgetItem(str(a))
                        self.tw_workers.setItem(self.tw_workers.rowCount() - 1, column, item)

        elif self.select_position == "Уволеные в этом месяце ИП":  # Вставляем уволеных
            self.tw_workers.clearContents()
            self.tw_workers.setRowCount(0)
            for row in range(len(self.staff_workers)):
                if self.staff_workers[row][5] == 1 and self.staff_workers[row][6].month == self.to_date.month() and\
                        self.staff_workers[row][6].year == self.to_date.year() and self.staff_workers[row][9] == 1:
                    self.tw_workers.insertRow(self.tw_workers.rowCount())
                    for column in range(5):
                        a = self.staff_workers[row][column]
                        item = QTableWidgetItem(str(a))
                        self.tw_workers.setItem(self.tw_workers.rowCount() - 1, column, item)

        elif self.select_position == "Не принятые":  # Вставляем уволеных
            self.tw_workers.clearContents()
            self.tw_workers.setRowCount(0)
            for row in range(len(self.staff_workers)):
                if self.staff_workers[row][5] == 2:
                    self.tw_workers.insertRow(self.tw_workers.rowCount())
                    for column in range(5):
                        a = self.staff_workers[row][column]
                        item = QTableWidgetItem(str(a))
                        self.tw_workers.setItem(self.tw_workers.rowCount() - 1, column, item)

        elif self.select_position == "Все работающие":  # Вставляем уволеных
            self.tw_workers.clearContents()
            self.tw_workers.setRowCount(0)
            for row in range(len(self.staff_workers)):
                if self.staff_workers[row][5] == 0:
                    self.tw_workers.insertRow(self.tw_workers.rowCount())
                    for column in range(5):
                        a = self.staff_workers[row][column]
                        item = QTableWidgetItem(str(a))
                        self.tw_workers.setItem(self.tw_workers.rowCount() - 1, column, item)

        else:
            self.tw_workers.clearContents()
            self.tw_workers.setRowCount(0)
            for row in range(len(self.staff_workers)):
                if self.staff_workers[row][7] == self.select_position and self.staff_workers[row][5] == 0:
                    self.tw_workers.insertRow(self.tw_workers.rowCount())
                    for column in range(5):
                        a = self.staff_workers[row][column]
                        item = QTableWidgetItem(str(a))
                        self.tw_workers.setItem(self.tw_workers.rowCount() - 1, column, item)

        self.tw_workers.setSortingEnabled(True)

    def ui_filter(self):
        if self.filter is None:
            self.filter = StaffFilter(self)
        self.filter.of_set_sql_query(self.query_table_all)
        self.filter.setWindowModality(Qt.ApplicationModal)
        self.filter.show()

    def fast_filter(self):
        # Блок условий артикула
        if self.le_fast_filter.text() != '':
            q_filter = "(staff_worker_info.Last_Name LIKE '%s')" % ("%" + self.le_fast_filter.text() + "%", )
            self.query_table_select = self.query_table_all.replace("ORDER BY", " WHERE " + q_filter + " ORDER BY")
        else:
            self.query_table_select = self.query_table_all

        self.set_info()

    def ui_export(self):
        path = QFileDialog.getSaveFileName(self, "Сохранение")
        if path[0]:
            to_excel.table_to_excel(self.tw_workers, path[0])

    def of_set_filter(self, sql):
        self.query_table_select = sql

        self.set_info()


class OneStaff(QMainWindow):
    def __init__(self, main, change=False):
        super(OneStaff, self).__init__()
        loadUi(getcwd() + '/ui/add_work.ui', self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.change = change  # Запоминаем это добаление работника или изменение
        self.m = main
        self.access_save_sql = True
        self.alert = []  # Массив для запоминания изменений

        self.id_info = None
        self.sql_beika = False

        self.access()

    def access(self):
        for item in User().access_list(self.__class__.__name__):
            a = getattr(self, item["atr1"])
            if item["atr2"]:
                a = getattr(a, item["atr2"])

            if item["value"]:
                if item["value"] == "True":
                    val = True
                elif item["value"] == "False":
                    val = False
                else:
                    val = item["value"]
                a(val)
            else:
                a()

    def access_save(self, bool):
        self.access_save_sql = bool

    def inspection_path(self, dir_name, sql_dir_name):  # Находим путь работника
        if not hasattr(self, 'path_work'):
            query = 'SELECT `Values` FROM program_settings_path WHERE Name = "%s"' % sql_dir_name
            info_sql = my_sql.sql_select(query)
            if "mysql.connector.errors" in str(type(info_sql)):
                QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                return False
            self.path_wor = info_sql[0][0]
            if not path.isdir("%s/%s" % (self.path_wor, dir_name)):
                try:
                    mkdir("%s/%s" % (self.path_wor, dir_name))
                    return "%s/%s" % (self.path_wor, dir_name)
                except:
                    QMessageBox.critical(self, "Ошибка файлы", "Нет доступа к корневому диалогу, файлы недоступны", QMessageBox.Ok)
                    return False
            else:
                return "%s/%s" % (self.path_wor, dir_name)

    def inspection_files(self, dir_name, sql_dir_name):   # Проверяем файлы и даем иконки
        self.path = self.inspection_path(dir_name, sql_dir_name)
        if self.path:
            self.lw_file.clear()
            files = listdir("%s/%s" % (self.path_wor, dir_name))
            for file in files:
                if "~" not in file:
                    r = path.splitext(file)  # Получаем название и расширение
                    if "xlsx" in r[1][1:] or "xlsm" in r[1] or "xls" in r[1] or "xlt" in r[1]:
                        ico = "xlsx"
                    elif "xml" in r[1][1:] or "docx" in r[1] or "doc" in r[1] or "docm" in r[1]:
                        ico = "xml"
                    elif "png" in r[1].lower() or "jpg" in r[1] or "jpeg" in r[1] or "jpe" in r[1] or "gif" in r[1] or "bmp" in r[1]:
                        ico = "image"
                    elif "pdf" in r[1]:
                        ico = "pdf"
                    else:
                        ico = "other"

                    list_item = QListWidgetItem(r[0] + r[1])
                    list_item.setIcon(QIcon(getcwd() + "/images/%s.ico" % ico))
                    self.lw_file.addItem(list_item)

    def set_add_settings(self):
        # Создадим группы для радио кнопок
        self.group_sex = QButtonGroup()
        self.group_leave = QButtonGroup()

        self.group_sex.addButton(self.rb_sex_m)
        self.group_sex.addButton(self.rb_sex_f)

        self.group_leave.addButton(self.rb_not_employed)
        self.group_leave.addButton(self.rb_employed)
        self.group_leave.addButton(self.rb_leave)

        # Начальные чеки
        self.de_info_leave.setEnabled(False)
        self.rb_sex_f.setChecked(True)
        self.rb_not_employed.setChecked(True)

        # Выставляем даты
        self.to_date = QDate.currentDate()
        self.de_info_birth.setDate(self.to_date)
        self.de_info_recruitment.setDate(self.to_date)
        self.de_info_leave.setDate(self.to_date)
        self.de_passport_issued.setDate(self.to_date)
        self.de_passport_ending.setDate(self.to_date)
        self.de_migration_validity_from.setDate(self.to_date)
        self.de_migration_validity_to.setDate(self.to_date)
        self.de_registration_validity_from.setDate(self.to_date)
        self.de_registration_validity_to.setDate(self.to_date)
        self.de_registration.setDate(self.to_date)
        self.de_patent_issued.setDate(self.to_date)
        self.de_patent_ending.setDate(self.to_date)
        self.de_migration.setDate(self.to_date)
        self.de_insurance_date.setDate(self.to_date)
        self.de_notification.setDate(self.to_date)

        # заполняем страны
        query = "SELECT Country_name FROM staff_country"
        self.country = my_sql.sql_select(query)
        if "mysql.connector.errors" in str(type(self.country)):
            QMessageBox.critical(self, "Ошибка sql", self.country.msg, QMessageBox.Ok)
            return False
        for country in self.country:
            self.cb_info_country.addItem(country[0])

        # заполняем должности
        query = "SELECT Name FROM staff_position"
        self.position = my_sql.sql_select(query)
        if "mysql.connector.errors" in str(type(self.position)):
            QMessageBox.critical(self, "Ошибка sql", self.position.msg, QMessageBox.Ok)
            return False
        for country in self.position:
            self.cb_info_position.addItem(country[0])

        self.alert = []  # Обнуляем массив для запоминания изменений

    def change_birth(self, birth_date):  # Подсчет колличества лет
        years = int(birth_date.daysTo(self.to_date) / 365)
        self.lb_info_years.setText("Возраст: %s" % years)
        self.alter_info()

    def change_patent_date(self, patent_date):  # автоматически продлевает патент на 1 год
        self.de_patent_ending.setDate(patent_date.addYears(1))

    def select_file(self, file):  # Открываем выбраный фаил
        if self.id_info:
            dir_name = self.id_info
            self.path = self.inspection_path(dir_name, 'Путь корень рабочие')
            if self.path:
                file_name = file.text()
                subprocess.Popen(r'%s/%s' % (self.path.replace("/", "\\"), file_name.replace("/", "\\")), shell=True)

    def open_dir(self):  # Открываем выбраную папку
        if self.id_info:
            dir_name = self.id_info
            self.path = self.inspection_path(dir_name, 'Путь корень рабочие')
            if self.path:
                # subprocess.Popen(['explorer "' + self.path.replace("/", "\\") + '"'])
                subprocess.Popen('explorer "%s"' % self.path.replace("/", "\\"))

    def add_file(self):  # Добавляем файлы
        info = AddFile()
        if info.exec() == 0:
            return False
        new_r = path.splitext(info.path_copy_file.text())[1]
        dir_name = self.id_info
        copy(info.path_copy_file.text(), self.inspection_path(dir_name, 'Путь корень рабочие') + "/" + info.le_new_file_name.text() + path.splitext(info.path_copy_file.text())[1])
        self.inspection_files(dir_name, 'Путь корень рабочие')

    def input_check(self):
        self.alert2 = []  # Запоминаем введеные даные
        self.delete = []  # Запоминаем что надо удалить
        # проверяем основную информацию
        if self.le_info_first_name.text() == "":
            QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели имя", QMessageBox.Ok)
            return False
        if self.le_info_last_name.text() == "":
            QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели фамилию", QMessageBox.Ok)
            return False
        if self.de_info_birth.date() > self.to_date:
            QMessageBox.critical(self, "Ошибка ввода", "Не верная дата рождения", QMessageBox.Ok)
            return False
        if self.rb_leave.isChecked() and self.de_info_leave.date() < self.de_info_recruitment.date():
            QMessageBox.critical(self, "Ошибка ввода", "Не верная дата увольнения", QMessageBox.Ok)
            return False
        if self.le_info_birthplace.text() == "":
            QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели местро рождения", QMessageBox.Ok)
            return False

        # проверка паспорта
        if self.le_passport_number.text() != "" or self.le_passport_issued.text() != "":
            self.alert2.append("passport")
            if self.le_passport_number.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели номер паспорта", QMessageBox.Ok)
                return False
            if self.le_passport_issued.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели кем выдан пасспорт", QMessageBox.Ok)
                return False
            if self.de_passport_ending.date() < self.de_passport_issued.date():
                QMessageBox.critical(self, "Ошибка ввода", "Не верная дата окончания паспорта", QMessageBox.Ok)
                return False
        elif self.le_passport_series.text() == "" and self.le_passport_number.text() == "" and self.le_passport_issued.text() == "":
            self.delete.append("passport")
            self.alert2.append("passport")

        # проверка миграционной карты
        if self.le_migration_serial.text() != "" or self.le_migration_number.text() != "":
            self.alert2.append("migration")
            if self.le_migration_serial.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели серию миграционной карты", QMessageBox.Ok)
                return False
            if self.le_migration_number.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели номер миграционной карты", QMessageBox.Ok)
                return False
            if self.le_migration_kpp.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели КПП выдачи минрационной карты", QMessageBox.Ok)
                return False
            if self.de_migration_validity_to.date() < self.de_migration_validity_from.date():
                QMessageBox.critical(self, "Ошибка ввода", "Не верная дата окончания миграционной карты", QMessageBox.Ok)
                return False
        elif self.le_migration_serial.text() == "" and self.le_migration_number.text() == "" and self.le_migration_kpp.text() == "":
            self.delete.append("migration")
            self.alert2.append("migration")

        # проверка регистрации
        if self.le_registration_address.text() != "":
            self.alert2.append("registration")
            if self.de_registration_validity_to.date() < self.de_registration_validity_from.date():
                QMessageBox.critical(self, "Ошибка ввода", "Не верная дата окончания регистрации", QMessageBox.Ok)
                return False
        elif self.le_registration_address.text() == "":
            self.delete.append("registration")
            self.alert2.append("registration")

        # проверка патента
        if self.le_patent_serial.text() != "" or self.le_patent_number.text() != "" or self.le_patent_additional_number.text() != "":
            self.alert2.append("patent")
            if self.le_patent_serial.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели серию патента", QMessageBox.Ok)
                return False
            if self.le_patent_number.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели номер патента", QMessageBox.Ok)
                return False
            if self.le_patent_additional_number.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели дополнительный номер патета", QMessageBox.Ok)
                return False
            if self.le_patent_issued.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели кем выдан патент", QMessageBox.Ok)
                return False
            if self.de_patent_ending.date() < self.de_patent_issued.date():
                QMessageBox.critical(self, "Ошибка ввода", "Не верная дата окончания патента", QMessageBox.Ok)
                return False
        elif self.le_patent_serial.text() == "" and self.le_patent_number.text() == "" and self.le_patent_additional_number.text() == "" and self.le_patent_issued.text() == "":
            self.delete.append("patent")
            self.alert2.append("patent")

        # проверка страховки
        if self.le_insurance_number.text() != "" or self.le_insurance_company.text() != "" :
            self.alert2.append("insurance")
            if self.le_insurance_number.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели номер страховой", QMessageBox.Ok)
                return False
            if self.le_insurance_company.text() == "":
                QMessageBox.critical(self, "Ошибка ввода", "Вы не ввели компанию страховой", QMessageBox.Ok)
                return False
        elif self.le_insurance_number.text() == "" and self.le_insurance_company.text() == "":
            self.delete.append("insurance")
            self.alert2.append("insurance")

        if self.le_login_login.text() == "" and self.le_login_password.text() == "":
            self.delete.append("login")

        return True

    def alter_info(self):  # Инфрмация об изменении
        if "info" not in self.alert:
            self.alert.append("info")

    def alter_passport(self):  # Инфрмация об изменении
        if "passport" not in self.alert:
            self.alert.append("passport")

    def alter_migration(self):  # Инфрмация об изменении
        if "migration" not in self.alert:
            self.alert.append("migration")

    def alter_registration(self):  # Инфрмация об изменении
        if "registration" not in self.alert:
            self.alert.append("registration")

    def alter_patent(self):  # Инфрмация об изменении
        if "patent" not in self.alert:
            self.alert.append("patent")

    def alter_insurance(self):  # Инфрмация об изменении
        if "insurance" not in self.alert:
            self.alert.append("insurance")

    def alter_notification(self):  # Инфрмация об изменении
        if "notification" not in self.alert:
            self.alert.append("notification")

    def alter_login(self):  # Инфрмация об изменении
        if "login" not in self.alert:
            self.alert.append("login")

    def alert_beika(self):  # Инфрмация об изменении разрешения на ввод бейки
        if "beika" not in self.alert:
            self.alert.append("beika")

    def insert_info(self, id_worker):
        # Обнуляем список заполненых полей
        self.update_sql = []
        # Вставляем основную информацию
        self.id_info = id_worker
        query = """SELECT staff_worker_info.Id, First_Name, Last_Name, Middle_Name, Sex, Date_Birth, Date_Recruitment, staff_worker_info.Leave,
                    Date_Leave, staff_country.Country_name, Phone, Address, staff_position.Name, INN, SNILS, Note, Birthplace, Ip
                    FROM staff_worker_info LEFT JOIN staff_country ON staff_worker_info.Country_Id = staff_country.Id
                    LEFT JOIN staff_position ON staff_worker_info.Position_Id = staff_position.Id  WHERE staff_worker_info.Id = %s"""
        sql_reply = my_sql.sql_select(query, (id_worker,))
        if "mysql.connector.errors" in str(type(sql_reply)):
            QMessageBox.critical(self, "Ошибка sql", sql_reply.msg, QMessageBox.Ok)
            return False
        self.lb_info_id.setText("Табельный номер: %s" % sql_reply[0][0])
        self.le_info_first_name.setText(sql_reply[0][1])
        self.le_info_last_name.setText(sql_reply[0][2])
        self.le_info_middle_name.setText(sql_reply[0][3])
        if sql_reply[0][4] == "M":
            self.rb_sex_m.setChecked(True)
        elif sql_reply[0][4] == "F":
            self.rb_sex_f.setChecked(True)
        self.le_info_birthplace.setText(sql_reply[0][16])
        self.de_info_birth.setDate(sql_reply[0][5])
        self.de_info_recruitment.setDate(sql_reply[0][6])
        if sql_reply[0][7] == 0:
            self.rb_employed.setChecked(True)
        elif sql_reply[0][7] == 2:
            self.rb_not_employed.setChecked(True)
        elif sql_reply[0][7] == 1:
            self.rb_leave.setChecked(True)
            self.de_info_leave.setEnabled(True)

        if sql_reply[0][8]:
            self.de_info_leave.setDate(sql_reply[0][8])
        self.cb_info_country.setCurrentText(sql_reply[0][9])
        self.le_info_phone.setText(sql_reply[0][10])
        self.le_info_address.setText(sql_reply[0][11])
        self.cb_info_position.setCurrentText(sql_reply[0][12])
        self.le_info_inn.setText(sql_reply[0][13])
        self.le_info_snils.setText(sql_reply[0][14])
        self.le_info_note.appendPlainText(sql_reply[0][15])

        if sql_reply[0][17]:
            self.cb_ip.setChecked(True)

        leave = sql_reply[0][7]

        #Запомним имя фамилию дату для проверки изменения дирректории
        self.sql_f_name = sql_reply[0][1]
        self.sql_l_name = sql_reply[0][2]
        self.sql_r_date = sql_reply[0][6]

        # Заполняем паспорт
        query = "SELECT Series, Number, Issued, Data_Issued, Date_Ending FROM staff_worker_passport WHERE Worker_Info_Id = %s"
        sql_reply = my_sql.sql_select(query, (id_worker,))
        if "mysql.connector.errors" in str(type(sql_reply)):
            QMessageBox.critical(self, "Ошибка sql", sql_reply.msg, QMessageBox.Ok)
            return False
        if sql_reply:
            self.update_sql.append("passport")
            self.le_passport_series.setText(sql_reply[0][0])
            self.le_passport_number.setText(sql_reply[0][1])
            self.le_passport_issued.setText(sql_reply[0][2])
            self.de_passport_issued.setDate(sql_reply[0][3])
            self.de_passport_ending.setDate(sql_reply[0][4])

        # Заполняем миграционку
        query = "SELECT Serial, Number, KPP, Date_Validity_From, Date_Validity_To, Date_migration FROM staff_worker_migration WHERE Worker_Info_Id = %s"
        sql_reply = my_sql.sql_select(query, (id_worker,))
        if "mysql.connector.errors" in str(type(sql_reply)):
            QMessageBox.critical(self, "Ошибка sql", sql_reply.msg, QMessageBox.Ok)
            return False
        if sql_reply:
            self.update_sql.append("migration")
            self.le_migration_serial.setText(sql_reply[0][0])
            self.le_migration_number.setText(sql_reply[0][1])
            self.le_migration_kpp.setText(sql_reply[0][2])
            self.de_migration_validity_from.setDate(sql_reply[0][3])
            self.de_migration_validity_to.setDate(sql_reply[0][4])
            self.de_migration.setDate(sql_reply[0][5])

        # Заполняем регистрацию
        query = "SELECT Address, Date_Registration, Date_Validity_From, Date_Validity_To FROM staff_worker_registraton WHERE Worker_Info_id = %s"
        sql_reply = my_sql.sql_select(query, (id_worker,))
        if "mysql.connector.errors" in str(type(sql_reply)):
            QMessageBox.critical(self, "Ошибка sql", sql_reply.msg, QMessageBox.Ok)
            return False
        if sql_reply:
            self.update_sql.append("registration")
            self.le_registration_address.setText(sql_reply[0][0])
            self.de_registration.setDate(sql_reply[0][1])
            self.de_registration_validity_from.setDate(sql_reply[0][2])
            self.de_registration_validity_to.setDate(sql_reply[0][3])

        # Заполняем патент
        query = "SELECT Serial, Number, Additional_Number, Issued, Data_Issued, Date_Ending FROM staff_worker_patent WHERE Worker_Info_Id = %s"
        sql_reply = my_sql.sql_select(query, (id_worker,))
        if "mysql.connector.errors" in str(type(sql_reply)):
            QMessageBox.critical(self, "Ошибка sql", sql_reply.msg, QMessageBox.Ok)
            return False
        if sql_reply:
            self.update_sql.append("patent")
            self.le_patent_serial.setText(sql_reply[0][0])
            self.le_patent_number.setText(sql_reply[0][1])
            self.le_patent_additional_number.setText(sql_reply[0][2])
            self.le_patent_issued.setText(sql_reply[0][3])
            self.de_patent_issued.setDate(sql_reply[0][4])
            self.de_patent_ending.setDate(sql_reply[0][5])

        # Заполняем страховку
        query = "SELECT Number, Company, Date FROM staff_worker_insurance WHERE Worker_Info_Id = %s"
        sql_reply = my_sql.sql_select(query, (id_worker,))
        if "mysql.connector.errors" in str(type(sql_reply)):
            QMessageBox.critical(self, "Ошибка sql", sql_reply.msg, QMessageBox.Ok)
            return False
        if sql_reply:
            self.update_sql.append("insurance")
            self.le_insurance_number.setText(sql_reply[0][0])
            self.le_insurance_company.setText(sql_reply[0][1])
            self.de_insurance_date.setDate(sql_reply[0][2])

        # Заполняем Оповещения
        query = "SELECT Date, Note FROM staff_worker_notification WHERE Worker_Info_Id = %s"
        sql_reply = my_sql.sql_select(query, (id_worker,))
        if "mysql.connector.errors" in str(type(sql_reply)):
            QMessageBox.critical(self, "Ошибка sql", sql_reply.msg, QMessageBox.Ok)
            return False
        if sql_reply:
            self.update_sql.append("notification")
            self.cb_notification.setChecked(True)
            self.le_notofication.setEnabled(True)
            self.de_notification.setDate(sql_reply[0][0])
            self.le_notofication.setPlainText(sql_reply[0][1])

        # Заполняем список файлов
        dir_name = self.id_info
        self.inspection_files(dir_name, 'Путь корень рабочие')

        self.pushButton.setEnabled(True)
        self.pushButton_4.setEnabled(True)
        self.pushButton_5.setEnabled(True)
        self.pushButton_10.setEnabled(True)
        self.pushButton_6.setEnabled(True)
        self.pushButton_12.setEnabled(True)
        self.pushButton_9.setEnabled(True)
        self.pushButton_13.setEnabled(True)
        self.pushButton_7.setEnabled(True)
        self.pushButton_14.setEnabled(True)
        self.pushButton_8.setEnabled(True)
        self.pushButton_15.setEnabled(True)
        self.pushButton_18.setEnabled(True)
        self.pushButton_19.setEnabled(True)
        self.le_login_login.setEnabled(True)
        self.le_login_password.setEnabled(True)

        # Заполняем логин
        if leave == 0:
            query = "SELECT Login, Password FROM staff_worker_login WHERE Worker_Info_Id = %s"
            sql_reply = my_sql.sql_select(query, (id_worker,))
            if "mysql.connector.errors" in str(type(sql_reply)):
                QMessageBox.critical(self, "Ошибка sql", sql_reply.msg, QMessageBox.Ok)
                return False
            if sql_reply:
                self.update_sql.append("login")
                self.le_login_login.setText(sql_reply[0][0])
                self.le_login_password.setText(sql_reply[0][1])

        self.alert = []

        # Смотрим можно ли вносить бекйку этому работнику
        self.cb_beika.setEnabled(True)
        query = "SELECT Atr_Value FROM access WHERE Worker_Id = %s AND Class = 'MainWindowOperation' AND Atr1 = 'pb_beika'"
        sql_reply = my_sql.sql_select(query, (id_worker,))
        if "mysql.connector.errors" in str(type(sql_reply)):
            QMessageBox.critical(self, "Ошибка sql получения разрешения на бейку", sql_reply.msg, QMessageBox.Ok)
            return False

        if sql_reply and sql_reply[0][0]:
            self.cb_beika.setChecked(True)

            self.sql_beika = True

        return True

    def check_login(self):
        query = "SELECT COUNT(*) FROM staff_worker_login WHERE Login = %s"
        info_sql = my_sql.sql_select(query, (self.le_login_login.text(), ))
        if "mysql.connector.errors" in str(type(info_sql)):
            QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
            return False
        if info_sql[0][0] > 0:
            QMessageBox.critical(self, "Занято", "Этот логин занят", QMessageBox.Ok)
        else:
            QMessageBox.information(self, "Свободно", "Этот логин свободен", QMessageBox.Ok)

    def free_login(self):
        query = """SELECT (staff_worker_login.Login+1) as `login`
                    FROM staff_worker_login
                    WHERE ( SELECT 1 FROM staff_worker_login as `st` WHERE `st`.Login = (staff_worker_login.Login + 1) ) IS NULL
                    ORDER BY staff_worker_login.Login LIMIT 1"""
        info_sql = my_sql.sql_select(query)
        if "mysql.connector.errors" in str(type(info_sql)):
            QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
            return False

        text = "Логин: %s свободен" % str(int(info_sql[0][0]))
        QMessageBox.information(self, "Логин", text, QMessageBox.Ok)

    def acc(self):  # Добаление информации в базу
        if self.change:  # Если мы изменяем а не добавляем работника
            if self.input_check():  # Проверка заполнености полей
                if "info" in self.alert:  # Добаление основной информации

                    if self.rb_sex_m.isChecked():  # Узнаем пол работника
                        self.sex = "M"
                    elif self.rb_sex_f.isChecked():
                        self.sex = "F"

                    if self.rb_employed.isChecked():  # Узнаем принят уволен непринят
                        self.leave = 0
                    elif self.rb_leave.isChecked():
                        self.leave = 1
                    elif self.rb_not_employed.isChecked():
                        self.leave = 2

                    if self.cb_ip.isChecked():
                        ip = 1
                    else:
                        ip = 0

                    id_country = my_sql.sql_select("SELECT Id FROM staff_country WHERE Country_name = %s", (self.cb_info_country.currentText(),))[0][0]
                    id_position = my_sql.sql_select("SELECT Id FROM staff_position WHERE Name = %s", (self.cb_info_position.currentText(),))[0][0]

                    query = """UPDATE staff_worker_info SET First_Name = %s, Last_Name = %s, Middle_Name = %s, Sex = %s, Date_Birth = %s, Birthplace = %s, Date_Recruitment = %s,
                        `Leave` = %s,  Date_Leave = %s, Country_Id = %s, Phone = %s, Address = %s, Position_Id = %s, INN = %s, SNILS = %s, Note = %s, Ip = %s
                         WHERE Id = %s"""
                    parametrs = (self.le_info_first_name.text(), self.le_info_last_name.text(), self.le_info_middle_name.text(), self.sex,
                                 self.de_info_birth.date().toString(Qt.ISODate), self.le_info_birthplace.text(),
                                 self.de_info_recruitment.date().toString(Qt.ISODate),
                                 self.leave, self.de_info_leave.date().toString(Qt.ISODate), id_country, self.le_info_phone.text(),
                                 self.le_info_address.text(), id_position, self.le_info_inn.text(), self.le_info_snils.text(), self.le_info_note.toPlainText(),
                                 ip, self.id_info)
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql i", info_sql.msg, QMessageBox.Ok)
                        return False

                if "passport" in self.alert and "passport" in self.alert2:
                    if "passport" in self.delete:  # Если пспорт надо удалить
                        query = "DELETE FROM staff_worker_passport WHERE Worker_Info_Id = %s"
                        parametrs = (self.id_info, )
                    elif "passport" in self.update_sql:  # Если пспорт надо обновить
                        query = "UPDATE staff_worker_passport SET Series = %s, Number = %s, Issued = %s, Data_Issued = %s, Date_Ending = %s WHERE Worker_Info_Id = %s"
                        parametrs = (self.le_passport_series.text(), self.le_passport_number.text(), self.le_passport_issued.text(),
                                     self.de_passport_issued.date().toString(Qt.ISODate), self.de_passport_ending.date().toString(Qt.ISODate), self.id_info)
                    else:  # Если пспорт надо добавить
                        query = "INSERT INTO staff_worker_passport (Worker_Info_Id, Series, Number, Issued, Data_Issued, Date_Ending) VALUES (%s, %s, %s, %s, %s, %s)"
                        parametrs = (self.id_info, self.le_passport_series.text(), self.le_passport_number.text(), self.le_passport_issued.text(),
                                     self.de_passport_issued.date().toString(Qt.ISODate), self.de_passport_ending.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql p", info_sql.msg, QMessageBox.Ok)
                        return False

                if "migration" in self.alert and "migration" in self.alert2:
                    if "migration" in self.delete:  # Если пспорт надо удалить
                        query = "DELETE FROM staff_worker_migration WHERE Worker_Info_Id = %s"
                        parametrs = (self.id_info, )
                    elif "migration" in self.update_sql:  # Если Миграцию надо обновить
                        query = """UPDATE staff_worker_migration SET Serial = %s, Number = %s, KPP = %s, Date_migration = %s, Date_Validity_From = %s,
                                 Date_Validity_To = %s WHERE Worker_Info_Id = %s"""
                        parametrs = (self.le_migration_serial.text(), self.le_migration_number.text(), self.le_migration_kpp.text(),
                                     self.de_migration.date().toString(Qt.ISODate),
                                     self.de_migration_validity_from.date().toString(Qt.ISODate), self.de_migration_validity_to.date().toString(Qt.ISODate),
                                     self.id_info)
                    else:  # Если миграцию надо добавить
                        query = "INSERT INTO staff_worker_migration (Worker_Info_Id, Serial, Number, KPP, Date_migration, Date_Validity_From, Date_Validity_To) VALUES (%s, %s, %s, %s, %s, %s, %s)"
                        parametrs = (self.id_info, self.le_migration_serial.text(), self.le_migration_number.text(), self.le_migration_kpp.text(),
                                     self.de_migration.date().toString(Qt.ISODate), self.de_migration_validity_from.date().toString(Qt.ISODate),
                                     self.de_migration_validity_to.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql m", info_sql.msg, QMessageBox.Ok)
                        return False

                if "registration" in self.alert and "registration" in self.alert2:
                    if "registration" in self.delete:  # Если пспорт надо удалить
                        query = "DELETE FROM staff_worker_registraton WHERE Worker_Info_Id = %s"
                        parametrs = (self.id_info, )
                    elif "registration" in self.update_sql:  # Если регистрацию надо обновить
                        query = "UPDATE staff_worker_registraton SET Address = %s, Date_Registration = %s, Date_Validity_From = %s, Date_Validity_To = %s WHERE Worker_Info_id = %s"
                        parametrs = (self.le_registration_address.text(), self.de_registration.date().toString(Qt.ISODate),
                                     self.de_registration_validity_from.date().toString(Qt.ISODate), self.de_registration_validity_to.date().toString(Qt.ISODate),
                                     self.id_info)
                    else:  # Если регистрацию надо добавить
                        query = "INSERT INTO staff_worker_registraton (Worker_Info_id, Address, Date_Registration, Date_Validity_From, Date_Validity_To) VALUES (%s, %s, %s, %s, %s)"
                        parametrs = (self.id_info, self.le_registration_address.text(), self.de_registration.date().toString(Qt.ISODate),
                                     self.de_registration_validity_from.date().toString(Qt.ISODate), self.de_registration_validity_to.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql r", info_sql.msg, QMessageBox.Ok)
                        return False

                if "patent" in self.alert and "patent" in self.alert2:
                    if "patent" in self.delete:  # Если пспорт надо удалить
                        query = "DELETE FROM staff_worker_patent WHERE Worker_Info_Id = %s"
                        parametrs = (self.id_info, )
                    elif "patent" in self.update_sql:  # Если патент надо обновить
                        query = """UPDATE staff_worker_patent SET Serial = %s, Number = %s, Additional_Number = %s, Issued = %s, Data_Issued = %s, Date_Ending = %s
                                   WHERE Worker_Info_Id = %s"""
                        parametrs = (self.le_patent_serial.text(), self.le_patent_number.text(), self.le_patent_additional_number.text(),
                                     self.le_patent_issued.text(), self.de_patent_issued.date().toString(Qt.ISODate),
                                     self.de_patent_ending.date().toString(Qt.ISODate), self.id_info)
                    else:  # Если патент надо добавить
                        query = "INSERT INTO staff_worker_patent (Worker_Info_Id, Serial, Number, Additional_Number, Issued,  Data_Issued, Date_Ending) VALUES (%s, %s, %s, %s, %s, %s, %s)"
                        parametrs = (self.id_info, self.le_patent_serial.text(), self.le_patent_number.text(), self.le_patent_additional_number.text(),
                                     self.le_patent_issued.text(), self.de_patent_issued.date().toString(Qt.ISODate),
                                     self.de_patent_ending.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql pt", info_sql.msg, QMessageBox.Ok)
                        return False

                if "insurance" in self.alert and "insurance" in self.alert2:
                    if "insurance" in self.delete:  # Если пспорт надо удалить
                        query = "DELETE FROM staff_worker_insurance WHERE Worker_Info_Id = %s"
                        parametrs = (self.id_info, )
                    elif "insurance" in self.update_sql:  # Если страховку надо обновить
                        query = """UPDATE staff_worker_insurance SET Number = %s, Company = %s, Date = %s WHERE Worker_Info_Id = %s"""
                        parametrs = (self.le_insurance_number.text(), self.le_insurance_company.text(),
                                     self.de_insurance_date.date().toString(Qt.ISODate), self.id_info)
                    else:  # Если страховку надо добавить
                        query = "INSERT INTO staff_worker_insurance (Worker_Info_Id, Number, Company, Date) VALUES (%s, %s, %s, %s)"
                        parametrs = (self.id_info, self.le_insurance_number.text(), self.le_insurance_company.text(), self.de_insurance_date.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql is", info_sql.msg, QMessageBox.Ok)
                        return False

                if "notification" in self.alert:
                    if "notification" in self.update_sql:  # Если напоминания надо обновить
                        if self.cb_notification.isChecked():
                            query = """UPDATE staff_worker_notification SET Date = %s, Note = %s WHERE Worker_Info_Id = %s"""
                            parametrs = (self.de_notification.date().toString(Qt.ISODate), self.le_notofication.toPlainText(), self.id_info)
                        else:
                            query = """DELETE FROM staff_worker_notification WHERE Worker_Info_Id = %s"""
                            parametrs = (self.id_info, )
                    else:  # Если напоминания надо добавить
                        if self.cb_notification.isChecked():
                            query = "INSERT INTO staff_worker_notification (Worker_Info_Id, Note, Date) VALUES (%s, %s, %s)"
                            parametrs = (self.id_info, self.le_notofication.toPlainText(), self.de_notification.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql nt", info_sql.msg, QMessageBox.Ok)
                        return False

                if "login" in self.update_sql:  # Если логин надо обновить
                    if "login" in self.alert and "login" not in self.delete and self.le_login_login.text():
                        query = """UPDATE staff_worker_login SET Login = %s, Password = %s WHERE Worker_Info_Id = %s"""
                        parametrs = (self.le_login_login.text(), self.le_login_password.text(), self.id_info)
                    elif "login" in self.delete:
                        query = """DELETE FROM staff_worker_login WHERE Worker_Info_Id = %s"""
                        parametrs = (self.id_info, )
                elif self.le_login_login.text():  # Если логин надо добавить
                    query = "INSERT INTO staff_worker_login (Worker_Info_Id, Login, Password) VALUES (%s, %s, %s)"
                    parametrs = (self.id_info, self.le_login_login.text(), self.le_login_password.text())
                try:
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        if info_sql.errno == 1062:
                            QMessageBox.critical(self, "Ошибка sql", "Такой логин уже есть он не сохраниться!", QMessageBox.Ok)
                            return False
                        else:
                            QMessageBox.critical(self, "Ошибка sql lg", info_sql.msg, QMessageBox.Ok)
                            return False
                except:
                    pass

                if "beika" in self.alert:  # Если разрешение на бейку надо изменить
                    if self.cb_beika.isChecked():
                        if not self.sql_beika:
                            query = "INSERT INTO access (Worker_Id, Class, Atr1, Atr2, Atr_Value) VALUES (%s, %s, %s, %s, %s)"
                            parametrs = (self.id_info, "MainWindowOperation", "pb_beika", "setEnabled", 'True')
                    else:
                        if self.sql_beika:
                            query = "DELETE FROM access WHERE Worker_Id = %s AND Class = 'MainWindowOperation' AND Atr1 = 'pb_beika'"
                            parametrs = (self.id_info, )

                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql изменения разрешения бейки", info_sql.msg, QMessageBox.Ok)
                        return False

                self.alert = []  # Обнуляем массив для запоминания изменений
                self.close()
                self.m.set_info()
                self.destroy()

        else:  # Если мы добавляем работника
            if self.input_check():  # Проверка заполнености полей
                if "info" in self.alert:  # Добаление основной информации
                    if self.rb_sex_m.isChecked():  # Узнаем пол работника
                        self.sex = "M"
                    elif self.rb_sex_f.isChecked():
                        self.sex = "F"

                    if self.rb_employed.isChecked():  # Узнаем принят уволен непринят
                        self.leave = 0
                    elif self.rb_leave.isChecked():
                        self.leave = 1
                    elif self.rb_not_employed.isChecked():
                        self.leave = 2

                    if self.cb_ip.isChecked():
                        ip = 1
                    else:
                        ip = 0

                    id_country = my_sql.sql_select("SELECT Id FROM staff_country WHERE Country_name = %s", (self.cb_info_country.currentText(),))[0][0]
                    id_position = my_sql.sql_select("SELECT Id FROM staff_position WHERE Name = %s", (self.cb_info_position.currentText(),))[0][0]

                    query = """INSERT INTO staff_worker_info (First_Name, Last_Name, Middle_Name, Sex, Date_Birth, Birthplace, Date_Recruitment, `Leave`,
                                                              Date_Leave, Country_Id, Phone, Address, Position_Id, INN, SNILS, Note, Ip)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                    parametrs = (self.le_info_first_name.text(), self.le_info_last_name.text(), self.le_info_middle_name.text(), self.sex,
                                 self.de_info_birth.date().toString(Qt.ISODate), self.le_info_birthplace.text(),
                                 self.de_info_recruitment.date().toString(Qt.ISODate),
                                 self.leave, self.de_info_leave.date().toString(Qt.ISODate), id_country, self.le_info_phone.text(),
                                 self.le_info_address.text(), id_position, self.le_info_inn.text(), self.le_info_snils.text(), self.le_info_note.toPlainText(), ip)
                    self.id_info = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(self.id_info)):
                        QMessageBox.critical(self, "Ошибка sql", self.id_info.msg, QMessageBox.Ok)
                        return False

                    # Авто выдача логина. Пока не знаю надо ли при добавлении работника!
                    # if self.rb_employed.isChecked():
                    #     query = "INSERT INTO staff_worker_login (Worker_Info_Id, Login, Password) VALUES (%s, %s, %s)"
                    #     parametrs = (self.id_info, self.id_info, "")
                    #     info_sql = my_sql.sql_change(query, parametrs)
                    #     if "mysql.connector.errors" in str(type(info_sql)):
                    #         QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                    #         return False

                if "passport" in self.alert and "passport" in self.alert2:
                    query = "INSERT INTO staff_worker_passport (Worker_Info_Id, Series, Number, Issued, Data_Issued, Date_Ending) VALUES (%s, %s, %s, %s, %s, %s)"
                    parametrs = (self.id_info, self.le_passport_series.text(), self.le_passport_number.text(), self.le_passport_issued.text(),
                                 self.de_passport_issued.date().toString(Qt.ISODate), self.de_passport_ending.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False

                if "migration" in self.alert and "migration" in self.alert2:
                    query = "INSERT INTO staff_worker_migration (Worker_Info_Id, Serial, Number, KPP, Date_migration, Date_Validity_From, Date_Validity_To) VALUES (%s, %s, %s, %s, %s, %s, %s)"
                    parametrs = (self.id_info, self.le_migration_serial.text(), self.le_migration_number.text(), self.le_migration_kpp.text(),
                                 self.de_migration.date().toString(Qt.ISODate), self.de_migration_validity_from.date().toString(Qt.ISODate),
                                 self.de_migration_validity_to.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False

                if "registration" in self.alert and "registration" in self.alert2:
                    query = "INSERT INTO staff_worker_registraton (Worker_Info_id, Address, Date_Registration, Date_Validity_From, Date_Validity_To) VALUES (%s, %s, %s, %s, %s)"
                    parametrs = (self.id_info, self.le_registration_address.text(), self.de_registration.date().toString(Qt.ISODate),
                                 self.de_registration_validity_from.date().toString(Qt.ISODate), self.de_registration_validity_to.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False

                if "patent" in self.alert and "patent" in self.alert2:
                    query = "INSERT INTO staff_worker_patent (Worker_Info_Id, Serial, Number, Additional_Number, Issued, Data_Issued, Date_Ending) VALUES (%s, %s, %s, %s, %s, %s, %s)"
                    parametrs = (self.id_info, self.le_patent_serial.text(), self.le_patent_number.text(), self.le_patent_additional_number.text(),
                                 self.le_patent_issued.text(), self.de_patent_issued.date().toString(Qt.ISODate),
                                 self.de_patent_ending.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False

                if "insurance" in self.alert and "insurance" in self.alert2:
                    query = "INSERT INTO staff_worker_insurance (Worker_Info_Id, Number, Company, Date) VALUES (%s, %s, %s, %s)"
                    parametrs = (self.id_info, self.le_insurance_number.text(), self.le_insurance_company.text(), self.de_insurance_date.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False

                if "notification" in self.alert:
                    query = "INSERT INTO staff_worker_notification (Worker_Info_Id, Note, Date) VALUES (%s, %s, %s)"
                    parametrs = (self.id_info, self.le_notofication.toPlainText(), self.de_notification.date().toString(Qt.ISODate))
                    info_sql = my_sql.sql_change(query, parametrs)
                    if "mysql.connector.errors" in str(type(info_sql)):
                        QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                        return False

                self.alert = []  # Обнуляем массив для запоминания изменений
                self.close()
                self.m.set_info()
                self.destroy()

    def exel_in(self):
        self.build_new_exel("in")

    def exel_out(self):
        self.build_new_exel("out")

    def exel_petition_in(self):
        self.build_word_petition("in")

    def exel_petition_out(self):
        self.build_word_petition("out")

    # Уведомление о заключении и расторжении трудового договора
    def build_new_exel(self, option):
        # Уведомление о заключении и расторжении трудового договора

        # Колонки EXCEL
        all_col = ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z",
               "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL")

        to_excel.patch_worksheet()

        self.statusBar().showMessage("Открываю шаблон")
        if option == "in":
            info = InfoDate(self.de_info_recruitment.date())
            if info.exec() == 0:
                return False
            book = openpyxl.load_workbook(filename=getcwd() + '/templates/staff/notif_in.xlsx')
        elif option == "out":
            book = openpyxl.load_workbook(filename=getcwd() + '/templates/staff/notif_out.xlsx')
        sheet = book['s2']

        # Если работник ИП то поменяем шапку
        if self.cb_ip.isChecked():

            sheet = book['s1']
            col = all_col[26:]
            for i, t in enumerate("46.64"):
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (col[i], 29)] = t

            for i, t in enumerate("ИНДИВИДУАЛЬНЫЙ ПРЕДПРИНИМАТЕЛЬ"):
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (all_col[i], 31)] = t

            for i, t in enumerate("РУБЛЁВ АЛЕКСАНДР    "):
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (all_col[i], 33)] = t

            for i, t in enumerate("АЛЕКСАНДРОВИЧ    "):
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (all_col[i], 35)] = t

            for i, t in enumerate("ОГРНИП 317774600502549"):
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (all_col[i], 37)] = t

            sheet = book['s2']

            for i, t in enumerate("ИНН 773013683314  "):
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (all_col[i], 7)] = t

            for i, t in enumerate("КПП 773001001  "):
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (all_col[i], 9)] = t

            for i, t in enumerate("121601 МОСКВА ФИЛЁВСКИЙ БУЛЬВАР   "):
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (all_col[i], 11)] = t

            for i, t in enumerate("ДОМ 40 КВАРТИРА 334   "):
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (all_col[i], 13)] = t

            col = all_col[8:]
            for i, t in enumerate("89299024574 "):
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (col[i], 15)] = t

            sheet = book['s4']
            sheet['R7'] = 'ИП Рублёв Александр Александрович'
            sheet = book['s2']

        col = all_col[8:]
        text = tuple(self.le_info_last_name.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Фамилия длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 19)] = t
            i += 1

        text = tuple(self.le_info_first_name.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Имя длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 21)] = t
            i += 1

        text = tuple(self.le_info_middle_name.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Отчество длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 23)] = t
            i += 1

        text = tuple(self.cb_info_country.currentText().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Место рождения длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 25)] = t
            i += 1

        col = [x + '27' for x in all_col[14:]]
        col += [x + '29' for x in all_col]

        text = tuple(self.le_info_birthplace.text().upper())
        n = 0
        if len(col) < len(text):
            m = len(col)
            while text[m] != " ":
                m -= 1
        else:
            m = len(text)
        i = 0
        while n < m:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s' % col[i]] = text[n]
            n += 1
            i += 1

        col = ("I", "J", "M", "N", "Q", "R", "S", "T")
        text = tuple(self.de_info_birth.date().toString("ddMMyyyy"))
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 31)] = t
            i += 1

        # Заполняем пасспорт
        col = all_col[3:12]
        text = tuple(self.le_passport_series.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Серия паспорта длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 35)] = t
            i += 1

        col = all_col[13:23]
        text = tuple(self.le_passport_number.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Номер паспорта длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 35)] = t
            i += 1

        col = ("AC", "AD", "AF", "AG", "AI", "AJ", "AK", "AL")
        text = tuple(self.de_passport_issued.date().toString("ddMMyyyy"))
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Дата выдачи паспорта длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 35)] = t
            i += 1

        col = [x + '37' for x in all_col[4:]]
        col += [x + '39' for x in all_col[4:]]
        col += [x + '41' for x in all_col[4:]]
        text = tuple(self.le_passport_issued.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Кем выдан паспорт, длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s' % (col[i])] = t
            i += 1

        # Миграционка
        # col = all_col[14:23]
        # text = tuple(self.le_migration_number.text().upper())
        # if len(text) > len(col):
        #     QMessageBox.critical(self, "Ошибка", "Номер миграционки длиннее строки ввода", QMessageBox.Ok)
        #     return False
        # i = 0
        # for t in text:
        #     self.statusBar().showMessage("Создаю %s" % i)
        #     sheet['%s%s' % (col[i], 7)] = t
        #     i += 1
        #
        # col = ("AE", "AF", "AI", "AJ", "AM", "AN", "AO", "AP")
        # text = tuple(self.de_migration.date().toString("ddMMyyyy"))
        # if len(text) > len(col):
        #     QMessageBox.critical(self, "Ошибка", "Дата выдачи миграционки длиннее строки ввода", QMessageBox.Ok)
        #     return False
        # i = 0
        # for t in text:
        #     self.statusBar().showMessage("Создаю %s" % i)
        #     sheet['%s%s' % (col[i], 7)] = t
        #     i += 1
        #
        # col = [x + '10' for x in all_col[17:]]
        # col += [x + '12' for x in all_col]
        # col += [x + '14' for x in all_col]
        # text = tuple(self.le_registration_address.text().upper())
        # n = 0
        # if len(col) < len(text):
        #     m = len(col)
        #     try:
        #         while text[m] != " ":
        #             m -= 1
        #     except IndexError:
        #         QMessageBox.critical(self, "Ошибка", "Не могу разделить место регистрации на строки, добавте символ пробела в месте переноса строки!", QMessageBox.Ok)
        #         return False
        # else:
        #     m = len(text)
        # i = 0
        # while n < m:
        #     self.statusBar().showMessage("Создаю %s" % i)
        #     sheet['%s' % col[i]] = text[n]
        #     n += 1
        #     i += 1
        #
        # col = ("T", "U", "X", "Y", "AB", "AC", "AD", "AE")
        # text = tuple(self.de_registration.date().toString("ddMMyyyy"))
        # if len(text) > len(col):
        #     QMessageBox.critical(self, "Ошибка", "Дата регистрации длиннее строки ввода", QMessageBox.Ok)
        #     return False
        # i = 0
        # for t in text:
        #     self.statusBar().showMessage("Создаю %s" % i)
        #     sheet['%s%s' % (col[i], 16)] = t
        #     i += 1


        # Патент
        sheet = book['s3']
        country = my_sql.sql_select("SELECT Patent, Act FROM staff_country WHERE Country_name = %s", (self.cb_info_country.currentText(),))[0]
        if country[0] == 1:
            col = all_col[11:]
            text = tuple("ПАТЕНТ " + self.le_patent_additional_number.text().upper())
            if len(text) > len(col):
                QMessageBox.critical(self, "Ошибка", "Доп номер патента длиннее строки ввода", QMessageBox.Ok)
                return False
            i = 0
            for t in text:
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (col[i], 3)] = t
                i += 1

            col = all_col[3:12]
            text = tuple(self.le_patent_serial.text().upper())
            if len(text) > len(col):
                QMessageBox.critical(self, "Ошибка", "Серия патента длиннее строки ввода", QMessageBox.Ok)
                return False
            i = 0
            for t in text:
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (col[i], 5)] = t
                i += 1

            col = all_col[13:23]
            text = tuple(self.le_patent_number.text().upper())
            if len(text) > len(col):
                QMessageBox.critical(self, "Ошибка", "Номер патента длиннее строки ввода", QMessageBox.Ok)
                return False
            i = 0
            for t in text:
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (col[i], 5)] = t
                i += 1

            col = ("AC", "AD", "AF", "AG", "AI", "AJ", "AK", "AL")
            text = tuple(self.de_patent_issued.date().toString("ddMMyyyy"))
            if len(text) > len(col):
                QMessageBox.critical(self, "Ошибка", "Дата выдачи патента длиннее строки ввода", QMessageBox.Ok)
                return False
            i = 0
            for t in text:
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (col[i], 5)] = t
                i += 1

            col = [x + '8' for x in all_col[5:]]
            col += [x + '10' for x in all_col]
            text = tuple(self.le_patent_issued.text().upper())
            if len(text) > len(col):
                QMessageBox.critical(self, "Ошибка", "Кем выдан патент длиннее строки ввода", QMessageBox.Ok)
                return False
            i = 0
            for t in text:
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s' % (col[i])] = t
                i += 1

            col = ("J", "K", "N", "O", "R", "S", "T", "U")
            text = tuple(self.de_patent_issued.date().toString("ddMMyyyy"))
            if len(text) > len(col):
                QMessageBox.critical(self, "Ошибка", "Срок С патента длиннее строки ввода", QMessageBox.Ok)
                return False
            i = 0
            for t in text:
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (col[i], 12)] = t
                i += 1

            col = ("Y", "Z", "AC", "AD", "AG", "AH", "AI", "AJ")
            text = tuple(self.de_patent_ending.date().toString("ddMMyyyy"))
            if len(text) > len(col):
                QMessageBox.critical(self, "Ошибка", "Срок ДО патента длиннее строки ввода", QMessageBox.Ok)
                return False
            i = 0
            for t in text:
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (col[i], 12)] = t
                i += 1
        else:

            col = []
            col.append([x + '20' for x in all_col])
            col.append([x + '22' for x in all_col])
            col.append([x + '24' for x in all_col])
            text = tuple(country[1].upper())
            m, n = 0, 0
            for t in text:
                if t == "\n":
                    m += 1
                    n = 0
                else:
                    self.statusBar().showMessage("Создаю %s" % n)
                    try:
                        sheet['%s' % (col[m][n])] = t
                        n += 1
                    except IndexError:
                        QMessageBox.critical(self, "Ошибка", "Текст патента длинее строк ввода!", QMessageBox.Ok)
                        return False

        # Должность
        position_number = my_sql.sql_select("SELECT Number FROM staff_position WHERE Name = %s", (self.cb_info_position.currentText(),))[0][0]
        text = tuple(self.cb_info_position.currentText().upper() + " " + position_number)
        if len(text) > len(all_col):
            QMessageBox.critical(self, "Ошибка", "Должность длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (all_col[i], 29)] = t
            i += 1

        # Дата работы / увольнения
        col = ("AA", "AB", "AD", "AE", "AG", "AH", "AI", "AJ")
        if option == "in":
            text = tuple(info.de_in.date().toString("ddMMyyyy"))
        elif option == "out":
            text = tuple(self.de_info_leave.date().toString("ddMMyyyy"))
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "дата договора о работе патента длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 40)] = t
            i += 1

        sheet = book['s4']
        # вставляем сегоднюшнюю дату
        date_now = QDate.currentDate()
        sheet['B11'] = date_now.toString("dd")
        sheet['E11'] = date_now.toString("MM")
        sheet['M11'] = date_now.toString("yy")

        dir_name = self.id_info
        self.path = self.inspection_path(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю")
            if option == "in":
                file_name = "Уведомление о приеме на работу %s.xlsx" % QDate.currentDate().toString("dd.MM.yyyy")
                book.save('%s/%s' % (self.path, file_name))
            elif option == "out":
                file_name = "Уведомление о расторжении договора %s.xlsx" % QDate.currentDate().toString("dd.MM.yyyy")
                book.save('%s/%s' % (self.path, file_name))
            dir_name = self.id_info
            self.path = self.inspection_path(dir_name, 'Путь корень рабочие')
            self.inspection_files(dir_name, 'Путь корень рабочие')
            self.statusBar().showMessage("Готово")
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    def build_new_exel_additional_sheet(self):
        to_excel.patch_worksheet()

        self.statusBar().showMessage("Открываю шаблон")
        book = openpyxl.load_workbook(filename=getcwd() + '/templates/staff/additional_sheet.xlsx')
        sheet = book['s1']

        col = ("I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z",
               "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP",)

        text = tuple(self.le_info_last_name.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Фамилия длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 12)] = t
            i += 1

        text = tuple(self.le_info_first_name.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Имя длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 14)] = t
            i += 1

        text = tuple(self.le_info_middle_name.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Отчество длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 16)] = t
            i += 1

        sheet = book['s2']
        date_now = QDate.currentDate()
        sheet['B58'] = date_now.toString("dd")
        sheet['E58'] = date_now.toString("MM")
        sheet['M58'] = date_now.toString("yy")

        dir_name = self.id_info
        self.path = self.inspection_path(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю")

            file_name = "Доп лист для уведомления %s.xlsx" % QDate.currentDate().toString("dd.MM.yyyy")
            book.save('%s/%s' % (self.path, file_name))

            dir_name = self.id_info
            self.path = self.inspection_path(dir_name, 'Путь корень рабочие')
            self.inspection_files(dir_name, 'Путь корень рабочие')
            self.statusBar().showMessage("Готово")
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Уведомление о регистрации
    def build_exel_registration(self):

        to_excel.patch_worksheet()

        info = ExelInfo(self.le_info_birthplace.text())
        if info.exec() == 0:
            return False
        self.statusBar().showMessage("Открываю шаблон")
        book = openpyxl.load_workbook(filename=getcwd() + '/templates/staff/registration.xlsx')
        sheet = book["s1"]
        col = ("N", "Q", "T", "W", "Z", "AC", "AF", "AI", "AL", "AO", "AR", "AU", "AX", "BA", "BD", "BG", "BJ", "BM", "BP", "BS", "BV", "BY", "CB", "CE", "CH", "CK",
               "CN", "CQ", "CT", "CW", "CZ", "DC", "DF", "DI", "DL")

        text = tuple(self.le_info_last_name.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Фамилия длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 16)] = t
            sheet['%s%s' % (col[i], 69)] = t
            i += 1

        text = tuple(self.le_info_first_name.text().upper() + " " + self.le_info_middle_name.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Имя + отчество длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 18)] = t
            sheet['%s%s' % (col[i], 71)] = t
            i += 1

        col = ("Q", "T", "W", "Z", "AC", "AF", "AI", "AL", "AO", "AR", "AU", "AX", "BA", "BD", "BG", "BJ", "BM", "BP", "BS", "BV", "BY", "CB", "CE", "CH", "CK",
               "CN", "CQ", "CT", "CW", "CZ", "DC", "DF", "DI", "DL")
        text = tuple(self.cb_info_country.currentText().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Гражданство длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 21)] = t
            sheet['%s%s' % (col[i], 74)] = t
            i += 1

        col = ("T", "W", "AF", "AI", "AO", "AR", "AU", "AX")
        text = tuple(self.de_info_birth.date().toString("ddMMyyyy"))
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 24)] = t
            sheet['%s%s' % (col[i], 77)] = t
            i += 1

        if self.rb_sex_m.isChecked():  # Узнаем пол работника
            sheet['%s%s' % ("BV", 24)] = "Х"
            sheet['%s%s' % ("BY", 77)] = "Х"
        elif self.rb_sex_f.isChecked():
            sheet['%s%s' % ("CK", 24)] = "Х"
            sheet['%s%s' % ("CN", 77)] = "Х"

        col = ("T", "W", "Z", "AC", "AF", "AI", "AL", "AO", "AR", "AU", "AX", "BA", "BD", "BG", "BJ", "BM", "BP", "BS", "BV", "BY", "CB", "CE", "CH", "CK",
               "CN", "CQ", "CT", "CW", "CZ", "DC", "DF", "DI", "DL")
        text = tuple(info.le_birthplace_country.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Государство места рождения длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 27)] = t
            i += 1

        text = tuple(info.le_birthplace_city.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Город места рождения длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 30)] = t
            i += 1

        col = ("BY", "CB", "CE", "CH")
        text = tuple(self.le_passport_series.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Серия паспорта длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 33)] = t
            sheet['%s%s' % (col[i], 80)] = t
            i += 1

        col = ("CN", "CQ", "CT", "CW", "CZ", "DC", "DF", "DI", "DL")
        text = tuple(self.le_passport_number.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Номер паспорта длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 33)] = t
            sheet['%s%s' % (col[i], 80)] = t
            i += 1

        col = ("Q", "T", "AC", "AF", "AL", "AO", "AR", "AU")
        text = tuple(self.de_passport_issued.date().toString("ddMMyyyy"))
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 35)] = t
            i += 1

        col = ("BM", "BP", "BY", "CB", "CH", "CK", "CN", "CQ")
        text = tuple(self.de_passport_ending.date().toString("ddMMyyyy"))
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 35)] = t
            i += 1

        col = ("N", "Q", "T", "W", "Z", "AC", "AF", "AI", "AL", "AO", "AR", "AU", "AX", "BA", "BD", "BG", "BJ", "BM", "BP", "BS", "BV", "BY", "CB")
        text = tuple(self.cb_info_position.currentText().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Должность длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 46)] = t
            i += 1

        col = ("W", "Z", "AI", "AL", "AR", "AU", "AX", "BA")
        text = tuple(info.de_in.date().toString("ddMMyyyy"))
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 48)] = t
            i += 1

        col = ("CH", "CK", "CT", "CW", "DC", "DF", "DI", "DL")
        text = tuple(info.de_from.date().toString("ddMMyyyy"))
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 48)] = t
            i += 1

        col = ("AC", "AF", "AO", "AR", "AX", "BA", "BD", "BG")
        text = tuple(info.de_from.date().toString("ddMMyyyy"))
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 94)] = t
            i += 1

        col = ("AC", "AF", "AI", "AL")
        text = tuple(self.le_migration_serial.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Серия миграционки длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 50)] = t
            i += 1

        col = ("AR", "AU", "AX", "BA", "BD", "BG", "BJ", "BM")
        text = tuple(self.le_migration_number.text().upper())
        if len(text) > len(col):
            QMessageBox.critical(self, "Ошибка", "Номер миграционки длиннее строки ввода", QMessageBox.Ok)
            return False
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 50)] = t
            i += 1

        # Заполнение предыдущего метса пребывания
        col = ("Q", "T", "W", "Z", "AC", "AF", "AI", "AL", "AO", "AR", "AU", "AX", "BA", "BD", "BG", "BJ", "BM", "BP", "BS")
        text = info.te_old_addres.toPlainText().upper()
        line_text = text.split('\n')
        if len(line_text) > 3:
            QMessageBox.critical(self, "Ошибка", "кол-во строк в прежнем месте больше 3", QMessageBox.Ok)
            return False
        row = 57
        for line in line_text:
            text = tuple(line)
            if len(text) > len(col):
                QMessageBox.critical(self, "Ошибка", "Прежний адрес длиннее строки ввода", QMessageBox.Ok)
                return False
            i = 0
            for t in text:
                self.statusBar().showMessage("Создаю %s" % i)
                sheet['%s%s' % (col[i], row)] = t
                i += 1
            row += 2

        sheet = book["s2"]

        col = ("CH", "CK", "CT", "CW", "DC", "DF", "DI", "DL")
        text = tuple(info.de_from.date().toString("ddMMyyyy"))
        i = 0
        for t in text:
            self.statusBar().showMessage("Создаю %s" % i)
            sheet['%s%s' % (col[i], 67)] = t
            i += 1

        # Вставляем черные квадраты
        self.path_templates = getcwd() + '/templates'

        sheet = book["s1"]

        dir_name = self.id_info
        self.path = self.inspection_path(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            file_name = "Уведомление для регистрации %s.xlsx" % QDate.currentDate().toString("dd.MM.yyyy")
            book.save('%s/%s' % (self.path, file_name))
            self.inspection_files(dir_name, 'Путь корень рабочие')
            self.statusBar().showMessage("Готово")
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Договор на работу
    def build_word_in(self):

        if not self.id_info:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет номера", QMessageBox.Ok)
            return False

        # Проверяем нужный номер документа
        self.statusBar().showMessage("проверяю SQL")
        doc_date_new = False
        query = "SELECT Number, Date FROM staff_worker_doc_number WHERE Worker_Info_Id = %s AND Name = %s"
        doc_number_sql = my_sql.sql_select(query, (self.id_info, "труд.дог."))
        if "mysql.connector.errors" in str(type(doc_number_sql)):
            QMessageBox.critical(self, "Ошибка sql", doc_number_sql.msg, QMessageBox.Ok)

        if doc_number_sql:
            doc_number = doc_number_sql[0][0]
            doc_date = doc_number_sql[0][1]
        else:
            query = "SELECT IFNULL(MAX(CAST(Number AS SIGNED)), 'No Number') FROM staff_worker_doc_number WHERE Name = %s AND YEAR(Date) = %s"
            doc_number = my_sql.sql_select(query, ("труд.дог.", QDate.currentDate().year()))
            if "mysql.connector.errors" in str(type(doc_number)):
                QMessageBox.critical(self, "Ошибка sql", doc_number.msg, QMessageBox.Ok)

            doc_number = doc_number[0][0]
            if "No Number" != doc_number:
                doc_number = int(doc_number) + 1
                doc_date = self.de_info_recruitment.date().toPyDate()
                doc_date_new = True
            else:
                doc_number = 1
                doc_date = self.de_info_recruitment.date().toPyDate()
                doc_date_new = True

        info = InfoDateNumber(doc_date, doc_number)
        if info.exec() == 0:
            self.statusBar().showMessage("Отмена")
            return False

        if info.de_in.date().toPyDate() != doc_date or info.le_number.text() != str(doc_number):
            doc_date = info.de_in.date().toPyDate()
            doc_number = info.le_number.text()
            doc_date_new = True

        # Нужен ли патент
        patent = my_sql.sql_select("SELECT Patent FROM staff_country WHERE Country_name = %s", (self.cb_info_country.currentText(),))[0][0]

        self.statusBar().showMessage("Открываю шаблон")
        f = open(getcwd() + '/templates/staff/contract.xml', "r", -1, "utf-8")
        xml = f.read()
        self.statusBar().showMessage("Закрываю шаблон")
        f.close()
        self.statusBar().showMessage("Создаю документ")
        number_xml = str(doc_number) + "/" + doc_date.strftime("%y")

        xml = xml.replace("НОМЕР", number_xml)
        xml = xml.replace("ДАТА", info.de_in.date().toString("dd.MM.yyyy"))
        xml = xml.replace("ПРОФЕССИЯ", self.cb_info_position.currentText())
        xml = xml.replace("ФИО", self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.le_info_middle_name.text())
        if self.rb_sex_m.isChecked():  # Узнаем пол работника
            xml = xml.replace("ЫЙ/АЯ", "ый")
        elif self.rb_sex_f.isChecked():
            xml = xml.replace("ЫЙ/АЯ", "ая")
        xml = xml.replace("ПРОФФ", self.cb_info_position.currentText().lower())
        if patent == 1:
            xml = xml.replace("ВРЕМЯ", "время действия патента серии")
            xml = xml.replace("ПАТЕНТ", self.le_patent_serial.text() + " № " + self.le_patent_number.text() + ".")
        else:
            xml = xml.replace("ВРЕМЯ", "неопределенный срок.")
            xml = xml.replace("ПАТЕНТ", "")
        xml = xml.replace("ДАТРОЖ", self.de_info_birth.date().toString("dd.MM.yyyy"))
        xml = xml.replace("ГРАЖДАНСТВО", self.cb_info_country.currentText())
        xml = xml.replace("ПАССПОРТ", self.le_passport_series.text().upper() + " " + self.le_passport_number.text().upper())
        if patent == 1:
            xml = xml.replace("?ПА", "Патент: ")
            xml = xml.replace("СТРАХОВКА", "ДМС: №" + self.le_insurance_number.text().upper() + ", от " + self.de_insurance_date.date().toString("dd.MM.yyyy") +
                              "г., страховая компания \"" + self.le_insurance_company.text() + '\"')
        else:
            xml = xml.replace("СТРАХОВКА", "")
            xml = xml.replace("?ПА", "")

        dir_name = self.id_info
        self.inspection_files(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            file_name = "Трудовой договор %s.doc" % QDate.currentDate().toString("dd.MM.yyyy")
            f = open('%s/%s' % (self.path, file_name), "w", -1, "utf-8")
            f.write(xml)
            f.close()
            if doc_date_new:
                query = """INSERT INTO staff_worker_doc_number (Worker_Info_Id, Name, Number, Date) VALUES (%s, %s, %s, %s)
                            ON DUPLICATE KEY UPDATE Number = %s, Date = %s"""
                parametrs = (self.id_info, "труд.дог.", doc_number, doc_date, doc_number, doc_date)
                info_sql = my_sql.sql_change(query, parametrs)
                if "mysql.connector.errors" in str(type(info_sql)):
                    QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                    return False
            self.statusBar().showMessage("Готово")
            self.inspection_files(dir_name, 'Путь корень рабочие')
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Заявления на прием и увольнение
    def build_word_petition(self, option):

        if option == "out":
            info = InfoDate(QDate().currentDate())
            info.label.setText("Дата увольнения")
            if info.exec() == 0:
                return False

        if not self.id_info:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет номера", QMessageBox.Ok)
            return False

        self.statusBar().showMessage("Открываю шаблон")
        if option == "out":
            f = open(getcwd() + '/templates/staff/petition_out.xml', "r", -1, "utf-8")
        else:
            f = open(getcwd() + '/templates/staff/petition_in.xml', "r", -1, "utf-8")
        xml = f.read()
        self.statusBar().showMessage("Закрываю шаблон")
        f.close()
        self.statusBar().showMessage("Создаю документ")

        # Если работник ИП то поменяем шапку
        if self.cb_ip.isChecked():
            xml = xml.replace("?КОМУ", """ИП Рублёву Александру Александровичу""")
        else:
            xml = xml.replace("?КОМУ", """Генеральному директору
                                          ООО «Авидевелопмент-М»	
                                          Ширяевой Н.А.""")

        xml = xml.replace("?ФИО", self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.le_info_middle_name.text())
        xml = xml.replace("?СЕРИЯ", self.le_passport_series.text().upper())
        xml = xml.replace("?НОМЕР", self.le_passport_number.text().upper())
        xml = xml.replace("?ВЫДАН", self.le_passport_issued.text())
        xml = xml.replace("?ДАТАВЫД", self.de_passport_issued.date().toString("dd.MM.yyyy"))
        xml = xml.replace("?ДАТАЗАЯВ", QDate().currentDate().toString("dd.MM.yyyy"))
        if option == "out":
            xml = xml.replace("?ДАТАУВОЛЬН", info.de_in.date().toString("dd.MM.yyyy"))
        else:
            xml = xml.replace("?ДОЛЖНОСТЬ", self.cb_info_position.currentText())

        dir_name = self.id_info
        self.inspection_files(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            if option == "out":
                file_name = "Заявление об увольнении %s.doc" % QDate.currentDate().toString("dd.MM.yyyy")
                f = open('%s/%s' % (self.path, file_name), "w", -1, "utf-8")
            else:
                file_name = "Заявление о приеме на работу %s.doc" % QDate.currentDate().toString("dd.MM.yyyy")
                f = open('%s/%s' % (self.path, file_name), "w", -1, "utf-8")
            f.write(xml)
            f.close()
            self.statusBar().showMessage("Готово")
            self.inspection_files(dir_name, 'Путь корень рабочие')
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Довереность на ЗП
    def build_word_proxy(self):

        if not self.id_info:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет номера", QMessageBox.Ok)
            return False

        self.statusBar().showMessage("Открываю шаблон")
        f = open(getcwd() + '/templates/staff/proxy.xml', "r", -1, "utf-8")
        xml = f.read()
        self.statusBar().showMessage("Закрываю шаблон")
        f.close()
        self.statusBar().showMessage("Создаю документ")

        xml = xml.replace("?ФИО", self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.le_info_middle_name.text())
        xml = xml.replace("?СЕРИЯ", self.le_passport_series.text().upper())
        xml = xml.replace("?НОМЕР", self.le_passport_number.text().upper())
        xml = xml.replace("?ДАТАЗАЯВ", QDate().currentDate().toString("dd.MM.yyyy"))

        dir_name = self.id_info
        self.inspection_files(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            file_name = "Доверенность на ЗП %s.doc" % QDate.currentDate().toString("dd.MM.yyyy")
            f = open('%s/%s' % (self.path, file_name), "w", -1, "utf-8")
            f.write(xml)
            f.close()
            self.statusBar().showMessage("Готово")
            self.inspection_files(dir_name, 'Путь корень рабочие')
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Временный пропуск
    def build_word_pass(self):

        if not self.id_info:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет номера", QMessageBox.Ok)
            return False

        self.statusBar().showMessage("Открываю шаблон")
        f = open(getcwd() + '/templates/staff/pass.xml', "r", -1, "utf-8")
        xml = f.read()
        self.statusBar().showMessage("Закрываю шаблон")
        f.close()
        self.statusBar().showMessage("Создаю документ")

        xml = xml.replace("?ФИО", self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.le_info_middle_name.text())

        dir_name = self.id_info
        self.inspection_files(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            file_name = "Временный пропуск %s.doc" % QDate.currentDate().toString("dd.MM.yyyy")
            f = open('%s/%s' % (self.path, file_name), "w", -1, "utf-8")
            f.write(xml)
            f.close()
            self.statusBar().showMessage("Готово")
            self.inspection_files(dir_name, 'Путь корень рабочие')
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Справка о приеме уведомления
    def build_word_admission_notice(self):

        info = InfoDate(QDate().currentDate())
        info.label.setText("Дата приема уведомления")
        if info.exec() == 0:
            return False

        if not self.id_info:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет номера", QMessageBox.Ok)
            return False

        self.statusBar().showMessage("Открываю шаблон")
        f = open(getcwd() + '/templates/staff/admission_notice.xml', "r", -1, "utf-8")
        xml = f.read()
        self.statusBar().showMessage("Закрываю шаблон")
        f.close()
        self.statusBar().showMessage("Создаю документ")

        xml = xml.replace("?ФИ", self.le_info_last_name.text() + " " + self.le_info_first_name.text())
        xml = xml.replace("?ДАТА", info.de_in.date().toString("dd.MM.yyyy"))

        dir_name = self.id_info
        self.inspection_files(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            file_name = "Справка о приеме уведомления %s.doc" % QDate.currentDate().toString("dd.MM.yyyy")
            f = open('%s/%s' % (self.path, file_name), "w", -1, "utf-8")
            f.write(xml)
            f.close()
            self.statusBar().showMessage("Готово")
            self.inspection_files(dir_name, 'Путь корень рабочие')
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Ходотайство о продлении регистрации
    def build_word_hodataistvo(self):

        if not self.id_info:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет номера", QMessageBox.Ok)
            return False

        info = InfoDate(QDate().currentDate())
        info.label.setText("Продлить до")
        if info.exec() == 0:
            return False

        # Проверяем нужный номер документа
        self.statusBar().showMessage("проверяю SQL")
        query = "SELECT IFNULL(MAX(CAST(Number AS SIGNED)), 'No Number') FROM staff_worker_doc_number WHERE Name = %s AND YEAR(Date) = %s"
        doc_number = my_sql.sql_select(query, ("ходатайство", info.de_in.date().year()))
        if "mysql.connector.errors" in str(type(doc_number)):
            QMessageBox.critical(self, "Ошибка sql", doc_number.msg, QMessageBox.Ok)
        doc_number = doc_number[0][0]

        if "No Number" != doc_number:
            doc_number = int(doc_number) + 1
        else:
            doc_number = 1

        # Узнаем номер договора
        query = "SELECT IFNULL(MAX(CAST(Number AS SIGNED)), 'No Number'), Date FROM staff_worker_doc_number WHERE Name = %s AND Worker_Info_Id = %s"
        contract_number = my_sql.sql_select(query, ("труд.дог.", self.id_info))
        if "mysql.connector.errors" in str(type(doc_number)):
            QMessageBox.critical(self, "Ошибка sql", contract_number.msg, QMessageBox.Ok)
            return False

        # Нужен ли патент
        patent = my_sql.sql_select("SELECT Patent FROM staff_country WHERE Country_name = %s", (self.cb_info_country.currentText(),))[0][0]

        if not self.id_info:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет номера", QMessageBox.Ok)
            return False

        self.statusBar().showMessage("Открываю шаблон")
        f = open(getcwd() + '/templates/staff/hodataistvo.xml', "r", -1, "utf-8")
        xml = f.read()
        self.statusBar().showMessage("Закрываю шаблон")
        f.close()
        self.statusBar().showMessage("Создаю документ")

        xml = xml.replace("?НОМЕРДОК", str(doc_number))
        xml = xml.replace("?ДАТАДОК", QDate().currentDate().toString("dd.MM.yyyy"))
        xml = xml.replace("?ДАТАДО", info.de_in.date().toString("dd.MM.yyyy"))
        xml = xml.replace("?РЕСПУБЛИКА", self.cb_info_country.currentText())
        xml = xml.replace("?ФИО", self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.le_info_middle_name.text())
        xml = xml.replace("?ДАТАРОЖ", self.de_info_birth.date().toString("dd.MM.yyyy"))
        xml = xml.replace("?ПАССПОРТ", self.le_passport_series.text().upper() + " " + self.le_passport_number.text().upper())
        xml = xml.replace("?ДАТАВЬЕЗД", self.de_migration.text())
        if contract_number[0][0] != 'No Number':
            xml = xml.replace("?ТРУДДОГ", "№ " + contract_number[0][0] + "/" + contract_number[0][1].strftime("%y") + " от " + contract_number[0][1].strftime("%d.%m.%Y"))
        else:
            xml = xml.replace("?ТРУДДОГ", "договор не найден")
        xml = xml.replace("?РЕГДО", self.de_registration_validity_to.date().toString("dd.MM.yyyy"))
        xml = xml.replace("?РАДР", self.le_registration_address.text())
        if patent == 1:
            xml = xml.replace("?ПАТЕНТ", "-патент на работу " + self.le_patent_serial.text() + " № " + self.le_patent_number.text() + ", действительный до " + self.de_patent_ending.date().toString("dd.MM.yyyy"))
        else:
            xml = xml.replace("?ПАТЕНТ", " ")

        dir_name = self.id_info
        self.inspection_files(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            file_name = "Ходатайство %s.doc" % QDate.currentDate().toString("dd.MM.yyyy")
            f = open('%s/%s' % (self.path, file_name), "w", -1, "utf-8")
            f.write(xml)
            f.close()
            query = """INSERT INTO staff_worker_doc_number (Worker_Info_Id, Name, Number, Date) VALUES (%s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE Number = %s, Date = %s"""
            parametrs = (self.id_info, "ходатайство", doc_number, QDate.currentDate().toString(Qt.ISODate), doc_number, info.de_in.date().toString(Qt.ISODate))
            info_sql = my_sql.sql_change(query, parametrs)
            if "mysql.connector.errors" in str(type(info_sql)):
                QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                return False
            self.statusBar().showMessage("Готово")
            self.inspection_files(dir_name, 'Путь корень рабочие')
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Приказ о приеме работника на работу
    def build_word_order_on_reception(self):
        if not self.id_info:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет номера", QMessageBox.Ok)
            return False

        # Проверяем нужный номер документа
        self.statusBar().showMessage("проверяю SQL")
        query = "SELECT IFNULL(MAX(CAST(Number AS SIGNED)), 'No Number') FROM staff_worker_doc_number WHERE Name = %s AND YEAR(Date) = %s"
        doc_number_sql = my_sql.sql_select(query, ("приказ.прием", QDate.currentDate().year()))
        if "mysql.connector.errors" in str(type(doc_number_sql)):
            QMessageBox.critical(self, "Ошибка sql", doc_number_sql.msg, QMessageBox.Ok)

        if ("No Number", ) not in doc_number_sql:
            doc_number = int(doc_number_sql[0][0]) + 1
            doc_date = self.de_info_recruitment.date().toPyDate()
        else:
            doc_number = 1
            doc_date = self.de_info_recruitment.date().toPyDate()

        info = InfoDateNumber(doc_date, doc_number)
        if info.exec() == 0:
            self.statusBar().showMessage("Отмена")
            return False

        if info.de_in.date().toPyDate() != doc_date or info.le_number.text() != str(doc_number):
            doc_date = info.de_in.date().toPyDate()
            doc_number = info.le_number.text()

        # Узнаем номер договора
        query = "SELECT IFNULL(MAX(CAST(Number AS SIGNED)), 'No Number'), Date FROM staff_worker_doc_number WHERE Name = %s AND Worker_Info_Id = %s"
        contract_number = my_sql.sql_select(query, ("труд.дог.", self.id_info))
        if "mysql.connector.errors" in str(type(contract_number)):
            QMessageBox.critical(self, "Ошибка sql", contract_number.msg, QMessageBox.Ok)
            return False

        if contract_number[0][1] is None:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет трудового договора", QMessageBox.Ok)
            return False

        self.statusBar().showMessage("Открываю шаблон")
        f = open(getcwd() + '/templates/staff/order_on_reception.xml', "r", -1, "utf-8")
        xml = f.read()
        self.statusBar().showMessage("Закрываю шаблон")
        f.close()
        self.statusBar().showMessage("Создаю документ")

        # Если работник ИП то поменяем шапку
        if self.cb_ip.isChecked():
            xml = xml.replace("?ФИРМА", 'ИП Рублёв Александр Александрович"')
            xml = xml.replace("?ПОДПИСАНТ", 'Рублёв А.А."')
        else:
            xml = xml.replace("?ФИРМА", 'Общество с ограниченной ответственностью "Авидевелопмент-М"')
            xml = xml.replace("?ПОДПИСАНТ", 'Ширяева Н.А.')

        xml = xml.replace("?НОМЕРДОК", str(doc_number))
        xml = xml.replace("?ДАТАСОСТ", doc_date.strftime("%d.%m.%Y"))
        xml = xml.replace("?ФИО", self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.le_info_middle_name.text())
        xml = xml.replace("?ТАБЕЛЬНОМ", str(self.id_info))
        xml = xml.replace("?ДОЛЖНОСТЬ", self.cb_info_position.currentText())
        xml = xml.replace("?ТРУДОВОЙДОГДАТА", contract_number[0][1].strftime("%d.%m.%Y"))
        xml = xml.replace("?ТРУДОВОЙДОГ", str(contract_number[0][0] + "/" + contract_number[0][1].strftime("%y")))

        dir_name = self.id_info
        self.inspection_files(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            file_name = "Приказ о приеме %s.doc" % QDate.currentDate().toString("dd.MM.yyyy")
            f = open('%s/%s' % (self.path, file_name), "w", -1, "utf-8")
            f.write(xml)
            f.close()

            query = """INSERT INTO staff_worker_doc_number (Worker_Info_Id, Name, Number, Date) VALUES (%s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE Number = %s, Date = %s"""
            parametrs = (self.id_info, "приказ.прием", doc_number, QDate.currentDate().toPyDate(), doc_number, QDate.currentDate().toPyDate())
            info_sql = my_sql.sql_change(query, parametrs)
            if "mysql.connector.errors" in str(type(info_sql)):
                QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                return False

            self.statusBar().showMessage("Готово")
            self.inspection_files(dir_name, 'Путь корень рабочие')
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Приказ о прекращении трудового договора
    def build_word_order_on_dismissal(self):
        if not self.id_info:
            QMessageBox.critical(self, "Ошибка", "У этого работника нет номера", QMessageBox.Ok)
            return False

        # Проверяем нужный номер документа
        self.statusBar().showMessage("проверяю SQL")
        query = "SELECT IFNULL(MAX(CAST(Number AS SIGNED)), 'No Number') FROM staff_worker_doc_number WHERE Name = %s AND YEAR(Date) = %s"
        doc_number_sql = my_sql.sql_select(query, ("приказ.увольне", QDate.currentDate().year()))
        if "mysql.connector.errors" in str(type(doc_number_sql)):
            QMessageBox.critical(self, "Ошибка sql", doc_number_sql.msg, QMessageBox.Ok)

        if ("No Number", ) not in doc_number_sql:
            doc_number = int(doc_number_sql[0][0]) + 1
            doc_date = self.de_info_recruitment.date().toPyDate()
        else:
            doc_number = 1
            doc_date = self.de_info_recruitment.date().toPyDate()

        info = InfoDateNumber(doc_date, doc_number)
        if info.exec() == 0:
            self.statusBar().showMessage("Отмена")
            return False

        if info.de_in.date().toPyDate() != doc_date or info.le_number.text() != str(doc_number):
            doc_date = info.de_in.date().toPyDate()
            doc_number = info.le_number.text()

        # Узнаем номер договора
        query = "SELECT IFNULL(MAX(CAST(Number AS SIGNED)), 'No Number'), Date FROM staff_worker_doc_number WHERE Name = %s AND Worker_Info_Id = %s"
        contract_number = my_sql.sql_select(query, ("труд.дог.", self.id_info))
        if "mysql.connector.errors" in str(type(contract_number)):
            QMessageBox.critical(self, "Ошибка sql", contract_number.msg, QMessageBox.Ok)
            return False

        if not contract_number[0][1]:
            QMessageBox.critical(self, "№ договора", "У этого работника нет трудового договора!\nВведите его в следующем окне!", QMessageBox.Ok)
            info = InfoDateNumber(QDate.currentDate(), "Нет")
            if info.exec() == 0:
                self.statusBar().showMessage("Отмена")
                return False

            contract_number = ((info.le_number.text(), info.de_in.date().toPyDate()), )

        self.statusBar().showMessage("Открываю шаблон")
        f = open(getcwd() + '/templates/staff/order_on_dismissal.xml', "r", -1, "utf-8")
        xml = f.read()
        self.statusBar().showMessage("Закрываю шаблон")
        f.close()
        self.statusBar().showMessage("Создаю документ")

        # Если работник ИП то поменяем шапку
        if self.cb_ip.isChecked():
            xml = xml.replace("?ФИРМА", 'ИП Рублёв Александр Александрович"')
            xml = xml.replace("?ПОДПИСАНТ", 'Рублёв А.А.')
        else:
            xml = xml.replace("?ФИРМА", 'Общество с ограниченной ответственностью "Авидевелопмент-М"')
            xml = xml.replace("?ПОДПИСАНТ", 'Ширяева Н.А.')

        xml = xml.replace("?НОМЕРДОК", str(doc_number))
        xml = xml.replace("?ДАТАСОСТ", doc_date.strftime("%d.%m.%Y"))
        xml = xml.replace("?ФИО", self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.le_info_middle_name.text())
        xml = xml.replace("?ТАБЕЛЬНОМ", str(self.id_info))
        xml = xml.replace("?ДОЛЖНОСТЬ", self.cb_info_position.currentText())
        xml = xml.replace("?ТРУДОВОЙДОГДАТА", contract_number[0][1].strftime("%d.%m.%Y"))
        xml = xml.replace("?ТРУДОВОЙДОГ", str(contract_number[0][0] + "/" + contract_number[0][1].strftime("%y")))

        dir_name = self.id_info
        self.inspection_files(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            file_name = "Приказ о расторжении догвора %s.doc" % QDate.currentDate().toString("dd.MM.yyyy")
            f = open('%s/%s' % (self.path, file_name), "w", -1, "utf-8")
            f.write(xml)
            f.close()

            query = """INSERT INTO staff_worker_doc_number (Worker_Info_Id, Name, Number, Date) VALUES (%s, %s, %s, %s)
                        ON DUPLICATE KEY UPDATE Number = %s, Date = %s"""
            parametrs = (self.id_info, "приказ.увольне", doc_number, QDate.currentDate().toPyDate(), doc_number, QDate.currentDate().toPyDate())
            info_sql = my_sql.sql_change(query, parametrs)
            if "mysql.connector.errors" in str(type(info_sql)):
                QMessageBox.critical(self, "Ошибка sql", info_sql.msg, QMessageBox.Ok)
                return False

            self.statusBar().showMessage("Готово")
            self.inspection_files(dir_name, 'Путь корень рабочие')
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    # Подарочный сертификат
    def build_word_sertificat(self):
        dialog = QInputDialog.getInt(self, "Введите сумму", "Сумма сертификата", 0, 0, 10000)
        if not dialog[1]:
            return False

        self.statusBar().showMessage("Открываю шаблон")
        f = open(getcwd() + '/templates/staff/sertificat.xml', "r", -1, "utf-8")
        xml = f.read()
        self.statusBar().showMessage("Закрываю шаблон")
        f.close()
        self.statusBar().showMessage("Создаю документ")

        if self.rb_sex_m.isChecked():  # Узнаем пол работника
            xml = xml.replace("?РОД", "Получил")
        elif self.rb_sex_f.isChecked():
            xml = xml.replace("?РОД", "Получила")

        xml = xml.replace("?ФИО", self.le_info_last_name.text() + " " + self.le_info_first_name.text() + " " + self.le_info_middle_name.text())
        xml = xml.replace("?СУМ", str(dialog[0]))

        dir_name = self.id_info
        self.inspection_files(dir_name, 'Путь корень рабочие')
        if self.path:
            self.statusBar().showMessage("Сохраняю фаил")
            file_name = "Подарочный сертификат %s.doc" % QDate.currentDate().toString("dd.MM.yyyy")
            f = open('%s/%s' % (self.path, file_name), "w", -1, "utf-8")
            f.write(xml)
            f.close()
            self.statusBar().showMessage("Готово")
            self.inspection_files(dir_name, 'Путь корень рабочие')
        else:
            self.statusBar().showMessage("Ошибка сохранения")
            return False

    def closeEvent(self, e):
        if self.alert and self.access_save_sql:
            result = QMessageBox.question(self, "Выйтиb?", "Сохранить изменения перед выходом?",
                                          QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel, QMessageBox.Yes)
            if result == 16384:
                self.acc()
                e.accept()
            elif result == 65536:
                e.accept()
            elif result == 4194304:
                e.ignore()
        else:

            e.accept()


class StaffFilter(QDialog):
    def __init__(self, main):
        super(StaffFilter, self).__init__()
        loadUi(getcwd() + '/ui/staff_filter.ui', self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))

        self.main = main

        # заполняем страны
        query = "SELECT Country_name, Id FROM staff_country"
        self.country = my_sql.sql_select(query)
        if "mysql.connector.errors" in str(type(self.country)):
            raise RuntimeError("Не смог получить страны")
        self.cb_info_country.addItem("", "")
        for country in self.country:
            self.cb_info_country.addItem(country[0], country[1])

    def ui_enabled_date(self, en):
        self.de_info_birth.setEnabled(en)

    def ui_acc(self):
        where = ""

        # Блок условий имя
        if self.le_info_first_name.text() != '':
            where = self.add_filter(where, "(staff_worker_info.First_Name LIKE '%s')" % ("%" + self.le_info_first_name.text() + "%", ))

        # Блок условий фамилия
        if self.le_info_last_name.text() != '':
            where = self.add_filter(where, "(staff_worker_info.Last_Name LIKE '%s')" % ("%" + self.le_info_last_name.text() + "%", ))

        # Блок условий отчество
        if self.le_info_middle_name.text() != '':
            where = self.add_filter(where, "(staff_worker_info.Middle_Name LIKE '%s')" % ("%" + self.le_info_middle_name.text() + "%", ))

        # Блок условий табельный номер
        if self.le_info_id.text() != '':
            where = self.add_filter(where, "(staff_worker_info.Id = %s)" % self.le_info_id.text())

        # Блок условий страна
        if self.cb_info_country.currentData() != '':
            where = self.add_filter(where, "(staff_worker_info.Country_Id = %s)" % self.cb_info_country.currentData())

        # Блок  условий мужчина или женщина
        where_item = ""
        if self.cb_sex_m.isChecked():
            where_item = self.add_filter(where_item, "staff_worker_info.Sex = 'M'", False)

        if self.cb_sex_f.isChecked():
            where_item = self.add_filter(where_item, "staff_worker_info.Sex = 'F'", False)

        if where_item:
            where_item = "(" + where_item + ")"
            where = self.add_filter(where, where_item)

        # Блок  условий даты рождения
        if self.cb_date.isChecked():
            where = self.add_filter(where_item, "staff_worker_info.Date_Birth = '%s'" % self.de_info_birth.date().toString(Qt.ISODate))

        # Блок  условий даты приема
        if self.gp_date_recruitment.isChecked():
            sql_date = "(staff_worker_info.Date_Recruitment >= '%s' AND staff_worker_info.Date_Recruitment <= '%s')" % \
                       (self.de_date_recruitment_from.date().toString(Qt.ISODate), self.de_date_recruitment_to.date().toString(Qt.ISODate))
            where = self.add_filter(where, sql_date)

        # Блок  условий даты увольнения
        if self.gp_date_leave.isChecked():
            sql_date = "(staff_worker_info.`Leave` = 1 AND staff_worker_info.Date_Leave >= '%s' AND staff_worker_info.Date_Leave <= '%s')" % \
                       (self.de_date_leave_from.date().toString(Qt.ISODate), self.de_date_leave_to.date().toString(Qt.ISODate))
            where = self.add_filter(where, sql_date)

        if where:
            self.sql_query_all = self.sql_query_all.replace("ORDER BY", " WHERE " + where + " ORDER BY")

        self.main.of_set_filter(self.sql_query_all)

        self.close()

    def ui_can(self):
        self.close()
        self.destroy()

    def add_filter(self, where, add, and_add=True):
        if where:
            if and_add:
                where += " AND " + add
            else:
                where += " OR " + add
        else:
            where = add

        return where

    def of_set_sql_query(self, sql):
        self.sql_query_all = sql


class Country(list.ListItems):
    def set_settings(self):
        self.setWindowTitle("Настройка стран")  # Имя окна
        self.toolBar.setStyleSheet("background-color: rgb(129, 66, 255);")  # Цвет бара
        self.title_new_window = "Страна"  # Имя вызываемых окон

        self.sql_list = "SELECT id, staff_country.Country_name FROM staff_country ORDER BY staff_country.Country_name"
        self.sql_add = "INSERT INTO staff_country (Country_name, Patent, Act) VALUES (%s, %s, %s)"
        self.sql_change_select = "SELECT Country_name, Patent, Act FROM staff_country WHERE Id = %s"
        self.sql_update_select = 'UPDATE staff_country SET Country_name = %s, Patent = %s, Act = %s WHERE Id = %s'
        self.sql_dell = "DELETE FROM staff_country WHERE Id = %s"

        self.set_new_win = {"WinTitle": "Страна",
                            "WinColor": "(129, 66, 255)",
                            "lb_name": "Название",
                            "lb_note": "Заметка"}

    def ui_add_item(self):
        self.add_country = ChangeCountry(self)
        self.add_country.setModal(True)
        self.add_country.show()

    def ui_change_item(self, id=False):
        if id:
            id_select = id
        else:
            try:
                id_select = self.lw_list.selectedItems()[0].data(3)
            except:
                QMessageBox.critical(self, "Ошибка", "Выберете элемент", QMessageBox.Ok)
                return False

        self.change_country = ChangeCountry(self, id_select)
        self.change_country.setModal(True)
        self.change_country.show()


class ChangeCountry(QDialog):  # Ввод и изменение гражданств
    def __init__(self, main, id=None):
        super(ChangeCountry, self).__init__()
        loadUi(getcwd() + '/ui/country.ui', self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))

        self.main = main
        self.id = id
        self.change_on = False

        if self.id:
            self.set_info()
            self.change_on = True
        else:
            self.change_on = False

    def set_info(self):
        self.country_id_change = self.id
        sql_ret = my_sql.sql_select(self.main.sql_change_select, (self.id, ))
        if "mysql.connector.errors" in str(type(sql_ret)):
            QMessageBox.critical(self, "Ошибка sql", sql_ret.msg, QMessageBox.Ok)
            return False
        else:
            self.le_name.setText(sql_ret[0][0])
            if not sql_ret[0][1]:
                self.cb_patent.setChecked(False)
                self.te_act.setEnabled(True)
                self.te_act.appendPlainText(sql_ret[0][2])
            else:
                self.te_act.setEnabled(False)
                self.cb_patent.setChecked(True)

    def check_patent(self, pat):
        if pat:
            self.te_act.setEnabled(False)
        else:
            self.te_act.setEnabled(True)

    def acc(self):
        name = self.le_name.text()
        if not self.cb_patent.isChecked():
            patent = 0
            act = self.te_act.toPlainText()
        else:
            patent = 1
            act = ""

        if self.change_on:
            sql_ret = my_sql.sql_change(self.main.sql_update_select, (name, patent, act, self.id))
        else:
            sql_ret = my_sql.sql_change(self.main.sql_add, (name, patent, act))
        if "mysql.connector.errors" in str(type(sql_ret)):
            QMessageBox.critical(self, "Ошибка sql", sql_ret.msg, QMessageBox.Ok)
            return False
        self.main.sql_set_list()
        self.close()
        self.destroy()

    def cancel(self):
        result = QMessageBox.question(self, "Выйти?", "Вы уверены что хотите выйти??", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if result == 16384:
            self.close()
            self.destroy()
        elif result == 65536:
            pass


class StaffPosition(list.ListItems):
    def set_settings(self):
        self.setWindowTitle("должностей")  # Имя окна
        self.toolBar.setStyleSheet("background-color: rgb(129, 66, 255);")  # Цвет бара
        self.title_new_window = "Должность"  # Имя вызываемых окон

        self.sql_list = "SELECT Id, staff_position.Name FROM staff_position"
        self.sql_add = ""
        self.sql_change_select = ""
        self.sql_update_select = ""
        self.sql_dell = "DELETE FROM staff_position WHERE Id = %s"

    def ui_add_item(self):
        self.add_position = ChangePosition(self)
        self.add_position.setModal(True)
        self.add_position.show()

    def ui_change_item(self, id=False):
        if id:
            id_select = id
        else:
            try:
                id_select = self.lw_list.selectedItems()[0].data(3)
            except:
                QMessageBox.critical(self, "Ошибка", "Выберете элемент", QMessageBox.Ok)
                return False
        self.change_provider = ChangePosition(self, id_select)
        self.change_provider.setModal(True)
        self.change_provider.show()


class ChangePosition(QDialog):
    def __init__(self, main, id=None):
        super(ChangePosition, self).__init__()
        self.main = main
        self.id = id
        loadUi(getcwd() + '/ui/staff_position.ui', self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.change_on = False

        if id:
            self.set_info()

    def set_info(self):
        self.change_on = True
        sql_ret = my_sql.sql_select("SELECT Name, Number FROM staff_position WHERE Id = %s", (self.id,))
        if "mysql.connector.errors" in str(type(sql_ret)):
            QMessageBox.critical(self, "Ошибка sql", sql_ret.msg, QMessageBox.Ok)
            return False

        self.le_name.setText(sql_ret[0][0])
        if str(sql_ret[0][1]) != "None":
            self.le_number.setText(sql_ret[0][1])

    def acc(self):
        name = self.le_name.text()
        number = self.le_number.text()
        if self.change_on:
            par = (name, number, self.id)
            sql_ret = my_sql.sql_change("UPDATE staff_position SET Name = %s, Number = %s WHERE Id = %s", par)
        else:
            par = (name, number)
            sql_ret = my_sql.sql_change("INSERT INTO staff_position(Name, Number) VALUES (%s, %s)", par)
        if "mysql.connector.errors" in str(type(sql_ret)):
            QMessageBox.critical(self, "Ошибка sql", sql_ret.msg, QMessageBox.Ok)
            return False
        self.main.sql_set_list()
        self.close()
        self.destroy()

    def cancel(self):
        result = QMessageBox.question(self, "Выйти?", "Вы уверены что хотите выйти??", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if result == 16384:
            self.close()
            self.destroy()
        elif result == 65536:
            pass


class ExelInfo(QDialog):
    def __init__(self, birthplace):
        super(ExelInfo, self).__init__()
        loadUi(getcwd() + '/ui/exel_info.ui', self)
        to_date = QDate.currentDate()
        self.de_in.setDate(to_date)
        self.de_from.setDate(to_date)
        self.setModal(True)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.le_birthplace_country.setText(birthplace)
        self.show()


class InfoDate(QDialog):
    def __init__(self, date_in):
        super(InfoDate, self).__init__()
        loadUi(getcwd() + '/ui/exel_info_date.ui', self)
        self.de_in.setDate(date_in)
        self.setModal(True)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.show()


class InfoDateNumber(QDialog):
    def __init__(self, date_in, number):
        super(InfoDateNumber, self).__init__()
        loadUi(getcwd() + '/ui/exel_info_date_and_number.ui', self)
        self.de_in.setDate(date_in)
        self.le_number.setText(str(number))
        self.setModal(True)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.show()


class AddFile(QDialog):
    def __init__(self):
        super(AddFile, self).__init__()
        loadUi(getcwd() + '/ui/work_add_file.ui', self)
        self.setModal(True)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.show()

    def copy_file_path(self):
        self.path_copy_file.setText(QFileDialog.getOpenFileName(self, "Выберите копируемый фаил")[0])

