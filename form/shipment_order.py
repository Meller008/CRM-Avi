from os import getcwd
from PyQt5.uic import loadUi
from PyQt5.QtWidgets import QMessageBox, QTableWidgetItem, QDialog, QFileDialog
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QIcon
from form import order
from classes import print_qt
from function import my_sql
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.worksheet.pagebreak import Break


class ShipmentOrder(QDialog):
    """Это окно строить документ на отгрузку"""
    def __init__(self):
        super(QDialog, self).__init__()
        loadUi(getcwd() + '/ui/shipment_order.ui', self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))

        self.set_start_settings()

    def set_start_settings(self):
        self.tw_order.horizontalHeader().resizeSection(0, 120)
        self.tw_order.horizontalHeader().resizeSection(1, 120)
        self.tw_order.horizontalHeader().resizeSection(2, 80)
        self.tw_order.horizontalHeader().resizeSection(3, 88)
        self.tw_order.horizontalHeader().resizeSection(4, 80)

    def ui_add_order(self):
        self.order_list = order.OrderList(self, True)
        self.order_list.setWindowModality(Qt.ApplicationModal)
        self.order_list.show()

    def ui_dell_order(self):
        try:
            self.tw_order.removeRow(self.tw_order.currentRow())
        except:
            QMessageBox.critical(self, "Ошибка Удаления", "Выделите заказ который хотите удалить", QMessageBox.Ok)
            return False

    def ui_export(self):
        position = self.calc_orders()

        path = QFileDialog.getSaveFileName(self, "Сохранение", filter="Excel(*.xlsx)")
        if not path[0]:
            return False

        border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        font_11 = Font(name="Arial", size=11)
        font_8 = Font(name="Arial", size=9)

        row_ex = 1
        book = openpyxl.load_workbook(filename='%s/shipment_order.xlsx' % (getcwd() + "/templates/shipment_order"))
        sheet = book['Отчет']

        # Заполняем первую строчку
        sheet.merge_cells("A%s:H%s" % (row_ex, row_ex))
        sheet["A1"] = "Лист на отгрузку"
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet["A1"].font = font_11
        row_ex += 1
        #
        # # Заполняем список заказов которые объеденили
        # for row in range(self.tw_order.rowCount()):
        #     order_data = (self.tw_order.item(row, 0).text(), self.tw_order.item(row, 1).text(),
        #                   self.tw_order.item(row, 2).text(), self.tw_order.item(row, 3).text(),
        #                   self.tw_order.item(row, 4).text())
        #
        #     order_text = "Клиент %s %s    № заказа %s    Дата отгрузки %s    №док. %s" % order_data
        #
        #     sheet.merge_cells("A%s:H%s" % (row_ex, row_ex))
        #     sheet["A%s" % row_ex] = order_text
        #     sheet["A%s" % row_ex].border = border_all
        #     sheet["A%s" % row_ex].font = font_8
        #     # sheet["A%s" % row_ex].alignment = Alignment(horizontal="center")
        #
        #     for row in sheet.iter_rows(min_row=row_ex, min_col=1, max_col=8, max_row=row_ex):
        #         for cell in row:
        #             cell.border = border_all
        #
        #     row_ex += 1

        # Заполняем позиции
        flag_go_left_column = True
        column_name = (("A", "B", "C"), ("E", "F", "G"))
        art_col, size_col, val_col = column_name[0]
        min_col = 1
        max_col = 3
        row_ex += 1
        beginning_row_article = row_ex

        for ka, va in position.items():
            start_row_article = row_ex

            for kp, vp in va.items():
                sheet["%s%s" % (size_col, row_ex)] = kp
                sheet["%s%s" % (size_col, row_ex)].alignment = Alignment(horizontal="right", vertical="center")
                sheet["%s%s" % (val_col, row_ex)] = vp
                sheet["%s%s" % (val_col, row_ex)].alignment = Alignment(horizontal="right", vertical="center")

                row_ex += 1

            sheet.merge_cells("%s%s:%s%s" % (art_col, start_row_article, art_col, row_ex-1))
            sheet["%s%s" % (art_col, start_row_article)] = ka
            sheet["%s%s" % (art_col, start_row_article)].alignment = Alignment(horizontal="center", vertical="center")
            sheet["%s%s" % (art_col, start_row_article)].font = font_11

            for row in sheet.iter_rows(min_row=start_row_article, min_col=min_col, max_col=max_col, max_row=row_ex-1):
                for cell in row:
                        cell.border = border_all
            row_ex += 1

            if row_ex > 70 and flag_go_left_column:
                art_col, size_col, val_col = column_name[1]
                min_col = 5
                max_col = 7
                row_ex = beginning_row_article
                flag_go_left_column = False




        book.save(path[0])

    def calc_orders(self):

        all_position = {}
        for row in range(self.tw_order.rowCount()):
            order_position = self.calc_order_position(self.tw_order.item(row, 0).data(-1))
            if order_position:
                for ka, va in order_position.items():
                    article = all_position.setdefault(ka, {})
                    for kp, vp in va.items():
                        value = article.setdefault(kp, 0)
                        article[kp] = value + vp

        return all_position

    def calc_order_position(self, order_id):
        """Собирает все артикула заказа в словарь вида
        {
            Артикул:{
                        Размер: кол-во
                        Размер: кол-во
                        ...
                    }
            ...
        }"""

        query = """SELECT product_article.Article, product_article_size.Size, order_position.Value
                      FROM order_position LEFT JOIN product_article_parametrs ON order_position.Product_Article_Parametr_Id = product_article_parametrs.Id
                        LEFT JOIN product_article_size ON product_article_parametrs.Product_Article_Size_Id = product_article_size.Id
                        LEFT JOIN product_article ON product_article_size.Article_Id = product_article.Id
                      WHERE Order_Id = %s"""
        sql_info = my_sql.sql_select(query, (order_id,))
        if "mysql.connector.errors" in str(type(sql_info)):
            QMessageBox.critical(self, "Ошибка sql получения позиций заказа", sql_info.msg, QMessageBox.Ok)
            return False

        all_position = {}
        for position_sql in sql_info:
            article = all_position.setdefault(position_sql[0], {})
            value = article.setdefault(position_sql[1], 0)
            article[position_sql[1]] = value + position_sql[2]

        return all_position

    def of_tree_select_order(self, order):
        query = """SELECT `order`.Id, clients.Name, clients_actual_address.Name, `order`.Number_Order, `order`.Date_Shipment, `order`.Number_Doc
                      FROM `order` LEFT JOIN clients ON `order`.Client_Id = clients.Id
                        LEFT JOIN clients_actual_address ON `order`.Clients_Adress_Id = clients_actual_address.Id
                      WHERE `order`.Id = %s"""
        sql_info = my_sql.sql_select(query, (order[1],))
        if "mysql.connector.errors" in str(type(sql_info)):
            QMessageBox.critical(self, "Ошибка sql получения информации о заказе", sql_info.msg, QMessageBox.Ok)
            return False

        if not sql_info:
            return False

        order = sql_info[0]

        self.tw_order.insertRow(self.tw_order.rowCount())
        table_item = QTableWidgetItem(str(order[1]))
        table_item.setData(-1, order[0])
        self.tw_order.setItem(self.tw_order.rowCount()-1, 0, table_item)

        table_item = QTableWidgetItem(str(order[2]))
        table_item.setData(-1, order[0])
        self.tw_order.setItem(self.tw_order.rowCount()-1, 1, table_item)

        table_item = QTableWidgetItem(str(order[3]))
        table_item.setData(-1, order[0])
        self.tw_order.setItem(self.tw_order.rowCount()-1, 2, table_item)

        table_item = QTableWidgetItem(order[4].strftime("%d.%m.%Y"))
        table_item.setData(-1, order[0])
        self.tw_order.setItem(self.tw_order.rowCount()-1, 3, table_item)

        table_item = QTableWidgetItem(str(order[5]))
        table_item.setData(-1, order[0])
        self.tw_order.setItem(self.tw_order.rowCount()-1, 4, table_item)