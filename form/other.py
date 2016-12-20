from os import getcwd
from PyQt5.uic import loadUiType
from PyQt5.QtWidgets import QFileDialog, QMessageBox, QMainWindow
from PyQt5.QtGui import QIcon
import json
import openpyxl
from openpyxl.styles import Border, Side, Alignment

order_edi_class = loadUiType(getcwd() + '/ui/input_order_edi.ui')[0]


class OrderEDI(QMainWindow, order_edi_class):
    def __init__(self):
        super(OrderEDI, self).__init__()
        self.setupUi(self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.start_settings()

    def start_settings(self):
        try:
            file = open(getcwd() + '/templates/edi/auchan_edi.json')
        except IOError as e:
            print(e)
            self.le_base_auchan.setText('<html><head/><body><p align="center"><span style=" font-size:18pt; font-weight:600; color:#ff0000;">Базы НЕТ</span></p></body></html>')
            self.pb_order.setEnabled(False)
            return False
        else:
            with file:
                if file.read() == "":
                    self.le_base_auchan.setText('<html><head/><body><p align="center"><span style=" font-size:18pt; font-weight:600; color:#ff0000;">База ПУСТА</span></p></body></html>')
                    self.pb_order.setEnabled(False)
                    return False
                else:
                    self.le_base_auchan.setText('<html><head/><body><p align="center"><span style=" font-size:18pt; font-weight:600; color:#57ff57;">База ЕСТЬ</span></p></body></html>')
                    self.pb_order.setEnabled(True)

    def add_base(self):
        base = {}
        file_dir = QFileDialog.getOpenFileName(self, "Выберите базу", "C:/", "*.xls *.xlsx")
        if file_dir[1] == "":
            return False

        book = openpyxl.load_workbook(filename=file_dir[0])
        sheet = book["Лист1"]
        row_count = sheet.max_row
        for row in range(2, row_count+1):
            article = sheet["A%s" % row].value
            size = sheet["B%s" % row].value
            kod = sheet["E%s" % row].value
            position = sheet["G%s" % row].value
            base_item = {kod: {"article": article, "size": size, "position": position}}
            base.update(base_item)
        with open(getcwd() + '/templates/edi/auchan_edi.json', 'w', encoding='utf-8') as fh:
            fh.write(json.dumps(base, ensure_ascii=False, indent=4))
        self.start_settings()

    def add_order(self):
        order = []
        file_dir = QFileDialog.getOpenFileName(self, "Выберите заказ", "C:/", "*.xls *.xlsx")
        if file_dir[1] == "":
            return False

        book = openpyxl.load_workbook(filename=file_dir[0])
        sheet = book["Sheet1"]

        order_number = sheet["C5"].value
        order_date = sheet["C6"].value
        delivery_date = sheet["C7"].value
        delivery_address = sheet["J6"].value

        row = 18
        kod_column = self.le_kod.text()
        value_column = self.le_value.text()
        while sheet["B%s" % row].value is not None:
            order_item = {"kod": sheet["%s%s" % (kod_column, row)].value, "value": sheet["%s%s" % (value_column, row)].value}
            order.append(order_item)
            row += 1

        with open(getcwd() + '/templates/edi/auchan_edi.json', 'r', encoding='utf-8') as fh:
            base = json.load(fh)

        new_order = {}
        error_kod = []

        for one_order in order:
            if base.get(one_order["kod"]) is not None:
                new_order_position = {"value": one_order["value"], "size": base.get(one_order["kod"])["size"], "position": base.get(one_order["kod"])["position"]}
                article = base.get(one_order["kod"])["article"]

                if new_order.get(article) is None:
                    new_order.setdefault(article, [])
                new_order[article].append(new_order_position)
            else:
                error_kod.append(str(one_order["kod"]))

        if error_kod:
            str_err = "\n".join(error_kod)
            QMessageBox.critical(self, "Не найдены коды", "В базе не нашлись следующие штрих коды: \n %s" % str_err, QMessageBox.Ok)

        book = openpyxl.load_workbook(filename=getcwd() + '/templates/edi/templates_auchan.xlsx')
        sheet = book["Лист1"]
        sheet["A1"] = "Заказ №%s %s/%s" % (order_number, order_date, delivery_date)
        sheet["A2"] = delivery_address

        row = 4
        a_char = "A"
        s_char = "B"
        v_char = "C"

        border = Border(left=Side(border_style='thin', color='FF000000'),
                         right=Side(border_style='thin', color='FF000000'),
                         top=Side(border_style='thin', color='FF000000'),
                         bottom=Side(border_style='thin', color='FF000000'),
                        )

        position = 0
        del_articcle = []
        while new_order and position < 100:
            for new_order_item in new_order.items():
                if new_order_item[1][0]["position"] == position:
                    article = new_order_item[0]

                    row_start = row
                    for a_size in new_order_item[1]:
                        sheet["%s%s" % (a_char, row)] = str(article)
                        sheet["%s%s" % (a_char, row)].border = border
                        sheet["%s%s" % (a_char, row)].alignment = Alignment(horizontal="center", vertical="center")
                        sheet["%s%s" % (s_char, row)] = str(a_size["size"])
                        sheet["%s%s" % (s_char, row)].border = border
                        sheet["%s%s" % (s_char, row)].alignment = Alignment(horizontal="center", vertical="center")
                        sheet["%s%s" % (v_char, row)] = str(a_size["value"])
                        sheet["%s%s" % (v_char, row)].border = border
                        sheet["%s%s" % (v_char, row)].alignment = Alignment(horizontal="center", vertical="center")
                        row += 1

                    sheet.merge_cells("%s%s:%s%s" % (a_char, row_start, a_char, row-1))
                    row += 1
                    del_articcle.append(article)
            for del_a in del_articcle:
                new_order.pop(del_a)
            del_articcle = []
            position += 1
        file_dir = QFileDialog.getExistingDirectory(self, "Куда сохранить заказ", "C:/")
        name = "Обработаный заказ Ашан %s.xlsx" % order_number
        book.save('%s/%s' % (file_dir, name))


