from os import getcwd
from PyQt5.uic import loadUi
from PyQt5.QtWidgets import QDialog, QMessageBox, QTableWidgetItem, QMainWindow, QFileDialog, QPushButton
from PyQt5.QtCore import Qt, QDate, QDateTime, QObject
from PyQt5.QtGui import QIcon, QBrush, QColor
import re, requests, xml.etree.cElementTree as xml_pars
from decimal import Decimal
import datetime
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.worksheet.pagebreak import Break
from copy import copy
from function import my_sql, to_excel, table_to_html, moneyfmt, calc_price, str_to
from classes import print_qt
from form.templates import table, list
from form import clients, article, print_label
import num2t4ru
from classes.my_class import User

import collections


class OrderList(table.TableList):
    def set_settings(self):

        self.setWindowTitle("Заказы")  # Имя окна
        self.resize(900, 270)
        self.pb_copy.deleteLater()
        self.pb_other.deleteLater()
        self.toolBar.setStyleSheet("background-color: rgb(126, 176, 127);")  # Цвет бара

        # Названия колонк (Имя, Длинна)
        self.table_header_name = (("Клиент", 120), ("№ закза", 70), ("Пункт разгрузки", 100), ("Дата поствки", 70), ("№ док.", 50), ("Позиций", 50),
                                  ("Стоимость", 105), ("Стоимость без ндс", 105), ("Примечание", 150), ("Отгр.", 40))

        self.filter = None
        self.query_table_all = """SELECT `order`.Id, clients.Name, `order`.Number_Order, clients_actual_address.Name,
                                      `order`.Date_Shipment, `order`.Number_Doc, COUNT(order_position.Id), ROUND(`order`.Sum_In_Nds, 2), ROUND(`order`.Sum_Off_Nds, 2),
                                      `order`.Note, IF(`order`.Shipped = 0, 'Нет', 'Да'), clients.No_Nds
                                    FROM `order` LEFT JOIN clients ON `order`.Client_Id = clients.Id
                                      LEFT JOIN clients_actual_address ON `order`.Clients_Adress_Id = clients_actual_address.Id
                                      LEFT JOIN order_position ON `order`.Id = order_position.Order_Id GROUP BY `order`.Id
                                    ORDER BY `order`.Date_Order DESC, `order`.Number_Doc DESC """

        #  нулевой элемент должен быть ID
        self.query_table_select = """SELECT `order`.Id, clients.Name, `order`.Number_Order, clients_actual_address.Name,
                                          `order`.Date_Shipment, `order`.Number_Doc, COUNT(order_position.Id), ROUND(`order`.Sum_In_Nds, 2), ROUND(`order`.Sum_Off_Nds, 2),
                                          `order`.Note, IF(`order`.Shipped = 0, 'Нет', 'Да'), clients.No_Nds
                                      FROM `order` LEFT JOIN clients ON `order`.Client_Id = clients.Id
                                          LEFT JOIN clients_actual_address ON `order`.Clients_Adress_Id = clients_actual_address.Id
                                          LEFT JOIN order_position ON `order`.Id = order_position.Order_Id GROUP BY `order`.Id
                                      ORDER BY `order`.Date_Order DESC, `order`.Number_Doc DESC """

        self.query_table_dell = "DELETE FROM `order` WHERE Id = %s"

    def ui_add_table_item(self):  # Добавить предмет
        id = False
        self.new_order = Order(self, id)
        self.new_order.setWindowModality(Qt.ApplicationModal)
        self.new_order.show()

    def ui_change_table_item(self, id=False):  # изменить элемент
        if id:
            item_id = id
        else:
            try:
                item_id = self.table_widget.selectedItems()[0].data(5)
            except:
                QMessageBox.critical(self, "Ошибка ", "Выделите элемент который хотите изменить", QMessageBox.Ok)
                return False

        self.order = Order(self, item_id)
        self.order.setWindowModality(Qt.ApplicationModal)
        self.order.show()

    def ui_double_click_table_item(self, item):  # Двойной клик по элементу
        if not self.dc_select:
            self.ui_change_table_item(item.data(5))
        else:
            # что хотим получить ставим всместо 0
            item = (self.table_widget.item(item.row(), 4).text(), item.data(5))
            self.main.of_tree_select_order(item)
            self.close()
            self.destroy()

    def set_table_info(self):
        self.table_items = my_sql.sql_select(self.query_table_select)
        if "mysql.connector.errors" in str(type(self.table_items)):
            QMessageBox.critical(self, "Ошибка sql получение таблицы", self.table_items.msg, QMessageBox.Ok)
            return False

        self.table_widget.clearContents()
        self.table_widget.setRowCount(0)

        if not self.table_items:
            return False

        for table_typle in self.table_items:
            self.table_widget.insertRow(self.table_widget.rowCount())
            if table_typle[10] == "Да":
                color = QBrush(QColor(62, 240, 130, 255))
            else:
                color = QBrush(QColor(228, 242, 99, 255))

            for column in range(1, len(table_typle)):

                if isinstance(table_typle[column], Decimal):
                    text = re.sub(r'(?<=\d)(?=(\d\d\d)+\b.)', ' ', str(table_typle[column]))
                    item = table.QTableWidgetItemFloat(text)

                elif isinstance(table_typle[column], datetime.date):
                    date = QDate(table_typle[column].year, table_typle[column].month, table_typle[column].day)
                    item = QTableWidgetItem()
                    item.setData(Qt.DisplayRole, date)

                else:
                    item = QTableWidgetItem()
                    item.setData(Qt.DisplayRole, table_typle[column])

                item.setData(5, table_typle[0])
                item.setBackground(color)
                self.table_widget.setItem(self.table_widget.rowCount() - 1, column - 1, item)

    def ui_filter(self):
        if self.filter is None:
            self.filter = OrderFilter(self)
        self.filter.of_set_sql_query(self.query_table_all)
        self.filter.setWindowModality(Qt.ApplicationModal)
        self.filter.show()

    def of_set_filter(self, sql):
        self.query_table_select = sql

        self.ui_update()


class Order(QMainWindow):
    def __init__(self, main_class=0, id=False):
        super(Order, self).__init__()
        loadUi(getcwd() + '/ui/order.ui', self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.id = id
        self.main = main_class
        self.access_save_sql = True

        self.save_change_order = False
        self.save_change_order_position = False

        self.start_settings()

        self.save_change_order = False
        self.save_change_order_position = False

        self.access()

    def access(self):
        for item in User().access_list(self.__class__.__name__):
            a = getattr(self, item["atr1"])
            if item["atr2"]:
                a = getattr(a, item["atr2"])

            if item["value"]:
                if item["value"] == "True":
                    val = True
                elif item["value"] == "False":
                    val = False
                else:
                    val = item["value"]
                a(val)
            else:
                a()

    def access_save(self, bool):
        self.access_save_sql = bool

    def start_settings(self):
        self.tw_position.horizontalHeader().resizeSection(0, 70)
        self.tw_position.horizontalHeader().resizeSection(1, 60)
        self.tw_position.horizontalHeader().resizeSection(2, 120)
        self.tw_position.horizontalHeader().resizeSection(3, 220)
        self.tw_position.horizontalHeader().resizeSection(4, 60)
        self.tw_position.horizontalHeader().resizeSection(5, 50)
        self.tw_position.horizontalHeader().resizeSection(6, 70)
        self.tw_position.horizontalHeader().resizeSection(7, 70)

        self.tw_position_label.horizontalHeader().resizeSection(0, 70)
        self.tw_position_label.horizontalHeader().resizeSection(1, 60)
        self.tw_position_label.horizontalHeader().resizeSection(2, 120)
        self.tw_position_label.horizontalHeader().resizeSection(3, 220)
        self.tw_position_label.horizontalHeader().resizeSection(4, 60)
        self.tw_position_label.horizontalHeader().resizeSection(5, 50)
        self.tw_position_label.horizontalHeader().resizeSection(6, 70)
        self.tw_position_label.horizontalHeader().resizeSection(7, 70)

        if self.id:
            self.start_set_sql_info()
        else:
            self.pb_add_position.setEnabled(False)
            self.pb_change_position.setEnabled(False)
            self.pb_dell_position.setEnabled(False)
            self.pb_check_warehouse.setEnabled(False)
            self.pb_check_warehouse.setEnabled(False)
            self.pb_edi.setEnabled(False)

            self.sql_shipped = False

            self.sql_date_shipment = None
            self.sql_number_doc = None

            self.de_date_order.setDate(QDate.currentDate())
            self.de_date_shipment.setDate(QDate.currentDate())

            self.ui_change_date_shipment()

    def start_set_sql_info(self):
        query = """SELECT `order`.Client_Id, clients.Name, `order`.Clients_Vendor_Id, `order`.Clients_Adress_Id, order_transport_company.Id,
                    order_transport_company.Name, `order`.Date_Order, `order`.Date_Shipment, `order`.Number_Order, `order`.Number_Doc, `order`.Note, `order`.Shipped,
                    `order`.Combined
                    FROM `order` LEFT JOIN order_transport_company ON `order`.Transport_Company_Id = order_transport_company.Id
                    LEFT JOIN clients ON `order`.Client_Id = clients.Id WHERE `order`.Id = %s"""
        sql_info = my_sql.sql_select(query, (self.id,))
        if "mysql.connector.errors" in str(type(sql_info)):
            QMessageBox.critical(self, "Ошибка sql получения информации о заказе", sql_info.msg, QMessageBox.Ok)
            return False
        self.of_list_clients([sql_info[0][0], sql_info[0][1]])

        find_index = self.cb_clients_vendor.findData(sql_info[0][2])
        if find_index >= 0:
            self.cb_clients_vendor.setCurrentIndex(find_index)

        find_index = self.cb_clients_adress.findData(sql_info[0][3])
        if find_index >= 0:
            self.cb_clients_adress.setCurrentIndex(find_index)

        self.le_transport_company.setWhatsThis(str(sql_info[0][4]))
        self.le_transport_company.setText(sql_info[0][5])

        self.de_date_order.setDate(sql_info[0][6])
        self.de_date_shipment.setDate(sql_info[0][7])

        # Запомним дату и номер для выдачи заказов
        self.sql_date_shipment = sql_info[0][7]
        self.sql_number_doc = str(sql_info[0][9])

        self.le_number_order.setText(sql_info[0][8])
        self.le_number_doc.setText(str(sql_info[0][9]))
        self.le_note.setText(sql_info[0][10])

        if sql_info[0][11] != 0:
            self.cb_shipping.setChecked(True)
            self.cb_shipping.setEnabled(True)
            self.pb_add_position.setEnabled(False)
            self.pb_change_position.setEnabled(False)
            self.pb_dell_position.setEnabled(False)
            self.sql_shipped = True
        else:
            self.cb_shipping.setChecked(False)
            self.sql_shipped = False

        if sql_info[0][12] == 1:
            self.cb_combined.setChecked(True)
        else:
            self.cb_combined.setChecked(False)

        query = """SELECT order_position.Id, product_article.Article, product_article_size.Size, product_article_parametrs.Id, product_article_parametrs.Name,
                    product_article_parametrs.Client_Name, order_position.Price, order_position.NDS, order_position.Value, order_position.In_On_Place,
                    order_position.Price * order_position.Value, product_article_parametrs.Client_code
                    FROM order_position LEFT JOIN product_article_parametrs ON order_position.Product_Article_Parametr_Id = product_article_parametrs.Id
                        LEFT JOIN product_article_size ON product_article_parametrs.Product_Article_Size_Id = product_article_size.Id
                        LEFT JOIN product_article ON product_article_size.Article_Id = product_article.Id
                    WHERE Order_Id = %s ORDER BY order_position.Id"""
        sql_info = my_sql.sql_select(query, (self.id,))
        if "mysql.connector.errors" in str(type(sql_info)):
            QMessageBox.critical(self, "Ошибка sql получения позиций заказа", sql_info.msg, QMessageBox.Ok)
            return False

        for position in sql_info:
            row = self.tw_position.rowCount()
            self.tw_position.insertRow(row)
            self.tw_position_label.insertRow(row)
            table_item = QTableWidgetItem(position[1])
            table_item.setData(-1, "set")
            table_item.setData(-2, position[0])
            self.tw_position.setItem(row, 0, table_item)
            table_item = QTableWidgetItem(position[1])
            table_item.setData(-1, "set")
            table_item.setData(-2, position[0])
            self.tw_position_label.setItem(row, 0, table_item)

            table_item = QTableWidgetItem(position[2])
            table_item.setData(-1, "set")
            table_item.setData(-2, position[0])
            self.tw_position.setItem(row, 1, table_item)
            table_item = QTableWidgetItem(position[2])
            table_item.setData(-1, "set")
            table_item.setData(-2, position[0])
            self.tw_position_label.setItem(row, 1, table_item)

            table_item = QTableWidgetItem(position[4])
            table_item.setData(-1, "set")
            table_item.setData(-2, position[0])
            table_item.setData(5, position[3])
            self.tw_position.setItem(row, 2, table_item)
            table_item = QTableWidgetItem(position[4])
            table_item.setData(-1, "set")
            table_item.setData(-2, position[0])
            table_item.setData(5, position[3])
            self.tw_position_label.setItem(row, 2, table_item)

            table_item = QTableWidgetItem(str(position[5]))
            table_item.setData(-1, "set")
            table_item.setData(-2, position[0])
            self.tw_position.setItem(row, 3, table_item)
            table_item = QTableWidgetItem(str(position[5]))
            table_item.setData(-1, "set")
            table_item.setData(-2, position[0])
            self.tw_position_label.setItem(row, 3, table_item)

            table_item = QTableWidgetItem(str(position[6]))
            table_item.setData(-1, "set")
            table_item.setData(-2, position[0])
            table_item.setData(5, position[7])
            self.tw_position.setItem(row, 4, table_item)
            table_item = QTableWidgetItem(str(position[6]))
            table_item.setData(-1, "set")
            table_item.setData(-2, position[0])
            table_item.setData(5, position[7])
            self.tw_position_label.setItem(row, 4, table_item)

            table_item = QTableWidgetItem(str(position[8]))
            table_item.setData(-1, "set")
            table_item.setData(-2, position[0])
            table_item.setData(5, position[9])
            self.tw_position.setItem(row, 5, table_item)
            table_item = QTableWidgetItem(str(position[8]))
            table_item.setData(-1, "set")
            table_item.setData(-2, position[0])
            table_item.setData(5, position[9])
            self.tw_position_label.setItem(row, 5, table_item)

            table_item = QTableWidgetItem(str(position[10]))
            table_item.setData(-1, "set")
            table_item.setData(-2, position[0])
            self.tw_position.setItem(row, 6, table_item)
            table_item = QTableWidgetItem(str(position[10]))
            table_item.setData(-1, "set")
            table_item.setData(-2, position[0])
            self.tw_position_label.setItem(row, 6, table_item)

            table_item = QTableWidgetItem(str(position[11]))
            table_item.setData(-1, "set")
            table_item.setData(-2, position[0])
            self.tw_position.setItem(row, 7, table_item)

        self.calc_sum()

    def ui_view_client(self):
        self.client_list = clients.ClientList(self, True)
        self.client_list.setWindowModality(Qt.ApplicationModal)
        self.client_list.show()

    def ui_view_transport_company(self):
        self.transport_company = TransportCompanyName(self, True)
        self.transport_company.setWindowModality(Qt.ApplicationModal)
        self.transport_company.show()

    def ui_add_position(self):
        self.position = Position()
        self.position.setModal(True)
        self.position.show()

        if not self.position.exec_():
            return False

        if self.position.le_parametr.whatsThis() == "":
            return False

        if not self.sql_shipped:
            self.cb_shipping.setChecked(self.sql_shipped)
            self.cb_shipping.setEnabled(self.sql_shipped)

        row = self.tw_position.rowCount()
        self.tw_position.insertRow(row)
        table_item = QTableWidgetItem(self.position.le_article.text())
        table_item.setData(-1, "new")
        self.tw_position.setItem(row, 0, table_item)

        table_item = QTableWidgetItem(self.position.le_size.text())
        table_item.setData(-1, "new")
        self.tw_position.setItem(row, 1, table_item)

        table_item = QTableWidgetItem(self.position.le_parametr.text())
        table_item.setData(-1, "new")
        table_item.setData(5, self.position.le_parametr.whatsThis())
        self.tw_position.setItem(row, 2, table_item)

        # вставка кода товара
        table_item = QTableWidgetItem(self.position.le_client_cod.text())
        table_item.setData(-1, "new")
        table_item.setData(5, self.position.le_parametr.whatsThis())
        self.tw_position.setItem(row, 7, table_item)

        query = "SELECT Client_Name FROM product_article_parametrs WHERE Id = %s"
        sql_info = my_sql.sql_select(query, (self.position.le_parametr.whatsThis(),))
        if "mysql.connector.errors" in str(type(sql_info)):
            QMessageBox.critical(self, "Ошибка sql получения имени артикула", sql_info.msg, QMessageBox.Ok)
            return False
        table_item = QTableWidgetItem(sql_info[0][0])
        table_item.setData(-1, "new")
        self.tw_position.setItem(row, 3, table_item)

        if self.position.nds_1.isChecked():
            nds = 18
        else:
            nds = 10
        if self.lb_client.whatsThis().find("no_nds") >= 0:
            table_item = QTableWidgetItem(self.position.le_price_no_nds.text())
            table_item.setData(-1, "new")
            table_item.setData(5, nds)
            self.tw_position.setItem(row, 4, table_item)

            table_item = QTableWidgetItem(self.position.le_sum_no_nds.text())
            table_item.setData(-1, "new")
            table_item.setData(5, nds)
            self.tw_position.setItem(row, 6, table_item)
        else:
            table_item = QTableWidgetItem(self.position.le_price.text())
            table_item.setData(-1, "new")
            table_item.setData(5, nds)
            self.tw_position.setItem(row, 4, table_item)

            table_item = QTableWidgetItem(self.position.le_sum_rub.text())
            table_item.setData(-1, "new")
            table_item.setData(5, nds)
            self.tw_position.setItem(row, 6, table_item)

        table_item = QTableWidgetItem(self.position.le_value.text())
        table_item.setData(-1, "new")
        table_item.setData(5, self.position.le_in_on_place.text())
        self.tw_position.setItem(row, 5, table_item)
        self.save_change_order_position = True
        try:
            self.main.pb_doc.deleteLater()
        except:
            pass

        self.tw_position.setCurrentItem(table_item)
        self.calc_sum()
        return True

    def ui_change_position(self, row_in=False):
        if row_in:
            select_row = row_in
        else:
            try:
                select_row = self.tw_position.currentRow()
            except:
                QMessageBox.critical(self, "Ошибка ", "Выделите элемент который хотите изменить", QMessageBox.Ok)
                return False

            if select_row == -1:
                return False

        self.position = Position()
        self.position.setModal(True)
        self.position.le_article.setText(self.tw_position.item(select_row, 0).text())
        self.position.le_size.setText(self.tw_position.item(select_row, 1).text())
        self.position.le_parametr.setText(self.tw_position.item(select_row, 2).text())
        self.position.le_parametr.setWhatsThis(str(self.tw_position.item(select_row, 2).data(5)))
        self.position.le_client_cod.setText(self.tw_position.item(select_row, 7).text())

        nds = self.tw_position.item(select_row, 4).data(5)
        if nds == 18:
            self.position.nds_1.setChecked(True)
        else:
            self.position.nds_2.setChecked(True)

        if self.lb_client.whatsThis().find("no_nds") >= 0:
            self.position.le_price_no_nds.setText(self.tw_position.item(select_row, 4).text())
            self.position.price_name_change = "no nds"
        else:
            self.position.le_price.setText(self.tw_position.item(select_row, 4).text())
            self.position.price_name_change = "nds"

        self.position.le_value.setText(self.tw_position.item(select_row, 5).text())
        self.position.le_in_on_place.setText(str(self.tw_position.item(select_row, 5).data(5)))
        self.position.ui_calculation()
        self.position.show()
        if not self.position.exec_():
            return False

        if not self.sql_shipped:
            self.cb_shipping.setChecked(self.sql_shipped)
            self.cb_shipping.setEnabled(self.sql_shipped)

        row = select_row
        if self.tw_position.item(row, 0).data(-1) == "new":
            status = "new"
            id = False
        else:
            status = "upd"
            id = self.tw_position.item(row, 0).data(-2)

        table_item = QTableWidgetItem(self.position.le_article.text())
        table_item.setData(-1, status)
        if id:
            table_item.setData(-2, id)
        self.tw_position.setItem(row, 0, table_item)

        table_item = QTableWidgetItem(self.position.le_size.text())
        table_item.setData(-1, status)
        if id:
            table_item.setData(-2, id)
        self.tw_position.setItem(row, 1, table_item)

        table_item = QTableWidgetItem(self.position.le_parametr.text())
        table_item.setData(-1, status)
        if id:
            table_item.setData(-2, id)
        table_item.setData(5, self.position.le_parametr.whatsThis())
        self.tw_position.setItem(row, 2, table_item)

        # вставка кода товара
        table_item = QTableWidgetItem(self.position.le_client_cod.text())
        table_item.setData(-1, "new")
        table_item.setData(5, self.position.le_parametr.whatsThis())
        self.tw_position.setItem(row, 7, table_item)

        query = "SELECT Client_Name FROM product_article_parametrs WHERE Id = %s"
        sql_info = my_sql.sql_select(query, (self.position.le_parametr.whatsThis(),))
        if "mysql.connector.errors" in str(type(sql_info)):
            QMessageBox.critical(self, "Ошибка sql получения имени артикула", sql_info.msg, QMessageBox.Ok)
            return False
        table_item = QTableWidgetItem(sql_info[0][0])
        table_item.setData(-1, status)
        if id:
            table_item.setData(-2, id)
        self.tw_position.setItem(row, 3, table_item)

        if self.position.nds_1.isChecked():
            nds = 18
        else:
            nds = 10
        if self.lb_client.whatsThis().find("no_nds") >= 0:
            table_item = QTableWidgetItem(self.position.le_price_no_nds.text())
            table_item.setData(-1, status)
            if id:
                table_item.setData(-2, id)
            table_item.setData(5, nds)
            self.tw_position.setItem(row, 4, table_item)

            table_item = QTableWidgetItem(self.position.le_sum_no_nds.text())
            table_item.setData(-1, status)
            if id:
                table_item.setData(-2, id)
            table_item.setData(5, nds)
            self.tw_position.setItem(row, 6, table_item)
        else:
            table_item = QTableWidgetItem(self.position.le_price.text())
            table_item.setData(-1, status)
            if id:
                table_item.setData(-2, id)
            table_item.setData(5, nds)
            self.tw_position.setItem(row, 4, table_item)

            table_item = QTableWidgetItem(self.position.le_sum_rub.text())
            table_item.setData(-1, status)
            if id:
                table_item.setData(-2, id)
            table_item.setData(5, nds)
            self.tw_position.setItem(row, 6, table_item)

        table_item = QTableWidgetItem(self.position.le_value.text())
        table_item.setData(-1, status)
        if id:
            table_item.setData(-2, id)
        table_item.setData(5, self.position.le_in_on_place.text())
        self.tw_position.setItem(row, 5, table_item)
        self.save_change_order_position = True
        try:
            self.main.pb_doc.deleteLater()
        except:
            pass

        self.calc_sum()
        return True

    def ui_double_click_position(self, row):
        if not self.sql_shipped:
            self.ui_change_position(row)

    def ui_del_position(self):
        try:
            row = self.tw_position.currentRow()
        except:
            QMessageBox.information(self, "Ошибка", "Выберете позицию для удаления", QMessageBox.Ok)
            return False

        if row == -1:
            return False

        result = QMessageBox.question(self, "Удалить?", "Точно удалить позицию?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if result == 16384:
            if not self.sql_shipped:
                self.cb_shipping.setChecked(self.sql_shipped)
                self.cb_shipping.setEnabled(self.sql_shipped)

            self.tw_position.setRowHidden(row, True)
            for col in range(4):
                self.tw_position.item(row, col).setData(-1, "del")
        self.save_change_order_position = True

        self.calc_sum()
        self.cb_shipping.setEnabled(False)
        return True

    def ui_change_date_shipment(self):

        # новый заказ
        if self.sql_date_shipment is None:
            query = "SELECT IFNULL(MAX(Number_Doc + 1), 'No Number') FROM `order` WHERE YEAR(Date_Shipment) = %s"
            sql_info = my_sql.sql_select(query, (self.de_date_shipment.date().year(),))
            if "mysql.connector.errors" in str(type(sql_info)):
                QMessageBox.critical(self, "Ошибка sql получения нового номера документа", sql_info.msg, QMessageBox.Ok)
                return False
            if sql_info[0][0] == "No Number":
                self.le_number_doc.setText("1")
            else:
                self.le_number_doc.setText(str(sql_info[0][0]))
        else:
            # Старый заказ и год отгрузки такой же как и в БД
            if self.sql_date_shipment.year == self.de_date_shipment.date().year():
                self.le_number_doc.setText(self.sql_number_doc)

            else:  # Заказ старый но новый год в дате
                query = "SELECT IFNULL(MAX(Number_Doc + 1), 'No Number') FROM `order` WHERE YEAR(Date_Shipment) = %s"
                sql_info = my_sql.sql_select(query, (self.de_date_shipment.date().year(),))
                if "mysql.connector.errors" in str(type(sql_info)):
                    QMessageBox.critical(self, "Ошибка sql получения нового номера документа", sql_info.msg, QMessageBox.Ok)
                    return False
                if sql_info[0][0] == "No Number":
                    self.le_number_doc.setText("1")
                else:
                    self.le_number_doc.setText(str(sql_info[0][0]))

        self.ui_order_info_edit()

    def ui_order_info_edit(self):
        if not self.save_change_order:
            self.save_change_order = True

    def ui_check_warehouse(self):
        position_article_id = {}

        # Переберем таблицу и получим нужные ID для проверки слада
        for row in range(self.tw_position.rowCount()):
            table_item = self.tw_position.item(row, 2)

            self.tw_position.item(row, 1).setBackground(QBrush(QColor(252, 141, 141, 255)))

            if table_item.data(-1) == "del":
               pass
            else:
                if position_article_id.get(int(table_item.data(5))) is None:
                    position_article_id.update({int(table_item.data(5)): int(self.tw_position.item(row, 5).text())})
                else:
                    position_article_id[int(table_item.data(5))] += int(self.tw_position.item(row, 5).text())

        # Получим остатки склада
        query = "SELECT Id_Article_Parametr, Value_In_Warehouse FROM product_article_warehouse WHERE Id_Article_Parametr IN %s" % str(
            tuple(position_article_id.keys())).replace(",)", ")")
        sql_info = my_sql.sql_select(query)
        if "mysql.connector.errors" in str(type(sql_info)):
            QMessageBox.critical(self, "Ошибка sql получение остатков склада", sql_info.msg, QMessageBox.Ok)
            return False

        if len(position_article_id) != len(sql_info):
            QMessageBox.critical(self, "Ошибка проверки", "Не равны списки заказа и баланса склада", QMessageBox.Ok)
            return False

        # переберем таблицу и сравним остатки
        color_yes = QBrush(QColor(150, 255, 161, 255))
        color_no = QBrush(QColor(252, 141, 141, 255))
        error = False

        for row in range(self.tw_position.rowCount()):
            table_item = self.tw_position.item(row, 2)
            if table_item.data(-1) != "del":

                warehouse_sql = [warehouse for warehouse in sql_info if warehouse[0] == int(table_item.data(5))][0]

                value = warehouse_sql[1] - position_article_id[int(table_item.data(5))]
                if value >= 0:
                    color = color_yes
                    note = "На складе %s" % warehouse_sql[1]
                else:
                    color = color_no
                    note = "Не хватает %s" % -value
                    error = True

                for col in range(7):
                    self.tw_position.item(row, col).setBackground(color)
                    self.tw_position.item(row, col).setToolTip(note)

        if not error and not self.save_change_order_position:
            self.cb_shipping.setEnabled(True)

    def ui_export(self):
        path = QFileDialog.getSaveFileName(self, "Сохранение")
        if path[0]:
            to_excel.table_to_excel(self.tw_position, path[0])

    def ui_import_edi(self):
        self.position = ImportEDI(self)
        self.position.setModal(True)
        self.position.show()

    def ui_document_list(self):
        self.position = OrderDocList(self)
        self.position.setModal(True)
        self.position.show()

    def ui_view_label(self, row):
        data = {
                "article": self.tw_position_label.item(row, 0).text(),
                "article_size": self.tw_position_label.item(row, 1).text(),
                "article_parametr": self.tw_position_label.item(row, 2).text(),
                "clients_vendor": self.cb_clients_vendor.currentText(),
                "date_order": self.de_date_order.date().toString("dd.MM.yyyy"),
                "number_order": self.le_number_order.text()}

        self.print_label = print_label.LabelFile(self.tw_position_label.item(row, 2).data(5), "Путь корень бирки", data)
        self.print_label.setModal(True)
        self.print_label.show()

    def ui_print_order(self):
        head = "Заказ"

        up_html = """
          <table>
          <caption>#caption#</caption>
          <tr>
          <th>Клиент</th><th>Постащик</th><th>Дата заказа</th><th>№ Заказа</th>
          </tr>
          <tr>
          <td>#client#</td><td>#clients_vendor#</td><td>#date_order#</td><td>#number_order#</td>
          </tr>
          <tr>
          </table>
          <table>
          <th>Пункт разгрузки</th><th>Транспортная</th><th>Дата отгрузки</th><th>№ Документа</th>
          </tr>
          <tr>
          <td>#clients_adress#</td><td>#transport_company#</td><td>#date_shipment#</td><td>#number_doc#</td>
          </tr>
          </table>
          <table>
          <th>Примечание</th>
          </tr>
          <tr>
          <td>#note#</td>
          </tr>
          </table>
          """

        up_html = up_html.replace("#caption#", head)
        up_html = up_html.replace("#client#", str(self.le_client.text()))
        up_html = up_html.replace("#clients_vendor#", str(self.cb_clients_vendor.currentText()))
        up_html = up_html.replace("#date_order#", self.de_date_order.date().toString("dd.MM.yyyy"))
        up_html = up_html.replace("#number_order#", str(self.le_number_order.text()))

        up_html = up_html.replace("#clients_adress#", str(self.cb_clients_adress.currentText()))
        up_html = up_html.replace("#transport_company#", str(self.le_transport_company.text()))
        up_html = up_html.replace("#date_shipment#", self.de_date_shipment.date().toString("dd.MM.yyyy"))
        up_html = up_html.replace("#number_doc#", str(self.le_number_doc.text()))

        up_html = up_html.replace("#note#", str(self.le_note.text()))

        html = table_to_html.tab_html(self.tw_position, up_template=up_html)
        self.print_class = print_qt.PrintHtml(self, html)

    def ui_acc(self):
        if self.save_sql():
            self.close()
            self.destroy()
            if self.main != 0:
                self.main.ui_update()

    def ui_can(self):
        if (self.save_change_order or self.save_change_order_position) and self.access_save_sql:
            result = QMessageBox.question(self, "Сохранить?", "Сохранить изменение перед выходом?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if result == 16384:
                self.save_sql()
        self.close()
        self.destroy()

    def save_sql(self):
        if not self.save_change_order:  # Если нечего сохранять то проверим цену!
            query = "SELECT Sum_In_Nds, Sum_Off_Nds FROM `order` WHERE Id = %s"
            sql_info = my_sql.sql_select(query, (self.id,))

            if not sql_info:
                self.save_change_order = True
                self.save_change_order_position = True

            if sql_info[0][0] != str_to.str_to_decimal(self.le_sum_in_nds.text()) or sql_info[0][1] != str_to.str_to_decimal(self.le_sum_no_nds.text()):
                self.save_change_order = True
                self.save_change_order_position = True

        if self.save_change_order:

            # Проверка номера документа
            query = "SELECT Id FROM `order` WHERE Number_Doc = %s AND YEAR(Date_Shipment) = %s"
            sql_info = my_sql.sql_select(query, (int(self.le_number_doc.text()), self.de_date_shipment.date().year()))
            if "mysql.connector.errors" in str(type(sql_info)):
                QMessageBox.critical(self, "Ошибка sql проверки номера документа", sql_info.msg, QMessageBox.Ok)
                return False

            if self.id:
                if len(sql_info) > 1:
                    QMessageBox.critical(self, "Ошибка номера документа", "Этот номер задвоен!!! Проверить номера!!!", QMessageBox.Ok)
                    return False
                elif len(sql_info) == 1 and sql_info[0][0] != int(self.id):
                    QMessageBox.critical(self, "Ошибка номера документа", "Такой номер в этом году уже используется!", QMessageBox.Ok)
                    return False
            else:
                if sql_info:
                    QMessageBox.critical(self, "Ошибка номера документа", "Такой номер в этом году уже есть!", QMessageBox.Ok)
                    return False

            if self.cb_shipping.isChecked():
                shipped = 1
            else:
                shipped = 0

            if self.cb_combined.isChecked():
                combined = 1
            else:
                combined = 0

            if self.id:
                if self.le_transport_company.text() == "":
                    tc_id = None
                else:
                    tc_id = self.le_transport_company.whatsThis()

                query = """UPDATE `order` SET Client_Id = %s, Clients_Vendor_Id = %s, Clients_Adress_Id = %s, Transport_Company_Id = %s, Date_Order = %s,
                            Date_Shipment = %s, Number_Order = %s, Number_Doc = %s, Note = %s, Sum_In_Nds = %s, Sum_Off_Nds = %s, Combined = %s WHERE Id = %s"""
                parametrs = (self.le_client.whatsThis(), self.cb_clients_vendor.currentData(), self.cb_clients_adress.currentData(),
                             tc_id, self.de_date_order.date().toString(Qt.ISODate), self.de_date_shipment.date().toString(Qt.ISODate), self.le_number_order.text(),
                             self.le_number_doc.text(), self.le_note.text(), self.le_sum_in_nds.text(), self.le_sum_no_nds.text(), combined, self.id)
                sql_info = my_sql.sql_change(query, parametrs)
                if "mysql.connector.errors" in str(type(sql_info)):
                    QMessageBox.critical(self, "Ошибка sql изменения заказа", sql_info.msg, QMessageBox.Ok)
                    return False
                self.new_id = False
            else:
                query = """INSERT INTO `order` (Client_Id, Clients_Vendor_Id, Clients_Adress_Id, Transport_Company_Id, Date_Order, Date_Shipment,
                                                Number_Order, Number_Doc, Note, Shipped, Sum_In_Nds, Sum_Off_Nds, Combined) 
                                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 0, %s, %s, %s)"""
                parametrs = (self.le_client.whatsThis(), self.cb_clients_vendor.currentData(), self.cb_clients_adress.currentData(),
                             self.le_transport_company.whatsThis(), self.de_date_order.date().toString(Qt.ISODate),
                             self.de_date_shipment.date().toString(Qt.ISODate), self.le_number_order.text(), self.le_number_doc.text(), self.le_note.text(),
                             self.le_sum_in_nds.text(), self.le_sum_no_nds.text(), combined)
                sql_info = my_sql.sql_change(query, parametrs)
                if "mysql.connector.errors" in str(type(sql_info)):
                    QMessageBox.critical(self, "Ошибка sql добавления заказа", sql_info.msg, QMessageBox.Ok)
                    return False
                self.new_id = sql_info

        if self.save_change_order_position or self.sql_shipped != self.cb_shipping.isChecked():

            if self.cb_shipping.isChecked():
                shipped = 1
            else:
                shipped = 0

            if self.id:
                sql_order_id = self.id
            elif self.new_id:
                sql_order_id = self.new_id
            else:
                QMessageBox.critical(self, "Ошибка id в save_sql", "Ошибка нету id заказа. Это не нормально", QMessageBox.Ok)
                return False

            sql_new = []
            sql_upd = []
            sql_del = []
            sql_shipped_position = []
            if not self.sql_shipped and shipped == 1:
                str_transaction = "Заказ %s - отгружен" % self.le_number_doc.text()
            elif self.sql_shipped and shipped == 0:
                str_transaction = "Заказ %s - отменен" % self.le_number_doc.text()
            else:
                str_transaction = ""
            for row in range(self.tw_position.rowCount()):
                table_item = self.tw_position.item(row, 2)
                if table_item.data(-1) == "new":
                    sql_new.append((sql_order_id, table_item.data(5), self.tw_position.item(row, 4).text(), self.tw_position.item(row, 4).data(5),
                                    self.tw_position.item(row, 5).text(), self.tw_position.item(row, 5).data(5)))
                    if str_transaction:
                        sql_shipped_position.append([table_item.data(5), int(self.tw_position.item(row, 5).text()), str_transaction])
                elif table_item.data(-1) == "upd":
                    sql_upd.append((sql_order_id, table_item.data(5), self.tw_position.item(row, 4).text(), self.tw_position.item(row, 4).data(5),
                                    self.tw_position.item(row, 5).text(), self.tw_position.item(row, 5).data(5), table_item.data(-2)))
                    if str_transaction:
                        sql_shipped_position.append([table_item.data(5), int(self.tw_position.item(row, 5).text()), str_transaction])
                elif table_item.data(-1) == "del":
                    sql_del.append((table_item.data(-2),))
                elif table_item.data(-1) == "set":
                    if str_transaction:
                        sql_shipped_position.append([table_item.data(5), int(self.tw_position.item(row, 5).text()), str_transaction])
                else:
                    QMessageBox.critical(self, "Ошибка состояния в save_sql", "Ошибка Непонятное sql состояние строки. Это не нормально!", QMessageBox.Ok)
                    return False

            if sql_new:
                query = "INSERT INTO order_position (Order_Id, Product_Article_Parametr_Id, Price, NDS, Value, In_On_Place) VALUES (%s, %s, %s, %s, %s, %s)"
                sql_info = my_sql.sql_many(query, sql_new)
                if "mysql.connector.errors" in str(type(sql_info)):
                    QMessageBox.critical(self, "Ошибка sql добавления позиций", sql_info.msg, QMessageBox.Ok)
                    return False
            elif sql_upd:
                query = "UPDATE order_position SET Order_Id = %s, Product_Article_Parametr_Id = %s, Price = %s, NDS = %s, Value = %s, In_On_Place = %s WHERE Id = %s"
                for sql_tuple in sql_upd:
                    sql_info = my_sql.sql_change(query, sql_tuple)
                    if "mysql.connector.errors" in str(type(sql_info)):
                        QMessageBox.critical(self, "Ошибка sql изменения позиции", sql_info.msg, QMessageBox.Ok)
                        return False
            elif sql_del:
                query = "DELETE FROM order_position WHERE Id = %s"
                for sql_tuple in sql_del:
                    sql_info = my_sql.sql_change(query, sql_tuple)
                    if "mysql.connector.errors" in str(type(sql_info)):
                        QMessageBox.critical(self, "Ошибка sql удаления позиции", sql_info.msg, QMessageBox.Ok)
                        return False

            if sql_shipped_position:
                sql_connect_transaction = my_sql.sql_start_transaction()
                if shipped == 0:
                    query = "UPDATE product_article_warehouse SET Value_In_Warehouse = Value_In_Warehouse + %s WHERE Id_Article_Parametr = %s"
                else:
                    query = "UPDATE product_article_warehouse SET Value_In_Warehouse = Value_In_Warehouse - %s WHERE Id_Article_Parametr = %s"
                for item in sql_shipped_position:
                    sql_info = my_sql.sql_change_transaction(sql_connect_transaction, query, (item[1], item[0]))
                    if "mysql.connector.errors" in str(type(sql_info)):
                        my_sql.sql_rollback_transaction(sql_connect_transaction)
                        QMessageBox.critical(self, "Ошибка sql изменения склада", sql_info.msg, QMessageBox.Ok)
                        return False

                if shipped == 0:
                    query = "INSERT INTO transaction_records_warehouse (Article_Parametr_Id, Date, Balance, Note, Code) VALUES (%s, NOW(), %s, %s, 311)"
                else:
                    query = "INSERT INTO transaction_records_warehouse (Article_Parametr_Id, Date, Balance, Note, Code) VALUES (%s, NOW(), -%s, %s, 310)"
                sql_info = my_sql.sql_many_transaction(sql_connect_transaction, query, sql_shipped_position)
                if "mysql.connector.errors" in str(type(sql_info)):
                    my_sql.sql_rollback_transaction(sql_connect_transaction)
                    QMessageBox.critical(self, "Ошибка sql добавления записей изменения склада", sql_info.msg, QMessageBox.Ok)
                    return False

                query = "UPDATE `order` SET Shipped = %s WHERE Id = %s"
                sql_info = my_sql.sql_change_transaction(sql_connect_transaction, query, (shipped, self.id))
                if "mysql.connector.errors" in str(type(sql_info)):
                    my_sql.sql_rollback_transaction(sql_connect_transaction)
                    QMessageBox.critical(self, "Ошибка sql добавления записей изменения склада", sql_info.msg, QMessageBox.Ok)
                    return False

                my_sql.sql_commit_transaction(sql_connect_transaction)

        return True

    def calc_sum(self):
        old_sum = self.le_sum_in_nds.text()
        if self.cb_combined.isChecked():
            query = """SELECT DISTINCT(product_article_parametrs.Client_code)
                          FROM order_position LEFT JOIN product_article_parametrs ON order_position.Product_Article_Parametr_Id = product_article_parametrs.Id
                          WHERE Order_Id = %s"""
            unite_code_sql = my_sql.sql_select(query, (self.id, ))
            if "mysql.connector.errors" in str(type(unite_code_sql)):
                QMessageBox.critical(self, "Ошибка sql получения уникальных кодов", unite_code_sql.msg, QMessageBox.Ok)
                return False

            product = collections.OrderedDict()

            for cod in unite_code_sql:
                product[str(cod[0])] = {"value": 0, "price": None, "price_no_nds": None, "sum_no_nds": 0, "nds": None, "sum": 0, "nds_sum": 0, "psb": 0}

            aricle = {}
            query = """SELECT product_article_parametrs.Id ,product_article_parametrs.Client_code
                          FROM order_position LEFT JOIN product_article_parametrs ON order_position.Product_Article_Parametr_Id = product_article_parametrs.Id
                          WHERE Order_Id = %s"""
            sql_info = my_sql.sql_select(query, (self.id, ))
            if "mysql.connector.errors" in str(type(sql_info)):
                QMessageBox.critical(self, "Ошибка sql получения кодов всех артикулов", sql_info.msg, QMessageBox.Ok)
                return False

            for i in sql_info:
                aricle.update({i[0]: i[1]})

            for row in range(self.tw_position.rowCount()):
                if self.tw_position.isRowHidden(row):
                    continue  # Если строка крыта то мы ее не считаем!!!
                # Проверяем находиться ли наный товар в списке объединяемых
                cod = aricle[int(self.tw_position.item(row, 2).data(5))]
                if cod not in product:
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Не найден код в словаре!", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли цены
                if product[cod]["price"] is not None and product[cod]["price"] != float(self.tw_position.item(row, 4).text()):
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разная цена одинаковых кодов", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли НДС
                if product[cod]["nds"] is not None and product[cod]["nds"] != int(self.tw_position.item(row, 4).data(5)):
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разный НДС одинаковых кодов", QMessageBox.Ok)
                    return False

                # Вставляем статические параметры
                product[cod]["price"] = float(self.tw_position.item(row, 4).text().replace(",", "."))
                product[cod]["nds"] = int(self.tw_position.item(row, 4).data(5))
                product[cod]["psb"] = int(self.tw_position.item(row, 5).data(5))

                # Сумируем кол-во
                product[cod]["value"] += int(self.tw_position.item(row, 5).text())

            # Пройдем по готовым обьединеным позициям и расчитаем цифры
            sum_in_nds = 0
            sum_of_nds = 0
            unit_position = 0
            for cod, cod_value in product.items():
                if self.lb_client.whatsThis().find("no_nds") >= 0:
                    cod_value["price_no_nds"], cod_value["sum_no_nds"], cod_value["nds_sum"], cod_value["sum"], cod_value["mest"] =\
                        calc_price.calc_no_nds(cod_value["price"], cod_value["value"], cod_value["nds"], cod_value["psb"])
                else:
                    cod_value["price_no_nds"], cod_value["sum_no_nds"], cod_value["nds_sum"], cod_value["sum"], cod_value["mest"] =\
                        calc_price.calc_nds(cod_value["price"], cod_value["value"], cod_value["nds"], cod_value["psb"])

                sum_of_nds += cod_value["sum_no_nds"]
                sum_in_nds += cod_value["sum"]
                unit_position += 1

            self.le_sum_position.setText(str(self.tw_position.rowCount()))
            self.le_sum_no_nds.setText(str(round(sum_of_nds, 2)))
            self.le_sum_in_nds.setText(str(round(sum_in_nds, 2)))
            self.le_sum_nds.setText(str(round(sum_in_nds - sum_of_nds, 2)))

            if old_sum != self.le_sum_in_nds.text():
                self.ui_order_info_edit()


        else:

            if self.tw_position.rowCount() < 1:
                return False

            sum_in_nds = 0
            sum_of_nds = 0
            for row in range(self.tw_position.rowCount()):
                if self.tw_position.isRowHidden(row):
                    continue  # Если строка крыта то мы ее не считаем!!!
                price = round(float(self.tw_position.item(row, 4).text().replace(",", ".")), 2)
                value = float(self.tw_position.item(row, 5).text().replace(",", "."))
                nds = float(self.tw_position.item(row, 4).data(5))

                if self.lb_client.whatsThis().find("no_nds") >= 0:
                    sum_of_nds += round(price * value, 2)
                    sum_in_nds += round((price * value) * (1 + nds / 100), 2)
                else:
                    sum_in_nds += round(price * value, 2)
                    sum_of_nds += round(price * value - (price * value * nds) / (100 + nds), 2)

            self.le_sum_position.setText(str(self.tw_position.rowCount()))
            self.le_sum_no_nds.setText(str(round(sum_of_nds, 2)))
            self.le_sum_in_nds.setText(str(round(sum_in_nds, 2)))
            self.le_sum_nds.setText(str(round(sum_in_nds - sum_of_nds, 2)))

            if old_sum != self.le_sum_in_nds.text():
                self.ui_order_info_edit()

    def of_list_clients(self, item):
        id_client, name_client = item
        self.le_client.setText(str(name_client))
        self.le_client.setWhatsThis(str(id_client))

        query = "SELECT No_Nds FROM clients WHERE Id = %s"
        parametrs = (id_client,)
        sql_info = my_sql.sql_select(query, parametrs)
        if "mysql.connector.errors" in str(type(sql_info)):
            QMessageBox.critical(self, "Ошибка sql получение НДС/Не НДС", sql_info.msg, QMessageBox.Ok)
            return False

        if sql_info[0][0]:
            self.lb_client.setText("<html><head/><body><p align='center'> Клиент (Без НДС) </p></body></html>")
            self.lb_client.setWhatsThis("no_nds")
        else:
            self.lb_client.setText("<html><head/><body><p align='center'> Клиент (C НДС) </p></body></html>")
            self.lb_client.setWhatsThis("nds")

        query = "SELECT Id, Name, Adres FROM clients_actual_address WHERE Client_Id = %s"
        parametrs = (id_client,)
        sql_info = my_sql.sql_select(query, parametrs)
        if "mysql.connector.errors" in str(type(sql_info)):
            QMessageBox.critical(self, "Ошибка sql получение адресов клиента", sql_info.msg, QMessageBox.Ok)
            return False

        if not sql_info:
            query = "SELECT Actual_Address FROM clients WHERE Id = %s"
            parametrs = (id_client,)
            sql_info = my_sql.sql_select(query, parametrs)
            if "mysql.connector.errors" in str(type(sql_info)):
                QMessageBox.critical(self, "Ошибка sql получение адреса клиента", sql_info.msg, QMessageBox.Ok)
                return False
            if sql_info[0][0] == "":
                self.cb_clients_adress.clear()
                self.cb_clients_adress.addItem("None", None)
                self.cb_clients_adress.setEnabled(False)
            else:
                self.cb_clients_adress.clear()
                self.cb_clients_adress.addItem(sql_info[0][0], None)
                self.cb_clients_adress.setEnabled(False)

        else:
            self.cb_clients_adress.clear()
            self.cb_clients_adress.setEnabled(True)
            for item in sql_info:
                self.cb_clients_adress.addItem(item[1], item[0])

        query = 'SELECT Id, CONCAT_WS(", ", Number, Contract, Data_From) FROM clients_vendor_number WHERE Client_Id = %s'
        parametrs = (id_client,)
        sql_info = my_sql.sql_select(query, parametrs)
        if "mysql.connector.errors" in str(type(sql_info)):
            QMessageBox.critical(self, "Ошибка sql получение номеров клиента", sql_info.msg, QMessageBox.Ok)
            return False

        if sql_info:
            self.cb_clients_vendor.clear()
            self.cb_clients_vendor.setEnabled(True)
            for item in sql_info:
                self.cb_clients_vendor.addItem(item[1], item[0])
        else:
            self.cb_clients_vendor.clear()
            self.cb_clients_vendor.addItem("None", None)
            self.cb_clients_vendor.setEnabled(False)

        self.pb_add_position.setEnabled(True)
        self.pb_change_position.setEnabled(True)
        self.pb_dell_position.setEnabled(True)

    def of_list_insert(self, item):
        self.le_transport_company.setText(item[1])
        self.le_transport_company.setWhatsThis(str(item[0]))

    def of_ex_torg12(self,edo, head, article, addres, unite, manager_name, no_pcb):
        path = QFileDialog.getSaveFileName(self, "Сохранение", filter="Excel(*.xlsx)")
        if not path[0]:
            return False

        border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        border_all_big = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

        font_8 = Font(name="Arial", size=8)

        ald_center = Alignment(horizontal="center")
        ald_right = Alignment(horizontal="right")

        wite = PatternFill(start_color='ffffff',
                              end_color='ffffff',
                              fill_type='solid')

        book = openpyxl.load_workbook(filename='%s/Накладная 2.xlsx' % (getcwd() + "/templates/order"))
        sheet = book['Отчет']

        sheet.oddHeader.right.text = "Продолжение накладной № %s от %s г." % (self.le_number_doc.text(), self.de_date_shipment.date().toString("dd.MM.yyyy"))
        sheet.oddHeader.right.size = 7

        # заполнение шапки
        if edo:
            sheet["C1"] = "   -   ЭДО   -   "
            sheet["C1"].border = border_all_big

        if not head:
            sheet["A2"] = ""

        query = "SELECT Name,  INN, KPP, Actual_Address, Legal_Address, Account, Bank, corres_Account, BIK, Full_Name FROM clients WHERE Id = %s"
        sql_info = my_sql.sql_select(query, (self.le_client.whatsThis(), ))
        if "mysql.connector.errors" in str(type(sql_info)):
            QMessageBox.critical(self, "Ошибка sql получения информации о клиенте", sql_info.msg, QMessageBox.Ok)
            return False

        if addres:
            adr = sql_info[0][3]
            kpp = sql_info[0][2]
        else:
            query = "SELECT Adres, KPP FROM clients_actual_address WHERE Id = %s"
            sql_adr = my_sql.sql_select(query, (self.cb_clients_adress.currentData(),))
            if "mysql.connector.errors" in str(type(sql_adr)):
                QMessageBox.critical(self, "Ошибка sql получения пункти разгрузки", sql_adr.msg, QMessageBox.Ok)
                return False

            if not sql_adr:
                QMessageBox.critical(self, "Нету адреса", "У клиента нету пункта разгрузки!", QMessageBox.Ok)
                return False
            adr = sql_adr[0][0]
            kpp = sql_adr[0][1]

        client = sql_info[0][9] + " ИНН " + str(sql_info[0][1])
        if sql_info[0][2]:
            client += " КПП " + str(kpp)
        client += ", " + adr + ",р/с " + str(sql_info[0][5]) + " в " + sql_info[0][6] + " к/с " + str(sql_info[0][7]) + " БИК " + str(sql_info[0][8])
        sheet["C8"] = client

        client = sql_info[0][9] + " ИНН " + str(sql_info[0][1])
        if sql_info[0][2]:
            client += " КПП " + str(sql_info[0][2])
        client += ", " + sql_info[0][4] + ",р/с " + str(sql_info[0][5]) + " в " + sql_info[0][6] + " к/с " + str(sql_info[0][7]) + " БИК " + str(sql_info[0][8])
        sheet["C12"] = client

        if self.cb_clients_vendor.currentData():
            query = "SELECT Number, Contract, Data_From FROM clients_vendor_number WHERE Id = %s"
            sql_info = my_sql.sql_select(query, (self.cb_clients_vendor.currentData(), ))
            if "mysql.connector.errors" in str(type(sql_info)):
                QMessageBox.critical(self, "Ошибка sql получения информации номера поставщика", sql_info.msg, QMessageBox.Ok)
                return False

            base = "№ заказа " + self.le_number_order.text() + ", Номер поставщика " + str(sql_info[0][0]) \
                   + " договор поставки № " + str(sql_info[0][1]) + " от " + sql_info[0][2].strftime("%d.%m.%Y")

            sheet["C14"] = base

        # заполнение середины
        sheet["G17"] = self.le_number_doc.text()
        sheet["T13"] = self.le_number_doc.text()
        sheet["I17"] = self.de_date_shipment.date().toString("dd.MM.yyyy")
        sheet["T14"] = self.de_date_shipment.date().toString("dd.MM.yyyy")
        sheet["G17"].border = border_all_big
        sheet["I17"].border = border_all_big

        if unite:
            query = """SELECT DISTINCT(product_article_parametrs.Client_code)
                          FROM order_position LEFT JOIN product_article_parametrs ON order_position.Product_Article_Parametr_Id = product_article_parametrs.Id
                          WHERE Order_Id = %s"""
            unite_code_sql = my_sql.sql_select(query, (self.id, ))
            if "mysql.connector.errors" in str(type(unite_code_sql)):
                QMessageBox.critical(self, "Ошибка sql получения уникальных кодов", unite_code_sql.msg, QMessageBox.Ok)
                return False

            product = collections.OrderedDict()

            for cod in unite_code_sql:
                product[str(cod[0])] = {"name": None, "cod": str(cod[0]), "psb": None, "mest": 0, "value": 0, "price": None, "price_no_nds": None,
                                        "sum_no_nds": 0, "nds": None, "sum": 0, "nds_sum": 0}

            for row in range(self.tw_position.rowCount()):
                query = "SELECT Client_code FROM product_article_parametrs WHERE Id = %s"
                sql_info = my_sql.sql_select(query, (self.tw_position.item(row, 2).data(5),))
                if "mysql.connector.errors" in str(type(sql_info)):
                    QMessageBox.critical(self, "Ошибка sql получения информации номера поставщика", sql_info.msg, QMessageBox.Ok)
                    return False

                # Проверяем находиться ли наный товар в списке объединяемых
                cod = sql_info[0][0]
                if cod not in product:
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Не найден код в словаре!", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли имена
                if product[cod]["name"] is not None and product[cod]["name"] != self.tw_position.item(row, 3).text():
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разные имена одинаковых кодов", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли цены
                if product[cod]["price"] is not None and product[cod]["price"] != float(self.tw_position.item(row, 4).text()):
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разная цена одинаковых кодов", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли psb
                if product[cod]["psb"] is not None and product[cod]["psb"] != int(self.tw_position.item(row, 5).data(5)):
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разный psb одинаковых кодов", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли НДС
                if product[cod]["nds"] is not None and product[cod]["nds"] != int(self.tw_position.item(row, 4).data(5)):
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разный НДС одинаковых кодов", QMessageBox.Ok)
                    return False

                # Вставляем статические параметры
                product[cod]["name"] = self.tw_position.item(row, 3).text()
                product[cod]["price"] = float(self.tw_position.item(row, 4).text())
                product[cod]["psb"] = int(self.tw_position.item(row, 5).data(5))
                product[cod]["nds"] = int(self.tw_position.item(row, 4).data(5))

                # Сумируем кол-во
                product[cod]["value"] += int(self.tw_position.item(row, 5).text())

            # Пройдем по готовым обьединеным позициям и расчитаем цифры
            for cod, cod_value in product.items():
                if self.lb_client.whatsThis().find("no_nds") >= 0:
                    cod_value["price_no_nds"], cod_value["sum_no_nds"], cod_value["nds_sum"], cod_value["sum"], cod_value["mest"] =\
                        calc_price.calc_no_nds(cod_value["price"], cod_value["value"], cod_value["nds"], cod_value["psb"])
                else:
                    cod_value["price_no_nds"], cod_value["sum_no_nds"], cod_value["nds_sum"], cod_value["sum"], cod_value["mest"] =\
                        calc_price.calc_nds(cod_value["price"], cod_value["value"], cod_value["nds"], cod_value["psb"])

        else:
            # если
            product = collections.OrderedDict()
            for row in range(self.tw_position.rowCount()):

                product[row] = {"name": None, "cod": None, "psb": None, "mest": 0, "value": 0, "price": None,
                                                                 "price_no_nds": None, "sum_no_nds": 0, "nds": None, "sum": 0, "nds_sum": 0}

                # Делаем ссылку дя удобства
                position = product[row]

                query = "SELECT Barcode, Client_code FROM product_article_parametrs WHERE Id = %s"
                sql_info = my_sql.sql_select(query, (self.tw_position.item(row, 2).data(5),))
                if "mysql.connector.errors" in str(type(sql_info)):
                    QMessageBox.critical(self, "Ошибка sql получения информации номера поставщика", sql_info.msg, QMessageBox.Ok)
                    return False
                if article:
                    position["name"] = self.tw_position.item(row, 3).text() + " а " + self.tw_position.item(row, 0).text() + \
                                       " р " + self.tw_position.item(row, 1).text() + " " + str(sql_info[0][0])
                    position["cod"] = sql_info[0][1]
                else:
                    position["name"] = self.tw_position.item(row, 3).text() + " " + str(sql_info[0][0])
                    position["cod"] = sql_info[0][1]

                position["nds"] = int(self.tw_position.item(row, 4).data(5))
                position["price"] = float(self.tw_position.item(row, 4).text())
                position["psb"] = int(self.tw_position.item(row, 5).data(5))
                position["value"] += int(self.tw_position.item(row, 5).text())

                if self.lb_client.whatsThis().find("no_nds") >= 0:
                    position["price_no_nds"], position["sum_no_nds"], position["nds_sum"], position["sum"], position["mest"] =\
                        calc_price.calc_no_nds(position["price"], position["value"], position["nds"], position["psb"])

                else:
                    position["price_no_nds"], position["sum_no_nds"], position["nds_sum"], position["sum"], position["mest"] =\
                        calc_price.calc_nds(position["price"], position["value"], position["nds"], position["psb"])

        all_value = 0
        all_no_nds = 0
        all_nds = 0
        all_sum = 0
        all_position = 0

        list_all = 1
        row_break = 14
        row_ex = 22

        num = 1
        for cod, position in product.items():
            sheet.merge_cells("B%s:D%s" % (row_ex, row_ex))
            sheet.merge_cells("L%s:M%s" % (row_ex, row_ex))
            sheet.merge_cells("N%s:O%s" % (row_ex, row_ex))
            sheet.merge_cells("P%s:Q%s" % (row_ex, row_ex))
            sheet.merge_cells("R%s:S%s" % (row_ex, row_ex))
            sheet.merge_cells("T%s:U%s" % (row_ex, row_ex))

            sheet["A%s" % row_ex] = num
            sheet["B%s" % row_ex] = position["name"]
            sheet["B%s" % row_ex].alignment = Alignment(wrapText=True)
            sheet["B%s" % row_ex].font = font_8
            sheet["E%s" % row_ex] = position["cod"]
            sheet["F%s" % row_ex] = "шт."
            sheet["G%s" % row_ex] = "796"
            sheet["H%s" % row_ex] = "кор."
            if not no_pcb:
                sheet["I%s" % row_ex] = position["psb"]
            sheet["J%s" % row_ex] = position["mest"]
            sheet["L%s" % row_ex] = position["value"]
            sheet["N%s" % row_ex] = moneyfmt.moneyfmt(position["price_no_nds"])
            sheet["N%s" % row_ex].alignment = ald_right
            sheet["P%s" % row_ex] = moneyfmt.moneyfmt(position["sum_no_nds"])
            sheet["P%s" % row_ex].alignment = ald_right
            sheet["R%s" % row_ex] = position["nds"]
            sheet["T%s" % row_ex] = moneyfmt.moneyfmt(position["nds_sum"])
            sheet["T%s" % row_ex].alignment = ald_right
            sheet["V%s" % row_ex] = moneyfmt.moneyfmt(position["sum"])
            sheet["V%s" % row_ex].alignment = ald_right

            all_position += position["mest"]
            all_value += position["value"]
            all_no_nds += position["sum_no_nds"]
            all_nds += round(position["nds_sum"], 2)
            all_sum += position["sum"]

            sheet.row_dimensions[row_ex].height = 23

            if row_break == 25 and len(product) > 16:
                sheet.page_breaks.append(Break(row_ex))
                list_all += 1
                row_break = 0

            row_break += 1
            row_ex += 1
            num += 1

        if row_break + 7 > 25:
            sheet.page_breaks.append(Break(row_ex-4))
            list_all += 1

        # Заполняем сумму
        sheet["I%s" % row_ex] = "Всего по накладной"
        sheet["I%s" % row_ex].alignment = ald_right

        sheet["J%s" % row_ex] = all_position
        sheet["J%s" % row_ex].alignment = ald_center

        sheet["K%s" % row_ex] = "X"
        sheet["K%s" % row_ex].alignment = ald_center

        sheet.merge_cells("L%s:M%s" % (row_ex, row_ex))
        sheet["L%s" % row_ex] = all_value
        sheet["L%s" % row_ex].alignment = ald_right

        sheet.merge_cells("N%s:O%s" % (row_ex, row_ex))
        sheet["N%s" % row_ex] = "X"
        sheet["N%s" % row_ex].alignment = ald_center

        sheet.merge_cells("P%s:Q%s" % (row_ex, row_ex))
        sheet["P%s" % row_ex] = moneyfmt.moneyfmt(all_no_nds)
        sheet["P%s" % row_ex].alignment = ald_right

        sheet.merge_cells("R%s:S%s" % (row_ex, row_ex))
        sheet["R%s" % row_ex] = "X"
        sheet["R%s" % row_ex].alignment = ald_center

        sheet.merge_cells("T%s:U%s" % (row_ex, row_ex))
        sheet["T%s" % row_ex] = moneyfmt.moneyfmt(all_nds)
        sheet["T%s" % row_ex].alignment = ald_right

        sheet["V%s" % row_ex] = moneyfmt.moneyfmt(all_sum)
        sheet["V%s" % row_ex].alignment = ald_right

        for row in sheet.iter_rows(min_row=row_ex, min_col=10, max_col=22):
            for cell in row:
                cell.border = border_all

        row_ex += 1

        # Формируем шапку
        for row in sheet.iter_rows(min_row=19, max_col=22, max_row=21):
            for cell in row:
                    cell.border = border_all_big

        for row in sheet.iter_rows(min_row=4, min_col=20, max_row=5):
            for cell in row:
                cell.border = border_all_big

        for row in sheet.iter_rows(min_row=17, min_col=7, max_col=10, max_row=17):
            for cell in row:
                cell.border = border_all_big

        for row in sheet.iter_rows(min_row=13, min_col=17, max_col=19, max_row=16):
            for cell in row:
                cell.border = border_all

        # Формируем шапку-правую колонку
        for row in sheet.iter_rows(min_row=13, min_col=20, max_col=22, max_row=14):
            for cell in row:
                cell.border = border_all
        sheet["T13"].border = Border(left=Side(style='medium'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        sheet["T14"].border = Border(left=Side(style='medium'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        sheet["V13"].border = Border(left=Side(style='thin'), right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='thin'))
        sheet["V14"].border = Border(left=Side(style='thin'), right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Формируем границы таблицы
        for row in sheet.iter_rows(min_row=22, max_col=22, max_row=row_ex-2):
            for cell in row:
                cell.border = border_all

        sheet2 = book['Низ']

        # Вставляем имя менеджера
        sheet2['G14'] = manager_name

        for row in sheet2.iter_rows(min_row=1, max_col=22, max_row=17):
            for cell in row:
                sheet["%s%s" % (cell.column, row_ex)] = cell.value
                sheet.row_dimensions[row_ex].height = sheet2.row_dimensions[cell.row].height
                if cell.has_style:
                    sheet["%s%s" % (cell.column, row_ex)].border = copy(cell.border)
                    sheet["%s%s" % (cell.column, row_ex)].font = copy(cell.font)
                    sheet["%s%s" % (cell.column, row_ex)].fill = wite

            row_ex += 1

        sheet["I%s" % (row_ex - 17)] = list_all

        # Числа прописью
        int_units = ((u'рубль', u'рубля', u'рублей'), 'm')
        exp_units = ((u'копейка', u'копейки', u'копеек'), 'f')
        sheet["D%s" % (row_ex-16)] = num2t4ru.num2text(len(product))
        sheet["D%s" % (row_ex-13)] = num2t4ru.num2text(all_position)
        sheet["D%s" % (row_ex-9)] = num2t4ru.decimal2text(Decimal(str(all_sum)), int_units=int_units, exp_units=exp_units)

        book.remove(sheet2)

        book.save(path[0])

    def of_ex_invoice(self, article, addres, unite):
        path = QFileDialog.getSaveFileName(self, "Сохранение", filter="Excel(*.xlsx)")
        if not path[0]:
            return False

        border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        font_7 = Font(name="Arial", size=7)
        alg_center = Alignment(horizontal="center")
        alg_right = Alignment(horizontal="right")
        wite = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')

        book = openpyxl.load_workbook(filename='%s/фактура.xlsx' % (getcwd() + "/templates/order"))
        sheet = book['Отчет']

        sheet.oddHeader.right.text = "Продолжение счета-фактуры № %s от %s г." % (self.le_number_doc.text(), self.de_date_shipment.date().toString("dd.MM.yyyy"))
        sheet.oddHeader.right.size = 7

        sheet["A2"] = "Счет-фактура № %s от %s" % (self.le_number_doc.text(), self.de_date_shipment.date().toString("dd.MM.yyyy"))

        query = "SELECT Name, INN, KPP, Actual_Address, Legal_Address, Full_Name FROM clients WHERE Id = %s"
        sql_info = my_sql.sql_select(query, (self.le_client.whatsThis(),))
        if "mysql.connector.errors" in str(type(sql_info)):
            QMessageBox.critical(self, "Ошибка sql получения информации о клиенте", sql_info.msg, QMessageBox.Ok)
            return False

        if addres == "fact":
            adr = sql_info[0][3]
            kpp = sql_info[0][2]
        elif addres == "adr list":
            query = "SELECT Adres, KPP FROM clients_actual_address WHERE Id = %s"
            sql_adr = my_sql.sql_select(query, (self.cb_clients_adress.currentData(),))
            if "mysql.connector.errors" in str(type(sql_adr)):
                QMessageBox.critical(self, "Ошибка sql получения пункти разгрузки", sql_adr.msg, QMessageBox.Ok)
                return False
            adr = sql_adr[0][0]
            kpp = sql_adr[0][1]
        else:
            adr = sql_info[0][4]
            kpp = sql_info[0][2]

        sheet["A8"] = "Грузополучатель и его адрес: " + sql_info[0][5] + " " + adr
        sheet["A10"] = "Покупатель: " + sql_info[0][5]
        sheet["A11"] = "Адрес: " + sql_info[0][4]
        sheet["A12"] = "ИНН/КПП покупателя: " + sql_info[0][1] + "/" + str(kpp)

        if self.cb_clients_vendor.currentData():
            query = "SELECT Number, Contract, Data_From FROM clients_vendor_number WHERE Client_Id = %s"
            sql_info = my_sql.sql_select(query, (self.le_client.whatsThis(), ))
            if "mysql.connector.errors" in str(type(sql_info)):
                QMessageBox.critical(self, "Ошибка sql получения информации номера поставщика", sql_info.msg, QMessageBox.Ok)
                return False
            base = "Идентификатор государственного контракта, договора (соглашения) (при наличии): Договор поставки № " + str(sql_info[0][1]) + \
                   " от " + sql_info[0][2].strftime("%d.%m.%Y")
            sheet["A14"] = base

            base = "№ поставщика: %s, № заказа: %s" % (str(sql_info[0][0]), self.le_number_order.text())
            sheet["A15"] = base

        if unite:
            query = """SELECT DISTINCT(product_article_parametrs.Client_code)
                          FROM order_position LEFT JOIN product_article_parametrs ON order_position.Product_Article_Parametr_Id = product_article_parametrs.Id
                          WHERE Order_Id = %s"""
            unite_code_sql = my_sql.sql_select(query, (self.id, ))
            if "mysql.connector.errors" in str(type(unite_code_sql)):
                QMessageBox.critical(self, "Ошибка sql получения уникальных кодов", unite_code_sql.msg, QMessageBox.Ok)
                return False

            product = collections.OrderedDict()

            for cod in unite_code_sql:
                product[str(cod[0])] = {"name": None, "cod": str(cod[0]), "psb": None, "mest": 0, "value": 0, "price": None, "price_no_nds": None,
                                        "sum_no_nds": 0, "nds": None, "sum": 0, "nds_sum": 0}

            for row in range(self.tw_position.rowCount()):
                query = "SELECT Client_code FROM product_article_parametrs WHERE Id = %s"
                sql_info = my_sql.sql_select(query, (self.tw_position.item(row, 2).data(5),))
                if "mysql.connector.errors" in str(type(sql_info)):
                    QMessageBox.critical(self, "Ошибка sql получения информации номера поставщика", sql_info.msg, QMessageBox.Ok)
                    return False

                # Проверяем находиться ли наный товар в списке объединяемых
                cod = sql_info[0][0]
                if cod not in product:
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Не найден код в словаре!", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли имена
                if product[cod]["name"] is not None and product[cod]["name"] != self.tw_position.item(row, 3).text():
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разные имена одинаковых кодов", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли цены
                if product[cod]["price"] is not None and product[cod]["price"] != float(self.tw_position.item(row, 4).text()):
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разная цена одинаковых кодов", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли psb
                if product[cod]["psb"] is not None and product[cod]["psb"] != int(self.tw_position.item(row, 5).data(5)):
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разный psb одинаковых кодов", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли НДС
                if product[cod]["nds"] is not None and product[cod]["nds"] != int(self.tw_position.item(row, 4).data(5)):
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разный НДС одинаковых кодов", QMessageBox.Ok)
                    return False

                # Вставляем статические параметры
                product[cod]["name"] = self.tw_position.item(row, 3).text()
                product[cod]["price"] = float(self.tw_position.item(row, 4).text())
                product[cod]["psb"] = int(self.tw_position.item(row, 5).data(5))
                product[cod]["nds"] = int(self.tw_position.item(row, 4).data(5))

                # Сумируем кол-во
                product[cod]["value"] += int(self.tw_position.item(row, 5).text())

            # Пройдем по готовым обьединеным позициям и расчитаем цифры
            for cod, cod_value in product.items():
                if self.lb_client.whatsThis().find("no_nds") >= 0:
                    cod_value["price_no_nds"], cod_value["sum_no_nds"], cod_value["nds_sum"], cod_value["sum"], cod_value["mest"] =\
                        calc_price.calc_no_nds(cod_value["price"], cod_value["value"], cod_value["nds"], cod_value["psb"])
                else:
                    cod_value["price_no_nds"], cod_value["sum_no_nds"], cod_value["nds_sum"], cod_value["sum"], cod_value["mest"] =\
                        calc_price.calc_nds(cod_value["price"], cod_value["value"], cod_value["nds"], cod_value["psb"])

        else:
            # если
            product = collections.OrderedDict()
            for row in range(self.tw_position.rowCount()):

                product[row] = {"name": None, "cod": None, "psb": None, "mest": 0, "value": 0, "price": None,
                                                                 "price_no_nds": None, "sum_no_nds": 0, "nds": None, "sum": 0, "nds_sum": 0}

                # Делаем ссылку дя удобства
                position = product[row]

                query = "SELECT Barcode, Client_code FROM product_article_parametrs WHERE Id = %s"
                sql_info = my_sql.sql_select(query, (self.tw_position.item(row, 2).data(5),))
                if "mysql.connector.errors" in str(type(sql_info)):
                    QMessageBox.critical(self, "Ошибка sql получения информации номера поставщика", sql_info.msg, QMessageBox.Ok)
                    return False
                if article:
                    position["name"] = self.tw_position.item(row, 3).text() + " а " + self.tw_position.item(row, 0).text() + \
                                       " р " + self.tw_position.item(row, 1).text() + " " + str(sql_info[0][0])
                    position["cod"] = sql_info[0][1]
                else:
                    position["name"] = self.tw_position.item(row, 3).text() + " " + str(sql_info[0][0])
                    position["cod"] = sql_info[0][1]

                position["nds"] = int(self.tw_position.item(row, 4).data(5))
                position["price"] = float(self.tw_position.item(row, 4).text())
                position["psb"] = int(self.tw_position.item(row, 5).data(5))
                position["value"] += int(self.tw_position.item(row, 5).text())

                if self.lb_client.whatsThis().find("no_nds") >= 0:
                    position["price_no_nds"], position["sum_no_nds"], position["nds_sum"], position["sum"], position["mest"] =\
                        calc_price.calc_no_nds(position["price"], position["value"], position["nds"], position["psb"])

                else:
                    position["price_no_nds"], position["sum_no_nds"], position["nds_sum"], position["sum"], position["mest"] =\
                        calc_price.calc_nds(position["price"], position["value"], position["nds"], position["psb"])

        all_no_nds = 0
        all_nds = 0
        all_sum = 0

        list_all = 1
        row_break = 12
        row_ex = 20
        num = 1
        for cod, position in product.items():
            sheet.merge_cells("A%s:D%s" % (row_ex, row_ex))

            sheet["A%s" % row_ex] = position["name"]
            sheet["A%s" % row_ex].font = font_7
            sheet["A%s" % row_ex].alignment = Alignment(wrapText=True)
            sheet["E%s" % row_ex] = "-"
            sheet["E%s" % row_ex].alignment = alg_center
            sheet["F%s" % row_ex] = position["cod"]

            sheet["G%s" % row_ex] = "796"
            sheet["H%s" % row_ex] = "Шт."
            sheet["I%s" % row_ex] = position["value"]
            sheet["J%s" % row_ex] = moneyfmt.moneyfmt(position["price_no_nds"])
            sheet["J%s" % row_ex].alignment = alg_right
            sheet["K%s" % row_ex] = moneyfmt.moneyfmt(position["sum_no_nds"])
            sheet["K%s" % row_ex].alignment = alg_right
            sheet["L%s" % row_ex] = "без акциза"
            sheet["L%s" % row_ex].font = font_7
            sheet["M%s" % row_ex] = position["nds"]
            sheet["N%s" % row_ex] = moneyfmt.moneyfmt(position["nds_sum"])
            sheet["N%s" % row_ex].alignment = alg_right
            sheet["O%s" % row_ex] = moneyfmt.moneyfmt(position["sum"])
            sheet["O%s" % row_ex].alignment = alg_right
            sheet["Q%s" % row_ex] = "РФ"
            sheet["Q%s" % row_ex].alignment = alg_center

            all_no_nds += position["sum_no_nds"]
            all_nds += round(position["nds_sum"], 2)
            all_sum += position["sum"]

            sheet.row_dimensions[row_ex].height = 23

            if row_break == 23 and len(product) > 14:
                sheet.page_breaks.append(Break(row_ex))
                list_all += 1
                row_break = 0

            row_break += 1
            row_ex += 1

        if row_break + 5 > 23:
            sheet.page_breaks.append(Break(row_ex-4))
            list_all += 1

        # Формируем границы таблицы
        for row in sheet.iter_rows(min_row=17, max_col=18, max_row=row_ex-1):
            for cell in row:
                cell.border = border_all

        # Запишем итог
        sheet.merge_cells("A%s:J%s" % (row_ex, row_ex))
        sheet["A%s" % row_ex] = "ВСЕГО К ОПЛАТЕ"
        sheet["K%s" % row_ex] = moneyfmt.moneyfmt(all_no_nds)
        sheet["K%s" % row_ex].alignment = alg_right
        sheet["L%s" % row_ex] = "X"
        sheet["L%s" % row_ex].alignment = alg_center
        sheet["M%s" % row_ex] = "X"
        sheet["M%s" % row_ex].alignment = alg_center
        sheet["N%s" % row_ex] = moneyfmt.moneyfmt(all_nds)
        sheet["N%s" % row_ex].alignment = alg_right
        sheet["O%s" % row_ex] = moneyfmt.moneyfmt(all_sum)
        sheet["O%s" % row_ex].alignment = alg_right
        for row in sheet.iter_rows(min_row=row_ex, max_col=15):
            for cell in row:
                cell.border = border_all

        row_ex += 2

        sheet2 = book['низ']

        for row in sheet2.iter_rows(min_row=1, max_col=16, max_row=7):
            for cell in row:
                sheet["%s%s" % (cell.column, row_ex)] = cell.value
                sheet.row_dimensions[row_ex].height = sheet2.row_dimensions[cell.row].height
                if cell.has_style:
                    sheet["%s%s" % (cell.column, row_ex)].border = copy(cell.border)
                    sheet["%s%s" % (cell.column, row_ex)].font = copy(cell.font)
                    sheet["%s%s" % (cell.column, row_ex)].fill = wite

            row_ex += 1

        book.remove(sheet2)

        try:
            book.save(path[0])
        except PermissionError:
            QMessageBox.critical(self, "Ошибка сохранения", "Не удалось сохранить документ\n Скорее всего во сохраняете заместо открытого документа!", QMessageBox.Ok)

    def of_ex_ttn(self, addres, article, auto, driver, unite, manager_name):
        path = QFileDialog.getSaveFileName(self, "Сохранение", filter="Excel(*.xlsx)")
        if not path[0]:
            return False

        book = openpyxl.load_workbook(filename='%s/ТТН.xlsx' % (getcwd() + "/templates/order"))
        sheet = book['Отчет']

        border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        border_all_big = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
        ald_center_full = Alignment(horizontal="center", vertical="center", wrapText=True)
        ald_center = Alignment(horizontal="center")
        font_8 = Font(name="Arial", size=8)
        wite = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')

        sheet.oddHeader.right.text = "Продолжение товарно-транспортной накладной № %s от %s г." % (self.le_number_doc.text(), self.de_date_shipment.date().toString("dd.MM.yyyy"))
        sheet.oddHeader.right.size = 7

        # строим первый лист
        sheet["A4"] = "Срок доставки груза " + self.de_date_shipment.date().toString("dd.MM.yyyy")
        sheet["AD3"] = self.le_number_doc.text()
        query = "SELECT Name, Details FROM order_transport_company WHERE Id = %s"
        sql_info = my_sql.sql_select(query, (self.le_transport_company.whatsThis(),))
        if "mysql.connector.errors" in str(type(sql_info)):
            QMessageBox.critical(self, "Ошибка sql получения информации транспортной компании", sql_info.msg, QMessageBox.Ok)
            return False
        sheet["B5"] = sql_info[0][0]
        sheet["B16"] = sql_info[0][1]

        query = "SELECT Adres FROM clients_actual_address WHERE Id = %s"
        sql_info = my_sql.sql_select(query, (self.cb_clients_adress.currentData(),))
        if "mysql.connector.errors" in str(type(sql_info)):
            QMessageBox.critical(self, "Ошибка sql получения пункти разгрузки", sql_info.msg, QMessageBox.Ok)
            return False
        sheet["R16"] = sql_info[0][0]

        sheet["B7"] = auto
        sheet["C11"] = driver

        for row in sheet.iter_rows(min_row=23, max_col=31, max_row=25):
            for cell in row:
                cell.border = border_all

        for row in sheet.iter_rows(min_row=38, max_col=31, max_row=43):
            for cell in row:
                cell.border = border_all

        for row in sheet.iter_rows(min_row=45, max_col=23, max_row=50):
            for cell in row:
                cell.border = border_all

        for row in sheet.iter_rows(min_row=52, max_col=23, max_row=57):
            for cell in row:
                cell.border = border_all

            sheet.page_breaks.append(Break(59))

        # Заполняем шапку второго листа
        sheet["AC65"] = self.le_number_doc.text()
        sheet["AC66"] = self.de_date_shipment.date().toString("dd.MM.yyyy")

        query = "SELECT Name, INN, KPP, Actual_Address, Legal_Address, Account, Bank, corres_Account, BIK, Full_Name FROM clients WHERE Id = %s"
        sql_info = my_sql.sql_select(query, (self.le_client.whatsThis(),))
        if "mysql.connector.errors" in str(type(sql_info)):
            QMessageBox.critical(self, "Ошибка sql получения информации о клиенте", sql_info.msg, QMessageBox.Ok)
            return False

        client = sql_info[0][9] + " ИНН " + str(sql_info[0][1])
        if sql_info[0][2]:
            client += " КПП " + str(sql_info[0][2])
        client += ", " + sql_info[0][4] + ",р/с " + str(sql_info[0][5]) + " в " + sql_info[0][6] + " к/с " + str(sql_info[0][7]) + " БИК " + str(sql_info[0][8])
        # sheet["C11"] = client
        sheet["E72"] = client

        if addres == "fact":
            adr = sql_info[0][3]
        elif addres == "adr list":
            query = "SELECT Adres FROM clients_actual_address WHERE Id = %s"
            sql_adr = my_sql.sql_select(query, (self.cb_clients_adress.currentData(),))
            if "mysql.connector.errors" in str(type(sql_adr)):
                QMessageBox.critical(self, "Ошибка sql получения пункти разгрузки", sql_adr.msg, QMessageBox.Ok)
                return False
            adr = sql_adr[0][0]
        else:
            adr = sql_info[0][4]
        sheet["E70"] = adr

        if unite:
            query = """SELECT DISTINCT(product_article_parametrs.Client_code)
                          FROM order_position LEFT JOIN product_article_parametrs ON order_position.Product_Article_Parametr_Id = product_article_parametrs.Id
                          WHERE Order_Id = %s"""
            unite_code_sql = my_sql.sql_select(query, (self.id, ))
            if "mysql.connector.errors" in str(type(unite_code_sql)):
                QMessageBox.critical(self, "Ошибка sql получения уникальных кодов", unite_code_sql.msg, QMessageBox.Ok)
                return False

            product = collections.OrderedDict()

            for cod in unite_code_sql:
                product[str(cod[0])] = {"name": None, "cod": str(cod[0]), "psb": None, "mest": 0, "value": 0, "price": None, "price_no_nds": None,
                                        "sum_no_nds": 0, "nds": None, "sum": 0, "nds_sum": 0}

            for row in range(self.tw_position.rowCount()):
                query = "SELECT Client_code FROM product_article_parametrs WHERE Id = %s"
                sql_info = my_sql.sql_select(query, (self.tw_position.item(row, 2).data(5),))
                if "mysql.connector.errors" in str(type(sql_info)):
                    QMessageBox.critical(self, "Ошибка sql получения информации номера поставщика", sql_info.msg, QMessageBox.Ok)
                    return False

                # Проверяем находиться ли наный товар в списке объединяемых
                cod = sql_info[0][0]
                if cod not in product:
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Не найден код в словаре!", QMessageBox.Ok)
                    return False

                #Проверяем совпадают ли имена
                if product[cod]["name"] is not None and product[cod]["name"] != self.tw_position.item(row, 3).text():
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разные имена одинаковых кодов", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли цены
                if product[cod]["price"] is not None and product[cod]["price"] != float(self.tw_position.item(row, 4).text()):
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разная цена одинаковых кодов", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли psb
                if product[cod]["psb"] is not None and product[cod]["psb"] != int(self.tw_position.item(row, 5).data(5)):
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разный psb одинаковых кодов", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли НДС
                if product[cod]["nds"] is not None and product[cod]["nds"] != int(self.tw_position.item(row, 4).data(5)):
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разный НДС одинаковых кодов", QMessageBox.Ok)
                    return False

                # Вставляем статические параметры
                product[cod]["name"] = self.tw_position.item(row, 3).text()
                product[cod]["price"] = float(self.tw_position.item(row, 4).text())
                product[cod]["psb"] = int(self.tw_position.item(row, 5).data(5))
                product[cod]["nds"] = int(self.tw_position.item(row, 4).data(5))

                # Сумируем кол-во
                product[cod]["value"] += int(self.tw_position.item(row, 5).text())

            # Пройдем по готовым обьединеным позициям и расчитаем цифры
            for cod, cod_value in product.items():
                if self.lb_client.whatsThis().find("no_nds") >= 0:
                    cod_value["price_no_nds"], cod_value["sum_no_nds"], cod_value["nds_sum"], cod_value["sum"], cod_value["mest"] =\
                        calc_price.calc_no_nds(cod_value["price"], cod_value["value"], cod_value["nds"], cod_value["psb"])
                else:
                    cod_value["price_no_nds"], cod_value["sum_no_nds"], cod_value["nds_sum"], cod_value["sum"], cod_value["mest"] =\
                        calc_price.calc_nds(cod_value["price"], cod_value["value"], cod_value["nds"], cod_value["psb"])

        else:
            # если
            product = collections.OrderedDict()
            for row in range(self.tw_position.rowCount()):

                product[row] = {"name": None, "cod": None, "psb": None, "mest": 0, "value": 0, "price": None,
                                                                 "price_no_nds": None, "sum_no_nds": 0, "nds": None, "sum": 0, "nds_sum": 0}

                # Делаем ссылку дя удобства
                position = product[row]

                query = "SELECT Barcode, Client_code FROM product_article_parametrs WHERE Id = %s"
                sql_info = my_sql.sql_select(query, (self.tw_position.item(row, 2).data(5),))
                if "mysql.connector.errors" in str(type(sql_info)):
                    QMessageBox.critical(self, "Ошибка sql получения информации номера поставщика", sql_info.msg, QMessageBox.Ok)
                    return False
                if article:
                    position["name"] = self.tw_position.item(row, 3).text() + " а " + self.tw_position.item(row, 0).text() + \
                                       " р " + self.tw_position.item(row, 1).text() + " " + str(sql_info[0][0])
                    position["cod"] = sql_info[0][1]
                else:
                    position["name"] = self.tw_position.item(row, 3).text() + " " + str(sql_info[0][0])
                    position["cod"] = sql_info[0][1]

                position["nds"] = int(self.tw_position.item(row, 4).data(5))
                position["price"] = float(self.tw_position.item(row, 4).text())
                position["psb"] = int(self.tw_position.item(row, 5).data(5))
                position["value"] += int(self.tw_position.item(row, 5).text())

                if self.lb_client.whatsThis().find("no_nds") >= 0:
                    position["price_no_nds"], position["sum_no_nds"], position["nds_sum"], position["sum"], position["mest"] =\
                        calc_price.calc_no_nds(position["price"], position["value"], position["nds"], position["psb"])

                else:
                    position["price_no_nds"], position["sum_no_nds"], position["nds_sum"], position["sum"], position["mest"] =\
                        calc_price.calc_nds(position["price"], position["value"], position["nds"], position["psb"])

        all_no_nds = 0
        all_mest = 0

        list_all = 1
        row_break = 11
        row_ex = 78
        for cod, position in product.items():

            sheet.merge_cells("B%s:E%s" % (row_ex, row_ex))
            sheet.merge_cells("F%s:I%s" % (row_ex, row_ex))
            sheet.merge_cells("J%s:K%s" % (row_ex, row_ex))
            sheet.merge_cells("L%s:M%s" % (row_ex, row_ex))
            sheet.merge_cells("N%s:S%s" % (row_ex, row_ex))
            sheet.merge_cells("U%s:V%s" % (row_ex, row_ex))
            sheet.merge_cells("W%s:X%s" % (row_ex, row_ex))
            sheet.merge_cells("Y%s:Z%s" % (row_ex, row_ex))
            sheet.merge_cells("AA%s:AC%s" % (row_ex, row_ex))
            sheet.merge_cells("AD%s:AE%s" % (row_ex, row_ex))

            sheet["A%s" % row_ex] = position["cod"]
            sheet["F%s" % row_ex] = position["cod"]
            sheet["J%s" % row_ex] = position["value"]
            sheet["L%s" % row_ex] = moneyfmt.moneyfmt(position["price_no_nds"])
            sheet["L%s" % row_ex].alignment = ald_center_full
            sheet["N%s" % row_ex] = position["name"]
            sheet["T%s" % row_ex] = "шт."
            sheet["U%s" % row_ex] = "Короб"
            sheet["W%s" % row_ex] = position["mest"]
            sheet["Y%s" % row_ex] = "---"
            sheet["AA%s" % row_ex] = moneyfmt.moneyfmt(position["sum_no_nds"])
            sheet["AA%s" % row_ex].alignment = ald_center_full

            all_no_nds += position["sum_no_nds"]
            all_mest += position["mest"]

            sheet.row_dimensions[row_ex].height = 23

            if row_break == 27:
                sheet.page_breaks.append(Break(row_ex))
                list_all += 1
                row_break = 0

            row_break += 1
            row_ex += 1

        if row_break + 9 > 27:
            sheet.page_breaks.append(Break(row_ex-4))
            list_all += 1

        for row in sheet.iter_rows(min_row=75, max_col=31, max_row=row_ex-1):
            for cell in row:
                cell.border = border_all
                cell.alignment = ald_center_full
                cell.font = font_8

        # низ ТТН
        sheet2 = book['Низ']

        # Вставляем имя менеджера
        sheet2['H17'] = manager_name

        for row in sheet2.iter_rows(min_row=1, max_col=31, max_row=20):
            for cell in row:
                sheet["%s%s" % (cell.column, row_ex)] = cell.value
                sheet.row_dimensions[row_ex].height = sheet2.row_dimensions[cell.row].height
                sheet["%s%s" % (cell.column, row_ex)].fill = wite
                if cell.has_style:
                    sheet["%s%s" % (cell.column, row_ex)].border = copy(cell.border)
                    sheet["%s%s" % (cell.column, row_ex)].font = copy(cell.font)
            row_ex += 1
        book.remove(sheet2)

        sheet["M25"] = all_mest
        sheet["H28"] = all_mest

        sheet["U%s" % (row_ex-20)] = self.le_number_order.text()
        sheet["U%s" % (row_ex-20)].font = font_8
        sheet["S%s" % (row_ex-19)] = list_all

        sheet.merge_cells("AA%s:AC%s" % (row_ex-20, row_ex-20))
        sheet["AA%s" % (row_ex-20)] = moneyfmt.moneyfmt(all_no_nds)
        sheet["AA%s" % (row_ex-20)].alignment = ald_center_full
        sheet["AB%s" % (row_ex-20)].border = border_all_big
        sheet["AC%s" % (row_ex-20)].border = border_all_big

        sheet["L%s" % (row_ex-18)] = num2t4ru.num2text(len(product)) + " позиций"
        sheet["L%s" % (row_ex-18)].alignment = ald_center

        sheet["F%s" % (row_ex-16)] = num2t4ru.num2text(len(product))
        sheet["F%s" % (row_ex-16)].alignment = ald_center

        sheet["F%s" % (row_ex-14)] = num2t4ru.num2text(all_mest)
        sheet["F%s" % (row_ex-14)].alignment = ald_center

        int_units = ((u'рубль', u'рубля', u'рублей'), 'm')
        exp_units = ((u'копейка', u'копейки', u'копеек'), 'f')
        sheet["C%s" % (row_ex-10)] = num2t4ru.decimal2text(Decimal(str(all_no_nds)), int_units=int_units, exp_units=exp_units)

        book.save(path[0])

    def of_ex_score(self, article, unite, bank):
        path = QFileDialog.getSaveFileName(self, "Сохранение", filter="Excel(*.xlsx)")
        if not path[0]:
            return False

        book = openpyxl.load_workbook(filename='%s/Счет.xlsx' % (getcwd() + "/templates/order"))
        sheet = book['Отчет']

        font_7 = Font(name="Arial", size=7)
        ald_center = Alignment(horizontal="center")
        border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        border_all_big = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

        if bank == "ВТБ":
            sheet["A2"] = "ИНН/КПП: 7703561330/773001001"
            sheet["A3"] = "Расчетный счет: 40702810417030008128 в Филиал № 7701 Банка ВТБ (ПАО)  г. Москва"
            sheet["A4"] = "Корр. счет: 30101810345250000745 БИК банка: 044525745"
        elif bank == "Бинбанк":
            sheet["A2"] = "ИНН/КПП: 7703561330/773001001"
            sheet["A3"] = "Расчетный счет: 40702810600140100104 в БИНБАНК (ПАО) г. Москва"
            sheet["A4"] = "Корр. счет: 30101810245250000117 БИК банка: 044525117"

        sheet["A5"] = "Счет № %s от %s г." % (self.le_number_doc.text(), self.de_date_order.date().toString("dd.MM.yyyy"))
        sheet["A6"] = "Плательщик: " + self.le_client.text()

        for row in sheet.iter_rows(min_row=5, max_col=7, max_row=5):
            for cell in row:
                cell.border = Border(top=Side(style='medium'), bottom=Side(style='medium'))

        if unite:
            query = """SELECT DISTINCT(product_article_parametrs.Client_code)
                          FROM order_position LEFT JOIN product_article_parametrs ON order_position.Product_Article_Parametr_Id = product_article_parametrs.Id
                          WHERE Order_Id = %s"""
            unite_code_sql = my_sql.sql_select(query, (self.id, ))
            if "mysql.connector.errors" in str(type(unite_code_sql)):
                QMessageBox.critical(self, "Ошибка sql получения уникальных кодов", unite_code_sql.msg, QMessageBox.Ok)
                return False

            product = collections.OrderedDict()

            for cod in unite_code_sql:
                product[str(cod[0])] = {"name": None, "cod": str(cod[0]), "psb": None, "mest": 0, "value": 0, "price": None, "price_no_nds": None,
                                        "sum_no_nds": 0, "nds": None, "sum": 0, "nds_sum": 0}

            for row in range(self.tw_position.rowCount()):
                query = "SELECT Client_code FROM product_article_parametrs WHERE Id = %s"
                sql_info = my_sql.sql_select(query, (self.tw_position.item(row, 2).data(5),))
                if "mysql.connector.errors" in str(type(sql_info)):
                    QMessageBox.critical(self, "Ошибка sql получения информации номера поставщика", sql_info.msg, QMessageBox.Ok)
                    return False

                # Проверяем находиться ли наный товар в списке объединяемых
                cod = sql_info[0][0]
                if cod not in product:
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Не найден код в словаре!", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли имена
                if product[cod]["name"] is not None and product[cod]["name"] != self.tw_position.item(row, 3).text():
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разные имена одинаковых кодов", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли цены
                if product[cod]["price"] is not None and product[cod]["price"] != float(self.tw_position.item(row, 4).text()):
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разная цена одинаковых кодов", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли psb
                if product[cod]["psb"] is not None and product[cod]["psb"] != int(self.tw_position.item(row, 5).data(5)):
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разный psb одинаковых кодов", QMessageBox.Ok)
                    return False

                # Проверяем совпадают ли НДС
                if product[cod]["nds"] is not None and product[cod]["nds"] != int(self.tw_position.item(row, 4).data(5)):
                    QMessageBox.critical(self, "Ошибка объединения кодов", "Разный НДС одинаковых кодов", QMessageBox.Ok)
                    return False

                # Вставляем статические параметры
                product[cod]["name"] = self.tw_position.item(row, 3).text()
                product[cod]["price"] = float(self.tw_position.item(row, 4).text())
                product[cod]["psb"] = int(self.tw_position.item(row, 5).data(5))
                product[cod]["nds"] = int(self.tw_position.item(row, 4).data(5))

                # Сумируем кол-во
                product[cod]["value"] += int(self.tw_position.item(row, 5).text())

            # Пройдем по готовым обьединеным позициям и расчитаем цифры
            for cod, cod_value in product.items():
                if self.lb_client.whatsThis().find("no_nds") >= 0:
                    cod_value["price_no_nds"], cod_value["sum_no_nds"], cod_value["nds_sum"], cod_value["sum"], cod_value["mest"] =\
                        calc_price.calc_no_nds(cod_value["price"], cod_value["value"], cod_value["nds"], cod_value["psb"])
                else:
                    cod_value["price_no_nds"], cod_value["sum_no_nds"], cod_value["nds_sum"], cod_value["sum"], cod_value["mest"] =\
                        calc_price.calc_nds(cod_value["price"], cod_value["value"], cod_value["nds"], cod_value["psb"])

        else:
            # если
            product = collections.OrderedDict()
            for row in range(self.tw_position.rowCount()):

                product[row] = {"name": None, "cod": None, "psb": None, "mest": 0, "value": 0, "price": None,
                                                                 "price_no_nds": None, "sum_no_nds": 0, "nds": None, "sum": 0, "nds_sum": 0}

                # Делаем ссылку дя удобства
                position = product[row]

                query = "SELECT Barcode, Client_code FROM product_article_parametrs WHERE Id = %s"
                sql_info = my_sql.sql_select(query, (self.tw_position.item(row, 2).data(5),))
                if "mysql.connector.errors" in str(type(sql_info)):
                    QMessageBox.critical(self, "Ошибка sql получения информации номера поставщика", sql_info.msg, QMessageBox.Ok)
                    return False
                if article:
                    position["name"] = self.tw_position.item(row, 3).text() + " а " + self.tw_position.item(row, 0).text() + \
                                       " р " + self.tw_position.item(row, 1).text() + " " + str(sql_info[0][0])
                    position["cod"] = sql_info[0][1]
                else:
                    position["name"] = self.tw_position.item(row, 3).text() + " " + str(sql_info[0][0])
                    position["cod"] = sql_info[0][1]

                position["nds"] = int(self.tw_position.item(row, 4).data(5))
                position["price"] = float(self.tw_position.item(row, 4).text())
                position["psb"] = int(self.tw_position.item(row, 5).data(5))
                position["value"] += int(self.tw_position.item(row, 5).text())

                if self.lb_client.whatsThis().find("no_nds") >= 0:
                    position["price_no_nds"], position["sum_no_nds"], position["nds_sum"], position["sum"], position["mest"] =\
                        calc_price.calc_no_nds(position["price"], position["value"], position["nds"], position["psb"])

                else:
                    position["price_no_nds"], position["sum_no_nds"], position["nds_sum"], position["sum"], position["mest"] =\
                        calc_price.calc_nds(position["price"], position["value"], position["nds"], position["psb"])

        all_nds = 0
        all_sum = 0

        row_ex = 9
        num = 1
        for cod, position in product.items():
            sheet.merge_cells("D%s:E%s" % (row_ex, row_ex))

            sheet["A%s" % row_ex] = num
            sheet["A%s" % row_ex].alignment = ald_center
            sheet["B%s" % row_ex] = position["name"]
            sheet["B%s" % row_ex].font = font_7
            sheet["B%s" % row_ex].alignment = Alignment(wrapText=True)
            sheet["C%s" % row_ex] = "шт."
            sheet["D%s" % row_ex] = position["value"]
            sheet["F%s" % row_ex] = position["price"]
            sheet["F%s" % row_ex].number_format = "#,##0.00"
            sheet["G%s" % row_ex] = position["sum"]
            sheet["G%s" % row_ex].number_format = "#,##0.00"

            all_nds += position["nds_sum"]
            all_sum += position["sum"]

            sheet.row_dimensions[row_ex].height = 23

            row_ex += 1
            num += 1

        for row in sheet.iter_rows(min_row=9, max_col=7, max_row=row_ex-1):
            for cell in row:
                cell.border = border_all

        # сумма товаров
        sheet.merge_cells("A%s:F%s" % (row_ex, row_ex))
        sheet["A%s" % row_ex] = "Итого:"
        sheet["G%s" % row_ex] = all_sum
        sheet["G%s" % row_ex].number_format = "#,##0.00"
        row_ex += 1

        sheet.merge_cells("A%s:F%s" % (row_ex, row_ex))
        sheet["A%s" % row_ex] = "В том числе НДС:"
        sheet["G%s" % row_ex] = all_nds
        sheet["G%s" % row_ex].number_format = "#,##0.00"
        for row in sheet.iter_rows(min_row=row_ex-1, max_col=7, max_row=row_ex):
            for cell in row:
                cell.border = border_all_big
        row_ex += 2

        # сумма прописью
        int_units = ((u'рубль', u'рубля', u'рублей'), 'm')
        exp_units = ((u'копейка', u'копейки', u'копеек'), 'f')
        sheet.merge_cells("A%s:F%s" % (row_ex, row_ex))
        sheet["A%s" % row_ex] = "Сумма к оплате: " + num2t4ru.decimal2text(Decimal(str(all_sum)), int_units=int_units, exp_units=exp_units)
        row_ex += 1

        int_units = ((u'рубль', u'рубля', u'рублей'), 'm')
        exp_units = ((u'копейка', u'копейки', u'копеек'), 'f')
        sheet.merge_cells("A%s:F%s" % (row_ex, row_ex))
        sheet["A%s" % row_ex] = "В том числе НДС: " + num2t4ru.decimal2text(Decimal(str(all_nds)), int_units=int_units, exp_units=exp_units)
        row_ex += 2

        # подписи
        sheet["A%s" % row_ex] = "М.П."
        sheet["A%s" % row_ex].alignment = Alignment(horizontal="center")
        sheet.merge_cells("B%s:D%s" % (row_ex, row_ex))
        sheet["B%s" % row_ex] = "Руководитель предприятия"
        sheet["B%s" % row_ex].alignment = Alignment(horizontal="right")
        row_ex += 1
        sheet.merge_cells("B%s:D%s" % (row_ex, row_ex))
        sheet["B%s" % row_ex] = "Главный бухгалтер"
        sheet["B%s" % row_ex].alignment = Alignment(horizontal="right")

        for row in sheet.iter_rows(min_row=row_ex-1, min_col=5, max_col=7, max_row=row_ex):
            for cell in row:
                cell.border = Border(bottom=Side(style='thin'))

        book.save(path[0])

    def of_word_reestr(self):
        path = QFileDialog.getSaveFileName(self, "Сохранение", filter="Word(*.doc)")
        if not path[0]:
            return False

        f = open(getcwd() + '/templates/order/reestr.xml', "r", -1, "utf-8")
        xml = f.read()
        f.close()

        xml = xml.replace("?НАКЛАДНОМЕР", self.le_number_doc.text())
        xml = xml.replace("?НАКЛАДДАТА", self.de_date_shipment.date().toString("dd.MM.yyyy"))

        f = open(path[0], "w", -1, "utf-8")
        f.write(xml)
        f.close()

    def of_word_act(self, price):
        path = QFileDialog.getSaveFileName(self, "Сохранение", filter="Word(*.doc)")
        if not path[0]:
            return False

        f = open(getcwd() + '/templates/order/act.xml', "r", -1, "utf-8")
        xml = f.read()
        f.close()

        query = "SELECT Full_Name, Legal_Address, INN, KPP FROM clients WHERE Id = %s"
        sql_info = my_sql.sql_select(query, (self.le_client.whatsThis(), ))
        if "mysql.connector.errors" in str(type(sql_info)):
            QMessageBox.critical(self, "Ошибка sql получения информации о клиенте", sql_info.msg, QMessageBox.Ok)
            return False

        adres = "%s %s ИНН/КПП %s/%s" % (sql_info[0][0], sql_info[0][1], sql_info[0][2], sql_info[0][3])
        nds_sum = round(price - (price * 100 / 118), 2)
        nds_text = "%s рублей %s копеек" % (int(nds_sum // 1), str(round(nds_sum - int(nds_sum), 2))[2:])

        xml = xml.replace("?НОМЕР", self.le_number_doc.text())
        xml = xml.replace("?ДАТА", self.de_date_shipment.date().toString("dd.MM.yyyy"))
        xml = xml.replace("?АДРЕС", adres)
        xml = xml.replace("?ЦЕНА", str(price))
        xml = xml.replace("?НДС", str(nds_sum))
        xml = xml.replace("?СНДС", nds_text)

        f = open(path[0], "w", -1, "utf-8")
        f.write(xml)
        f.close()


class OrderFilter(QDialog):
    def __init__(self, main):
        super(OrderFilter, self).__init__()
        loadUi(getcwd() + '/ui/order_filter.ui', self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))

        self.main = main

    def ui_view_client(self):
        self.client_list = clients.ClientList(self, True)
        self.client_list.setWindowModality(Qt.ApplicationModal)
        self.client_list.show()

    def ui_del_client(self):
        self.le_client.setWhatsThis("")
        self.le_client.setText("")

    def ui_acc(self):
        where = ""

        # Блок  условий выбора клиента
        if self.le_client.whatsThis() != '':
            where = self.add_filter(where, "(`order`.Client_Id = %s)" % self.le_client.whatsThis())

        # Блок  условий состояния отгрузки заказа
        where_item = ""
        if self.cb_shipped.isChecked():
            where_item = self.add_filter(where_item, "`order`.Shipped = 1", False)

        if self.cb_no_shipped.isChecked():
            where_item = self.add_filter(where_item, "`order`.Shipped = 0", False)

        if where_item:
            where_item = "(" + where_item + ")"
            where = self.add_filter(where, where_item)

        # Блок  условий проверки заказа НДС или без
        where_item = ""
        if self.cb_nds.isChecked():
            where_item = self.add_filter(where_item, "clients.No_Nds = 0", False)

        if self.cb_no_nds.isChecked():
            where_item = self.add_filter(where_item, "clients.No_Nds = 1", False)

        if where_item:
            where_item = "(" + where_item + ")"
            where = self.add_filter(where, where_item)

        # Блок условий номера заказа
        if self.le_order_number.text() != '':
            where = self.add_filter(where, "(`order`.Number_Order LIKE '%s')" % ("%" + self.le_order_number.text() + "%", ))

        # Блок условий номера документа
        if self.le_order_doc.text() != '':
            where = self.add_filter(where, "(`order`.Number_Doc LIKE '%s')" % ("%" + self.le_order_doc.text() + "%", ))

        # Блок  условий суммы заказа
        if self.gp_sum.isChecked():
            sql_date = "(SUM(order_position.Value * order_position.Price) BETWEEN %s AND %s)" % \
                       (self.le_sum_from.text(), self.le_sum_to.text())
            where = self.add_filter(where, sql_date)

        # Блок  условий даты закза
        if self.gp_date_order.isChecked():
            sql_date = "(`order`.Date_Order >= '%s' AND `order`.Date_Order <= '%s')" % \
                       (self.de_date_order_from.date().toString(Qt.ISODate), self.de_date_order_to.date().toString(Qt.ISODate))
            where = self.add_filter(where, sql_date)

        # Блок  условий даты отгрузки
        if self.gp_date_shipped.isChecked():
            sql_date = "(`order`.Date_Shipment >= '%s' AND `order`.Date_Shipment <= '%s')" % \
                       (self.de_date_shipped_from.date().toString(Qt.ISODate), self.de_date_shipped_to.date().toString(Qt.ISODate))
            where = self.add_filter(where, sql_date)

        # Делаем замену так как Were должно быть перед Group by
        if where:
            self.sql_query_all = self.sql_query_all.replace("GROUP BY", " WHERE " + where + " GROUP BY")

        self.main.of_set_filter(self.sql_query_all)

        self.close()

    def ui_can(self):
        self.close()
        self.destroy()

    def add_filter(self, where, add, and_add=True):
        if where:
            if and_add:
                where += " AND " + add
            else:
                where += " OR " + add
        else:
            where = add

        return where

    def of_set_sql_query(self, sql):
        self.sql_query_all = sql

    def of_list_clients(self, item):
        id_client, name_client = item
        self.le_client.setText(str(name_client))
        self.le_client.setWhatsThis(str(id_client))


class Position(QDialog):
    def __init__(self):
        super(Position, self).__init__()
        loadUi(getcwd() + '/ui/order_position.ui', self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))

    def ui_calculation(self):
        if self.nds_1.isChecked():
            nds = 18
        elif self.nds_2.isChecked():
            nds = 10
        else:
            nds = 0

        try:
            value = int(self.le_value.text().replace(",", "."))
        except:
            value = 0
        try:
            if self.price_name_change == "nds":
                price = round(float(self.le_price.text().replace(",", ".")), 4)
                price_no_nds = round(price - (price * nds) / (100 + nds), 4)
            elif self.price_name_change == "no nds":
                price_no_nds = round(float(self.le_price_no_nds.text().replace(",", ".")), 4)
                price = round(price_no_nds * (0.01 * nds + 1), 4)
            else:
                price = 0.0
                price_no_nds = 0.0
        except:
            price = 0.0
            price_no_nds = 0.0
        try:
            in_on_place = int(self.le_in_on_place.text().replace(",", "."))
        except:
            in_on_place = 0

        if value:
            if self.price_name_change == "nds":
                self.le_price_no_nds.setText(str(round(price_no_nds, 4)))
            elif self.price_name_change == "no nds":
                self.le_price.setText(str(round(price, 4)))
            sum_rub = round(value * price, 4)
            self.le_sum_rub.setText(str(sum_rub))
            if nds:
                self.le_sum_no_nds.setText(str(round(price_no_nds * value, 4)))
                self.le_sum_nds.setText(str(round((price * nds) / (100 + nds) * value, 4)))

        if in_on_place:
            self.le_place_all.setText(str(round(value / in_on_place, 1)))

    def ui_price_nds_change(self):
        self.price_name_change = "nds"
        self.ui_calculation()

    def ui_price_no_nds_change(self):
        self.price_name_change = "no nds"
        self.ui_calculation()

    def ui_view_article(self):
        self.article_list = article.ArticleList(self, True)
        self.article_list.setWindowModality(Qt.ApplicationModal)
        self.article_list.show()

    def ui_acc(self):
        self.done(1)
        self.close()
        self.destroy()

    def ui_cancel(self):
        self.done(0)
        self.close()
        self.destroy()

    def of_tree_select_article(self, article):
        self.article_list.close()
        self.article_list.destroy()
        self.le_article.setText(article["article"])
        self.le_size.setText(article["size"])
        self.le_parametr.setText(article["parametr"])
        self.le_parametr.setWhatsThis(str(article["parametr_id"]))
        self.le_price.setText(article["price"])
        self.le_in_on_place.setText(article["in on place"])
        self.le_client_cod.setText(article["client_cod"])
        if article["nds"] == 18:
            self.nds_1.setChecked(True)
        else:
            self.nds_2.setChecked(True)

        self.price_name_change = "nds"
        self.ui_calculation()


class TransportCompanyName(list.ListItems):
    def set_settings(self):
        self.setWindowTitle("Список транспорт компаний")  # Имя окна
        self.toolBar.setStyleSheet("background-color: rgb(211, 49, 60);")  # Цвет бара
        self.title_new_window = "ТК"  # Имя вызываемых окон
        self.resize(300, 400)

        self.sql_list = "SELECT Id, Name FROM order_transport_company"
        self.sql_add = "INSERT INTO order_transport_company (Name, Details) VALUES (%s, %s)"
        self.sql_change_select = "SELECT Name, Details FROM order_transport_company WHERE Id = %s"
        self.sql_update_select = 'UPDATE order_transport_company SET Name = %s, Details = %s WHERE Id = %s'
        self.sql_dell = "DELETE FROM order_transport_company WHERE Id = %s"

        self.set_new_win = {"WinTitle": "ТК",
                            "WinColor": "(211, 49, 60)",
                            "lb_name": "Название",
                            "lb_note": "Подробность"}


class OrderDocList(QDialog):
    def __init__(self, main):
        super(OrderDocList, self).__init__()
        loadUi(getcwd() + '/ui/order_doc_list.ui', self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))

        self.main = main

    def ui_select_doc(self, item):
        doc_name = item.text()

        if doc_name == "Накладная":
            self.sw_main.setCurrentIndex(1)
        elif doc_name == "Счет фактура":
            self.sw_main.setCurrentIndex(2)
        elif doc_name == "ТТН":
            self.sw_main.setCurrentIndex(3)
        elif doc_name == "Счет":
            self.sw_main.setCurrentIndex(4)
        elif doc_name == "Реестр":
            self.sw_main.setCurrentIndex(0)
        elif doc_name == "Акт доставки":
            self.sw_main.setCurrentIndex(5)

    def ui_acc(self):
        if self.lw_main.selectedItems()[0].text() == "Накладная":
            if self.cb_edo.isChecked():
                edo = True
            else:
                edo = False

            if self.cb_head.isChecked():
                head = True
            else:
                head = False

            if self.cb_unite_1.isChecked():
                unite = True
            else:
                unite = False

            if self.rb_addres_1.isChecked():
                addres = True
            else:
                addres = False

            if self.rb_name_1.isChecked():
                article = False
            else:
                article = True

            manager_name = self.cb_manager_name_1.currentText()
            no_pcb = self.cb_no_pcb_1.isChecked()

            self.main.of_ex_torg12(edo, head, article, addres, unite, manager_name, no_pcb)
            self.close()
            self.destroy()

        elif self.lw_main.selectedItems()[0].text() == "Счет фактура":
            if self.cb_unite_2.isChecked():
                unite = True
            else:
                unite = False

            if self.rb_fact_addres_1.isChecked():
                addres = "fact"
            elif self.rb_fact_addres_2.isChecked():
                addres = "adr list"
            else:
                addres = "legal"

            if self.rb_fact_name_1.isChecked():
                article = False
            else:
                article = True

            self.main.of_ex_invoice(article, addres, unite)
            self.close()
            self.destroy()

        elif self.lw_main.selectedItems()[0].text() == "ТТН":

            if self.cb_unite_3.isChecked():
                unite = True
            else:
                unite = False

            if self.rb_ttn_addres_1.isChecked():
                addres = "fact"
            elif self.rb_ttn_addres_2.isChecked():
                addres = "adr list"
            else:
                addres = "legal"

            if self.rb_ttn_name_1.isChecked():
                article = False
            else:
                article = True

            manager_name = self.cb_manager_name_3.currentText()
            self.main.of_ex_ttn(addres, article, self.le_ttn_auto.text(), self.le_ttn_driver.text(), unite, manager_name)
            self.close()
            self.destroy()

        elif self.lw_main.selectedItems()[0].text() == "Счет":
            if self.rb_score_name_1.isChecked():
                article = False
            else:
                article = True

            if self.cb_unite_4.isChecked():
                unite = True
            else:
                unite = False

            bank = self.rb_score_bank.currentText()

            self.main.of_ex_score(article, unite, bank)
            self.close()
            self.destroy()

        elif self.lw_main.selectedItems()[0].text() == "Реестр":
            self.main.of_word_reestr()
            self.close()
            self.destroy()

        elif self.lw_main.selectedItems()[0].text() == "Акт доставки":
            try:
                price = float(self.le_act_price.text())
            except:
                return False
            self.main.of_word_act(price)
            self.close()
            self.destroy()

    def ui_can(self):
        self.close()
        self.destroy()


class ImportEDI(QDialog):
    def __init__(self, main):
        super(ImportEDI, self).__init__()
        loadUi(getcwd() + '/ui/order_import_edi.ui', self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))

        self.login = {'Name': '4607191149998EC', 'Password': '3NxK8bAyG'}
        self.main = main
        self.set_size_table()
        self.start_edi()

    def set_size_table(self):
        self.tw_edi_1.horizontalHeader().resizeSection(0, 150)
        self.tw_edi_1.horizontalHeader().resizeSection(1, 130)
        self.tw_edi_1.horizontalHeader().resizeSection(2, 410)

        self.tw_edi_2.horizontalHeader().resizeSection(0, 90)
        self.tw_edi_2.horizontalHeader().resizeSection(1, 95)
        self.tw_edi_2.horizontalHeader().resizeSection(2, 130)

        self.tw_edi_3.horizontalHeader().resizeSection(0, 240)
        self.tw_edi_3.horizontalHeader().resizeSection(1, 90)
        self.tw_edi_3.horizontalHeader().resizeSection(2, 70)
        self.tw_edi_3.horizontalHeader().resizeSection(3, 45)
        self.tw_edi_3.horizontalHeader().resizeSection(4, 45)
        self.tw_edi_3.horizontalHeader().resizeSection(5, 40)
        self.tw_edi_3.horizontalHeader().resizeSection(6, 60)
        self.tw_edi_3.horizontalHeader().resizeSection(7, 60)
        self.tw_edi_3.horizontalHeader().resizeSection(8, 200)
        self.tw_edi_3.horizontalHeader().resizeSection(9, 40)

    def start_edi(self):
        url = "https://www.ecod.pl/webserv2/EDIservice.asmx/Relationships"
        values = self.login
        values.update( {"Timeout": 2000} )
        req = requests.get(url, params=values)

        text = req.text.replace("&lt;", "<")
        text = text.replace("&gt;", ">")

        tree = xml_pars.fromstring(text)

        if tree[0].text != "00000000":
            QMessageBox.critical(self, "Ошибка EDI", "Ошибка EDI %s" % tree[0].text, QMessageBox.Ok)
            self.close()
            self.destroy()
            return False

        self.tw_edi_1.setRowCount(0)
        for relation in tree[1][0]:
            if relation[3].text == "IN" and relation[4].text == "ORDER":
                info = {"PartnerIln": relation[1].text,
                        "DocumentType": relation[4].text,
                        "DocumentVersion": relation[5].text,
                        "DocumentStandard": relation[6].text,
                        "DocumentTest": relation[7].text
                        }

                self.tw_edi_1.insertRow(self.tw_edi_1.rowCount())

                item = QTableWidgetItem(str(relation[2].text))
                item.setData(5, info)
                self.tw_edi_1.setItem(self.tw_edi_1.rowCount()-1, 0, item)

                item = QTableWidgetItem(str(relation[4].text))
                self.tw_edi_1.setItem(self.tw_edi_1.rowCount()-1, 1, item)

                item = QTableWidgetItem(str(relation[8].text))
                self.tw_edi_1.setItem(self.tw_edi_1.rowCount()-1, 2, item)

    def ui_edi1_acc(self):
        try:
            item_info = self.tw_edi_1.selectedItems()[0].data(5)
        except:
            QMessageBox.critical(self, "Ошибка ", "Выделите клиента", QMessageBox.Ok)
            return False

        url = "https://www.ecod.pl/webserv2/EDIservice.asmx/ListMBEx"
        item_info.update(self.login)
        item_info.update({'DocumentStatus': 'A', 'Timeout': 3000, "DateFrom": "", "DateTo": "", "ItemFrom": "", "ItemTo": 60})
        req = requests.get(url, params=item_info)

        text = req.text.replace("&lt;", "<")
        text = text.replace("&gt;", ">")

        tree = xml_pars.fromstring(text)

        if tree[0].text != "00000000":
            QMessageBox.critical(self, "Ошибка EDI", "Ошибка EDI %s" % tree[0].text, QMessageBox.Ok)
            self.close()
            self.destroy()
            return False

        self.tw_edi_2.setRowCount(0)
        for relation in tree[1][0]:
            info = {"PartnerIln": relation[0].text,
                    "DocumentType": relation[2].text,
                    "DocumentStandard": relation[4].text,
                    "TrackingId": relation[1].text
                    }

            self.tw_edi_2.insertRow(self.tw_edi_2.rowCount())

            item = QTableWidgetItem(str(relation[7].text))
            item.setData(5, info)
            self.tw_edi_2.setItem(self.tw_edi_2.rowCount()-1, 0, item)

            item = QTableWidgetItem(QDate.fromString(relation[8].text, Qt.ISODate).toString("dd.MM.yyyy"))
            self.tw_edi_2.setItem(self.tw_edi_2.rowCount()-1, 1, item)

            item = QTableWidgetItem(QDateTime.fromString(relation[10].text, "yyyy-MM-dd HH:mm:ss").toString("dd.MM.yyyy HH:mm:ss"))
            self.tw_edi_2.setItem(self.tw_edi_2.rowCount()-1, 2, item)

        self.sw_main.setCurrentIndex(1)

    def ui_edi2_to_edi1(self):
        self.start_edi()
        self.sw_main.setCurrentIndex(0)

    def ui_edi2_acc(self):
        try:
            item_info = self.tw_edi_2.selectedItems()[0].data(5)
        except:
            QMessageBox.critical(self, "Ошибка ", "Выделите заказ", QMessageBox.Ok)
            return False

        url = "https://www.ecod.pl/webserv2/EDIservice.asmx/Receive"
        item_info.update(self.login)
        item_info.update({'ChangeDocumentStatus': 'R', 'Timeout': 5000})
        req = requests.get(url, params=item_info)

        text = req.text.replace("&lt;", "<")
        text = text.replace("&gt;", ">")

        tree = xml_pars.fromstring(text)

        if tree[0].text != "00000000":
            QMessageBox.critical(self, "Ошибка EDI", "Ошибка EDI %s" % tree[0].text, QMessageBox.Ok)
            self.close()
            self.destroy()
            return False

        self.tw_edi_3.setRowCount(0)
        color = QBrush(QColor(252, 141, 141, 255))

        for element in tree.iter('{http://www.comarch.com/}OrderNumber'):
            self.lb_order_number.setText(element.text)

        for element in tree.iter('{http://www.comarch.com/}OrderDate'):
            self.le_order_date.setText(element.text)

        for element in tree.iter('{http://www.comarch.com/}TotalLines'):
            self.le_order_position.setText(element.text)

        for element in tree.iter('{http://www.comarch.com/}TotalOrderedAmount'):
            self.le_order_value.setText(element.text)

        if self.cb_division_position.isChecked():  # Если позиции надо разделять
            for element in tree.iter('{http://www.comarch.com/}Line-Item'):  # Начинаем перебирать позиции

                # Находим все артикула с одинаковывм кодом клиента
                query = """SELECT product_article_parametrs.Id, product_article.Article, product_article_size.Size, product_article_parametrs.Name,
                                product_article_parametrs.Client_Name, product_article_parametrs.Price, product_article_parametrs.NDS, product_article_parametrs.In_On_Place,
                                product_article_parametrs.Barcode
                              FROM product_article_parametrs
                                LEFT JOIN product_article_size ON product_article_parametrs.Product_Article_Size_Id = product_article_size.Id
                                LEFT JOIN product_article ON product_article_size.Article_Id = product_article.Id
                              WHERE product_article_parametrs.Client_code = %s"""
                sql_info = my_sql.sql_select(query, (element.find('{http://www.comarch.com/}BuyerItemCode').text.lstrip("0"),))
                if "mysql.connector.errors" in str(type(sql_info)):
                    QMessageBox.critical(self, "Ошибка sql получения артикула", sql_info.msg, QMessageBox.Ok)
                    return False

                # разделем заказаное кол-во на найденое кол-во строк
                # Найдем кол-во в зависимости от того есть ли оно в документе ЕДИ
                if self.cb_no_all_value.isChecked():
                    value = int(float(element.find('{http://www.comarch.com/}OrderedUnitPacksize').text) * float(element.find('{http://www.comarch.com/}OrderedQuantity').text))
                else:
                    value = int(float(element.find('{http://www.comarch.com/}OrderedQuantity').text))
                value_one_position = value / len(sql_info)

                # вставляем поочередно все найденые совпадения для кода клиента
                for sql_position in sql_info:
                    self.tw_edi_3.insertRow(self.tw_edi_3.rowCount())

                    item = QTableWidgetItem(sql_position[4])
                    self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 0, item)

                    item = QTableWidgetItem(sql_position[8])
                    self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 1, item)

                    item = QTableWidgetItem(element.find('{http://www.comarch.com/}BuyerItemCode').text)
                    self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 2, item)

                    item = QTableWidgetItem(str(value_one_position))
                    self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 3, item)

                    OrderedUnitPacksize = element.find('{http://www.comarch.com/}OrderedUnitPacksize').text
                    item = QTableWidgetItem(OrderedUnitPacksize)
                    self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 5, item)

                    item = QTableWidgetItem(str(int(value_one_position / float(OrderedUnitPacksize))))
                    self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 4, item)

                    item = QTableWidgetItem(element.find('{http://www.comarch.com/}OrderedUnitNetPrice').text)
                    self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 6, item)

                    price = round(sql_position[5] - (sql_position[5] * sql_position[6]) / (100 + sql_position[6]), 2)
                    item_product = QTableWidgetItem(sql_position[1] + " " + sql_position[2] + " " + sql_position[3])
                    item_product.setData(5, {"article": sql_position[1],
                                             "size": sql_position[2],
                                             "parametr": sql_position[3],
                                             "parametr_id": sql_position[0],
                                             "client_Name": sql_position[4],
                                             "price": sql_position[5]})

                    if str(price) == element.find('{http://www.comarch.com/}OrderedUnitNetPrice').text:
                        item_price = QTableWidgetItem(str(price))
                        item_price.setData(5, sql_position[6])
                    else:
                        item_price = QTableWidgetItem(str(price))
                        item_price.setData(5, sql_position[6])
                        item_price.setBackground(color)

                    self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 8, item_product)
                    self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 7, item_price)

                    butt = QPushButton("Изм.")
                    butt.setProperty("row", self.tw_edi_3.rowCount()-1)
                    butt.clicked.connect(self.change_material_name)
                    self.tw_edi_3.setCellWidget(self.tw_edi_3.rowCount()-1, 9, butt)

        else:  # Если позиции не надо разделять
            for element in tree.iter('{http://www.comarch.com/}Line-Item'):

                query = """SELECT product_article_parametrs.Id, product_article.Article, product_article_size.Size, product_article_parametrs.Name,
                                product_article_parametrs.Client_Name, product_article_parametrs.Price, product_article_parametrs.NDS, product_article_parametrs.In_On_Place
                              FROM product_article_parametrs
                                LEFT JOIN product_article_size ON product_article_parametrs.Product_Article_Size_Id = product_article_size.Id
                                LEFT JOIN product_article ON product_article_size.Article_Id = product_article.Id
                              WHERE product_article_parametrs.Client_code = %s"""
                sql_info = my_sql.sql_select(query, (element.find('{http://www.comarch.com/}BuyerItemCode').text.lstrip("0"),))
                if "mysql.connector.errors" in str(type(sql_info)):
                    QMessageBox.critical(self, "Ошибка sql получения артикула", sql_info.msg, QMessageBox.Ok)
                    return False

                self.tw_edi_3.insertRow(self.tw_edi_3.rowCount())

                # Найдем кол-во в зависимости от того есть ли оно в документе ЕДИ
                if self.cb_no_all_value.isChecked():
                    value = int(float(element.find('{http://www.comarch.com/}OrderedUnitPacksize').text) * float(element.find('{http://www.comarch.com/}OrderedQuantity').text))
                else:
                    value = int(float(element.find('{http://www.comarch.com/}OrderedQuantity').text))

                item = QTableWidgetItem(element.find('{http://www.comarch.com/}ItemDescription').text)
                self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 0, item)

                item = QTableWidgetItem(element.find('{http://www.comarch.com/}EAN').text)
                self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 1, item)

                item = QTableWidgetItem(element.find('{http://www.comarch.com/}BuyerItemCode').text)
                self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 2, item)

                item = QTableWidgetItem(str(value))
                self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 3, item)

                OrderedUnitPacksize = element.find('{http://www.comarch.com/}OrderedUnitPacksize').text
                item = QTableWidgetItem(OrderedUnitPacksize)
                self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 5, item)

                item = QTableWidgetItem(str(value / int(float(OrderedUnitPacksize))))
                self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 4, item)

                item = QTableWidgetItem(element.find('{http://www.comarch.com/}OrderedUnitNetPrice').text)
                self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 6, item)

                if sql_info:
                    if len(sql_info) == 1:  # Одно совпадение

                        price = round(sql_info[0][5] - (sql_info[0][5] * sql_info[0][6]) / (100 + sql_info[0][6]), 2)

                        item_product = QTableWidgetItem(sql_info[0][1] + " " + sql_info[0][2] + " " + sql_info[0][3])
                        item_product.setData(5, {"article": sql_info[0][1],
                                                 "size": sql_info[0][2],
                                                 "parametr": sql_info[0][3],
                                                 "parametr_id": sql_info[0][0],
                                                 "client_Name": sql_info[0][4],
                                                 "price": sql_info[0][5]})

                        if str(price) == element.find('{http://www.comarch.com/}OrderedUnitNetPrice').text:
                            item_price = QTableWidgetItem(str(price))
                            item_price.setData(5, sql_info[0][6])
                        else:
                            item_price = QTableWidgetItem(str(price))
                            item_price.setData(5, sql_info[0][6])
                            item_price.setBackground(color)
                    else:  # Несколько совпадений
                        item_product = QTableWidgetItem("Несколько совпадений")
                        item_price = QTableWidgetItem("")
                else:  # Нет совпадений
                    item_product = QTableWidgetItem("Нет совпадений")
                    item_price = QTableWidgetItem("")

                self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 8, item_product)
                self.tw_edi_3.setItem(self.tw_edi_3.rowCount()-1, 7, item_price)

                butt = QPushButton("Изм.")
                butt.setProperty("row", self.tw_edi_3.rowCount()-1)
                butt.clicked.connect(self.change_material_name)
                self.tw_edi_3.setCellWidget(self.tw_edi_3.rowCount()-1, 9, butt)

        self.sw_main.setCurrentIndex(2)

    def ui_edi3_to_edi2(self):
        self.sw_main.setCurrentIndex(1)

    def ui_edi3_acc(self):
        # Проверим что бы количество было целым числом
        for row in range(self.tw_edi_3.rowCount()):
            if self.tw_edi_3.item(row, 8).data(5):
                value = float(self.tw_edi_3.item(row, 3).text())
                if value % 1 != 0:
                    QMessageBox.critical(self, "Ошибка количества", "Где то не целое колличество товара", QMessageBox.Ok)
                    return False

        for row in range(self.tw_edi_3.rowCount()):
            if self.tw_edi_3.item(row, 8).data(5):
                row_main = self.main.tw_position.rowCount()

                article = self.tw_edi_3.item(row, 8).data(5)

                self.main.tw_position.insertRow(row_main)

                table_item = QTableWidgetItem(article["article"])
                table_item.setData(-1, "new")
                self.main.tw_position.setItem(row_main, 0, table_item)

                table_item = QTableWidgetItem(article["size"])
                table_item.setData(-1, "new")
                self.main.tw_position.setItem(row_main, 1, table_item)

                table_item = QTableWidgetItem(article["parametr"])
                table_item.setData(-1, "new")
                table_item.setData(5, article["parametr_id"])
                self.main.tw_position.setItem(row_main, 2, table_item)

                table_item = QTableWidgetItem(article["client_Name"])
                table_item.setData(-1, "new")
                self.main.tw_position.setItem(row_main, 3, table_item)

                nds = self.tw_edi_3.item(row, 7).data(5)

                if self.main.lb_client.whatsThis().find("no_nds") >= 0:
                    price_no_nds = round(float(article["price"]) - (float(article["price"]) * int(nds)) / (100 + int(nds)), 4)
                    table_item = QTableWidgetItem(str(price_no_nds))
                    table_item.setData(-1, "new")
                    table_item.setData(5, nds)
                    self.main.tw_position.setItem(row, 4, table_item)

                    table_item = QTableWidgetItem(str(round(price_no_nds * int(float(self.tw_edi_3.item(row, 3).text())), 4)))
                    table_item.setData(-1, "new")
                    table_item.setData(5, nds)
                    self.main.tw_position.setItem(row, 6, table_item)
                else:
                    table_item = QTableWidgetItem(str(article["price"]))
                    table_item.setData(-1, "new")
                    table_item.setData(5, nds)
                    self.main.tw_position.setItem(row, 4, table_item)

                    table_item = QTableWidgetItem(str(round(float(article["price"]) * int(float(self.tw_edi_3.item(row, 3).text())), 4)))
                    table_item.setData(-1, "new")
                    table_item.setData(5, nds)
                    self.main.tw_position.setItem(row, 6, table_item)

                table_item = QTableWidgetItem(self.tw_edi_3.item(row, 3).text())
                table_item.setData(-1, "new")
                table_item.setData(5, self.tw_edi_3.item(row, 5).text())
                self.main.tw_position.setItem(row, 5, table_item)

        self.main.save_change_order_position = True
        self.main.calc_sum()
        try:
            self.main.pb_doc.deleteLater()
        except:
            pass
        self.close()
        self.destroy()
        return True

    def change_material_name(self):
        butt = QObject.sender(self)
        self.row_change_material = butt.property("row")
        self.article_list = article.ArticleList(self, True)
        self.article_list.setWindowModality(Qt.ApplicationModal)
        self.article_list.show()

    def of_tree_select_article(self, article):
        self.article_list.close()
        self.article_list.destroy()
        item = QTableWidgetItem(article["article"] + " " + article["size"] + " " + article["parametr"])
        item.setData(5, article)
        self.tw_edi_3.setItem(self.row_change_material, 8, item)

        price = round(float(article["price"]) - (float(article["price"]) * int(article["nds"])) / (100 + int(article["nds"])), 2)
        item = QTableWidgetItem(str(price))
        item.setData(5, article["nds"])
        if str(price) == self.tw_edi_3.item(self.row_change_material, 6).text():
            self.tw_edi_3.setItem(self.row_change_material, 7, item)
        else:
            color = QBrush(QColor(252, 141, 141, 255))
            item.setBackground(color)
            self.tw_edi_3.setItem(self.row_change_material, 7, item)



